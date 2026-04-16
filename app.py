"""
Supply Chain & Operations Helper v5.0
SR — Supply Chain Intelligence Platform
All improvements: Custom Logo, Transport Logic, Green-Red Heatmap,
Report Generation, Free AI (Groq), Enhanced UI
"""

# ═══════════════════════════════════════════════════════════════
# IMPORTS
# ═══════════════════════════════════════════════════════════════
import streamlit as st
import streamlit.components.v1 as components
import pandas as pd
import numpy as np
import networkx as nx
import plotly.graph_objects as go
import plotly.express as px
from plotly.subplots import make_subplots
from dataclasses import dataclass
from datetime import datetime, timedelta, date
from sklearn.ensemble import RandomForestRegressor, GradientBoostingRegressor
from sklearn.metrics import mean_squared_error, mean_absolute_error
import json, math, io, re, requests
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.utils import get_column_letter
import warnings
warnings.filterwarnings("ignore")

# ═══════════════════════════════════════════════════════════════
# CUSTOM SVG LOGO (SR + Supply Chain Network)
# ═══════════════════════════════════════════════════════════════
LOGO_SVG = """<svg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 120 120' width='60' height='60'>
  <defs>
    <linearGradient id='bg' x1='0%' y1='0%' x2='100%' y2='100%'>
      <stop offset='0%' style='stop-color:#1B4F72'/>
      <stop offset='100%' style='stop-color:#0F6E56'/>
    </linearGradient>
  </defs>
  <circle cx='60' cy='60' r='58' fill='url(#bg)' stroke='#FFFFFF' stroke-width='2.5'/>
  <circle cx='30' cy='38' r='8' fill='#F39C12' stroke='white' stroke-width='1.5'/>
  <circle cx='60' cy='22' r='8' fill='#27AE60' stroke='white' stroke-width='1.5'/>
  <circle cx='90' cy='38' r='8' fill='#E74C3C' stroke='white' stroke-width='1.5'/>
  <circle cx='38' cy='68' r='8' fill='#9B59B6' stroke='white' stroke-width='1.5'/>
  <circle cx='82' cy='68' r='8' fill='#2980B9' stroke='white' stroke-width='1.5'/>
  <circle cx='60' cy='88' r='8' fill='#F39C12' stroke='white' stroke-width='1.5'/>
  <line x1='30' y1='38' x2='60' y2='22' stroke='#F8C471' stroke-width='2.5'/>
  <line x1='60' y1='22' x2='90' y2='38' stroke='#7DCEA0' stroke-width='2.5'/>
  <line x1='30' y1='38' x2='38' y2='68' stroke='#F8C471' stroke-width='2.5'/>
  <line x1='90' y1='38' x2='82' y2='68' stroke='#7DCEA0' stroke-width='2.5'/>
  <line x1='38' y1='68' x2='60' y2='88' stroke='#AED6F1' stroke-width='2.5'/>
  <line x1='82' y1='68' x2='60' y2='88' stroke='#AED6F1' stroke-width='2.5'/>
  <line x1='38' y1='68' x2='82' y2='68' stroke='white' stroke-width='1.5' opacity='0.5' stroke-dasharray='3,2'/>
  <text x='60' y='110' text-anchor='middle' font-family='Arial,sans-serif' font-size='13' font-weight='900' fill='white' letter-spacing='3'>SR</text>
</svg>"""

LOGO_SIDEBAR = """<svg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 120 120' width='42' height='42'>
  <defs><linearGradient id='bg2' x1='0%' y1='0%' x2='100%' y2='100%'><stop offset='0%' style='stop-color:#1B4F72'/><stop offset='100%' style='stop-color:#0F6E56'/></linearGradient></defs>
  <circle cx='60' cy='60' r='58' fill='url(#bg2)' stroke='#FFFFFF' stroke-width='2'/>
  <circle cx='30' cy='38' r='7' fill='#F39C12' stroke='white' stroke-width='1.5'/>
  <circle cx='60' cy='22' r='7' fill='#27AE60' stroke='white' stroke-width='1.5'/>
  <circle cx='90' cy='38' r='7' fill='#E74C3C' stroke='white' stroke-width='1.5'/>
  <circle cx='38' cy='68' r='7' fill='#9B59B6' stroke='white' stroke-width='1.5'/>
  <circle cx='82' cy='68' r='7' fill='#2980B9' stroke='white' stroke-width='1.5'/>
  <circle cx='60' cy='88' r='7' fill='#F39C12' stroke='white' stroke-width='1.5'/>
  <line x1='30' y1='38' x2='60' y2='22' stroke='#F8C471' stroke-width='2'/>
  <line x1='60' y1='22' x2='90' y2='38' stroke='#7DCEA0' stroke-width='2'/>
  <line x1='30' y1='38' x2='38' y2='68' stroke='#F8C471' stroke-width='2'/>
  <line x1='90' y1='38' x2='82' y2='68' stroke='#7DCEA0' stroke-width='2'/>
  <line x1='38' y1='68' x2='60' y2='88' stroke='#AED6F1' stroke-width='2'/>
  <line x1='82' y1='68' x2='60' y2='88' stroke='#AED6F1' stroke-width='2'/>
  <text x='60' y='110' text-anchor='middle' font-family='Arial,sans-serif' font-size='12' font-weight='900' fill='white' letter-spacing='2'>SR</text>
</svg>"""

# ═══════════════════════════════════════════════════════════════
# DATA MODELS
# ═══════════════════════════════════════════════════════════════
@dataclass
class Node:
    id: str; name: str; node_type: str; capacity: float
    location: str = ""; x: float = 0.0; y: float = 0.0

@dataclass
class Edge:
    source: str; target: str; capacity: float
    cost: float = 1.0; active: bool = True

# ═══════════════════════════════════════════════════════════════
# TRANSPORT LOGIC
# ═══════════════════════════════════════════════════════════════
TRANSPORT_MODES = {
    "road":  {"speed_kmh": 60,  "cost_per_km": 2.5,  "label": "Road (Truck)",  "icon": "Truck"},
    "rail":  {"speed_kmh": 80,  "cost_per_km": 1.2,  "label": "Rail",          "icon": "Rail"},
    "air":   {"speed_kmh": 800, "cost_per_km": 12.0,  "label": "Air",           "icon": "Air"},
    "sea":   {"speed_kmh": 35,  "cost_per_km": 0.8,  "label": "Sea/Inland",    "icon": "Sea"},
}

def haversine_km(lat1, lon1, lat2, lon2):
    """Distance between two lat/lon points in km"""
    R = 6371
    dlat = math.radians(lat2 - lat1)
    dlon = math.radians(lon2 - lon1)
    a = math.sin(dlat/2)**2 + math.cos(math.radians(lat1))*math.cos(math.radians(lat2))*math.sin(dlon/2)**2
    return R * 2 * math.asin(math.sqrt(a))

def get_state(location_str):
    """Extract state from 'City, State' format"""
    parts = location_str.split(",")
    return parts[-1].strip().lower() if len(parts) > 1 else location_str.strip().lower()

def recommend_transport(src_node, tgt_node, load_units, urgent=False):
    """
    Smart transport mode recommendation based on:
    - Same city/state → road
    - Load size → truck / rail / air
    - Urgent → air
    - Distance → affects mode
    """
    src_loc = src_node.location; tgt_loc = tgt_node.location
    src_state = get_state(src_loc); tgt_state = get_state(tgt_loc)
    same_state = (src_state == tgt_state)

    # Compute distance if coordinates available
    dist_km = 0
    if src_node.x and src_node.y and tgt_node.x and tgt_node.y:
        dist_km = haversine_km(src_node.y, src_node.x, tgt_node.y, tgt_node.x)

    recommendations = []

    if urgent:
        mode = "air"
        reason = "Urgent delivery — air freight selected"
    elif same_state or dist_km < 150:
        mode = "road"
        reason = "Same region — road transport (cheapest)"
    elif load_units >= 500:
        mode = "rail"
        reason = f"High load ({load_units} units) — rail preferred"
    elif load_units >= 150:
        # Offer both
        mode = "rail"
        reason = f"Medium load ({load_units} units) — rail recommended"
    else:
        mode = "road"
        reason = f"Small load ({load_units} units) — road (truck)"

    m = TRANSPORT_MODES[mode]
    if dist_km > 0:
        cost = dist_km * m["cost_per_km"]
        hours = dist_km / m["speed_kmh"]
        days = hours / 24
    else:
        cost = None; hours = None; days = None

    return {
        "mode": mode, "label": m["label"], "icon": m["icon"],
        "reason": reason, "dist_km": round(dist_km, 1) if dist_km else None,
        "cost_estimate": round(cost, 0) if cost else None,
        "hours": round(hours, 1) if hours else None,
        "days": round(days, 2) if days else None,
        "same_state": same_state,
    }

# ═══════════════════════════════════════════════════════════════
# GRAPH ENGINE
# ═══════════════════════════════════════════════════════════════
class SupplyChainGraph:
    def __init__(self):
        self.nodes: dict = {}
        self.edges: list = []

    def add_node(self, node): self.nodes[node.id] = node

    def add_edge(self, edge):
        if edge.source not in self.nodes or edge.target not in self.nodes:
            raise ValueError(f"Node {edge.source} or {edge.target} not found")
        for e in self.edges:
            if e.source == edge.source and e.target == edge.target: return
        self.edges.append(edge)

    def remove_edge(self, src, tgt):
        self.edges = [e for e in self.edges if not (e.source==src and e.target==tgt)]

    def toggle_edge(self, src, tgt, active):
        for e in self.edges:
            if e.source==src and e.target==tgt: e.active = active

    def to_nx(self):
        G = nx.DiGraph()
        for nid, n in self.nodes.items(): G.add_node(nid, **vars(n))
        for e in self.edges:
            if e.active:
                G.add_edge(e.source, e.target, capacity=e.capacity, weight=e.cost)
        return G

    def geo_shortest_path(self, src, tgt):
        """Shortest path using real lat/lon distances as weights"""
        G = nx.DiGraph()
        for nid, n in self.nodes.items(): G.add_node(nid, **vars(n))
        for e in self.edges:
            if not e.active: continue
            sn = self.nodes[e.source]; tn = self.nodes[e.target]
            if sn.x and sn.y and tn.x and tn.y:
                dist = haversine_km(sn.y, sn.x, tn.y, tn.x)
                w = dist if dist > 0 else e.cost
            else:
                w = e.cost
            G.add_edge(e.source, e.target, capacity=e.capacity, weight=w, cost=e.cost)
        try:
            path = nx.dijkstra_path(G, src, tgt, weight="weight")
            total_dist = sum(
                haversine_km(self.nodes[path[i]].y, self.nodes[path[i]].x,
                             self.nodes[path[i+1]].y, self.nodes[path[i+1]].x)
                if (self.nodes[path[i]].x and self.nodes[path[i+1]].x) else 0
                for i in range(len(path)-1)
            )
            total_cost = sum(G[path[i]][path[i+1]].get("cost", 1) for i in range(len(path)-1))
            return {"found": True, "path": path, "dist_km": round(total_dist, 1), "cost": round(total_cost, 2)}
        except:
            return {"found": False, "path": [], "dist_km": None, "cost": None}

    def shortest_path(self, src, tgt):
        return self.geo_shortest_path(src, tgt)

    def all_shortest_paths(self, src, tgt, k=3):
        G = self.to_nx(); results = []
        try:
            for path in nx.shortest_simple_paths(G, src, tgt, weight="weight"):
                cost = sum(G[path[i]][path[i+1]]["weight"] for i in range(len(path)-1))
                results.append({"path": path, "cost": round(cost, 2)})
                if len(results) >= k: break
        except: pass
        return results

    def demand_fulfillment(self):
        G = self.to_nx()
        plants  = [n for n, d in self.nodes.items() if d.node_type=="plant"]
        demands = [n for n, d in self.nodes.items() if d.node_type=="demand"]
        results = {}
        for d_id in demands:
            req = self.nodes[d_id].capacity; total, reach = 0.0, []
            for p in plants:
                try:
                    fv, _ = nx.maximum_flow(G, p, d_id, capacity="capacity")
                    if fv > 0: total += fv; reach.append(p)
                except: pass
            ful = min(total, req)
            results[d_id] = {"required": req, "fulfilled": round(ful, 2),
                             "pct": round(ful/req*100 if req>0 else 100, 1),
                             "reachable_from": reach}
        return results

    def simulate_disruption(self, src, tgt):
        baseline = self.demand_fulfillment()
        self.toggle_edge(src, tgt, False)
        disrupted = self.demand_fulfillment()
        self.toggle_edge(src, tgt, True)
        impact = {}
        for d_id, base in baseline.items():
            dis = disrupted[d_id]; drop = base["pct"] - dis["pct"]
            impact[d_id] = {
                "demand_name": self.nodes[d_id].name,
                "before_pct": base["pct"], "after_pct": dis["pct"],
                "drop_pct": round(drop,1), "before_fulfilled": base["fulfilled"],
                "after_fulfilled": dis["fulfilled"],
                "lost_units": round(base["fulfilled"]-dis["fulfilled"],2),
                "severity": "critical" if drop>=50 else "high" if drop>=25 else "medium" if drop>0 else "none"
            }
        alt_paths = {}
        for d_id, imp in impact.items():
            if imp["drop_pct"] > 0:
                self.toggle_edge(src, tgt, False)
                paths = []
                for p in [n for n, nd in self.nodes.items() if nd.node_type=="plant"]:
                    r = self.shortest_path(p, d_id)
                    if r["found"]:
                        paths.append({"from_plant": self.nodes[p].name,
                                      "path": [self.nodes[n].name for n in r["path"]],
                                      "path_ids": r["path"], "cost": r["cost"]})
                paths.sort(key=lambda x: x["cost"])
                alt_paths[d_id] = paths[:3]
                self.toggle_edge(src, tgt, True)
        avg_drop = sum(v["drop_pct"] for v in impact.values())/len(impact) if impact else 0
        return {"removed_edge": f"{self.nodes[src].name} → {self.nodes[tgt].name}",
                "removed_src": src, "removed_tgt": tgt,
                "resilience_score": round(max(0, 100-avg_drop), 1),
                "impact": impact, "alt_paths": alt_paths}

    def rank_critical_edges(self):
        ranking = []
        for e in self.edges:
            if not e.active: continue
            r = self.simulate_disruption(e.source, e.target)
            avg = sum(v["drop_pct"] for v in r["impact"].values())/len(r["impact"]) if r["impact"] else 0
            ranking.append({"source":e.source,"target":e.target,"label":r["removed_edge"],
                "avg_fulfillment_drop":round(avg,1),"resilience_score":r["resilience_score"],
                "severity":"critical" if avg>=50 else "high" if avg>=25 else "medium" if avg>=5 else "low"})
        return sorted(ranking, key=lambda x: x["avg_fulfillment_drop"], reverse=True)

    def to_dict(self):
        return {"nodes":[vars(n) for n in self.nodes.values()],"edges":[vars(e) for e in self.edges]}


# ═══════════════════════════════════════════════════════════════
# INVENTORY MANAGER
# ═══════════════════════════════════════════════════════════════
class InventoryManager:
    def __init__(self):
        self.items = {}; self.stock = {}

    def add_item(self, iid, name, unit="units"):
        self.items[iid] = {"name": name, "unit": unit}

    def set_stock(self, node_id, iid, current, safety, reorder, daily_demand=1.0):
        if node_id not in self.stock: self.stock[node_id] = {}
        self.stock[node_id][iid] = {"current":float(current),"safety":float(safety),
                                     "reorder":float(reorder),"daily_demand":float(daily_demand)}

    def update_stock(self, node_id, iid, delta):
        if node_id in self.stock and iid in self.stock[node_id]:
            self.stock[node_id][iid]["current"] = max(0.0, self.stock[node_id][iid]["current"]+delta)
            return True
        return False

    def coverage_days(self, node_id, iid):
        if node_id in self.stock and iid in self.stock[node_id]:
            s = self.stock[node_id][iid]; dd = s.get("daily_demand",1)
            return round(s["current"]/dd,1) if dd>0 else 999
        return None

    def get_alerts(self, sc_nodes=None):
        alerts = []
        for node_id, items in self.stock.items():
            nn = sc_nodes[node_id].name if sc_nodes and node_id in sc_nodes else node_id
            for iid, s in items.items():
                iname = self.items.get(iid,{}).get("name",iid)
                if s["current"] <= s["reorder"]:
                    alerts.append({"node_id":node_id,"node_name":nn,"item_id":iid,"item_name":iname,
                        "level":"critical" if s["current"]<=s["safety"] else "warning",
                        "current":s["current"],"safety":s["safety"],"reorder":s["reorder"],
                        "coverage":self.coverage_days(node_id,iid)})
        return sorted(alerts, key=lambda x:(x["level"]!="critical", x["coverage"] or 999))

    def find_alternatives(self, demand_id, iid, required, sc, disrupted=None):
        alts = []
        for node_id, items in self.stock.items():
            if node_id==demand_id or iid not in items: continue
            s=items[iid]; avail=max(0.0, s["current"]-s["safety"])
            if avail<=0: continue
            if disrupted: sc.toggle_edge(disrupted[0], disrupted[1], False)
            r = sc.shortest_path(node_id, demand_id)
            if disrupted: sc.toggle_edge(disrupted[0], disrupted[1], True)
            if r["found"]:
                alts.append({"node_id":node_id,
                    "node_name":sc.nodes[node_id].name if node_id in sc.nodes else node_id,
                    "available":avail,"can_cover":avail>=required,
                    "coverage_pct":min(100, round(avail/max(required,1)*100,1)),
                    "path":[sc.nodes[n].name for n in r["path"] if n in sc.nodes],
                    "path_ids":r["path"],"route_cost":r["cost"],
                    "coverage_days":self.coverage_days(node_id,iid)})
        return sorted(alts, key=lambda x:(-x["coverage_pct"], x["route_cost"]))[:5]

    def to_df(self, sc_nodes=None):
        rows = []
        for node_id, items in self.stock.items():
            nn = sc_nodes[node_id].name if sc_nodes and node_id in sc_nodes else node_id
            nt = sc_nodes[node_id].node_type if sc_nodes and node_id in sc_nodes else ""
            for iid, s in items.items():
                iname=self.items.get(iid,{}).get("name",iid); unit=self.items.get(iid,{}).get("unit","units")
                dd=s.get("daily_demand",1); cov=round(s["current"]/dd,1) if dd>0 else 999
                status="Critical" if s["current"]<=s["safety"] else "Low" if s["current"]<=s["reorder"] else "Normal"
                rows.append({"Node":nn,"Type":nt.capitalize(),"Item":iname,"Unit":unit,
                    "Current Stock":s["current"],"Safety Stock":s["safety"],"Reorder Point":s["reorder"],
                    "Daily Demand":dd,"Coverage Days":cov,
                    "Available (above safety)":max(0,s["current"]-s["safety"]),"Status":status})
        return pd.DataFrame(rows)


# ═══════════════════════════════════════════════════════════════
# DEMAND FORECASTER
# ═══════════════════════════════════════════════════════════════
class DemandForecaster:
    def __init__(self):
        self.models={}; self.history={}; self.forecasts={}; self.metrics={}

    def generate_synthetic_history(self, node_id, node_name, base_demand, days=365, seed=None):
        if seed is None: seed=hash(node_id)%1000
        np.random.seed(seed); t=np.arange(days)
        demand=np.maximum(0, base_demand+0.05*t+base_demand*0.15*np.sin(2*np.pi*t/7)+
                          base_demand*0.20*np.sin(2*np.pi*t/365)+np.random.normal(0,base_demand*0.08,days))
        dates=[datetime.today()-timedelta(days=days-i) for i in range(days)]
        df=pd.DataFrame({"date":dates,"demand":demand,"node_id":node_id,"node_name":node_name})
        self.history[node_id]=df; return df

    def _make_features(self, y, lags=[1,7,14,30]):
        ml=max(lags); X,yo=[],[]
        for i in range(ml, len(y)):
            feats=[i,i%7,i%30,i//30]
            for lag in lags: feats.append(float(y[i-lag]))
            feats.append(float(np.mean(y[max(0,i-7):i]))); feats.append(float(np.mean(y[max(0,i-30):i])))
            X.append(feats); yo.append(float(y[i]))
        return np.array(X,dtype=float), np.array(yo,dtype=float)

    def train(self, node_id, horizon=30):
        if node_id not in self.history: return None
        df=self.history[node_id]; y=df["demand"].values
        X,yt=self._make_features(y); split=max(int(len(X)*0.80),1)
        rf=RandomForestRegressor(n_estimators=100,max_depth=6,random_state=42,n_jobs=-1)
        gbm=GradientBoostingRegressor(n_estimators=80,max_depth=4,learning_rate=0.1,random_state=42)
        rf.fit(X[:split],yt[:split]); gbm.fit(X[:split],yt[:split])
        if split<len(X):
            ep=0.5*rf.predict(X[split:])+0.5*gbm.predict(X[split:])
            yv=yt[split:]
            rmse=float(np.sqrt(mean_squared_error(yv,ep))); mae=float(mean_absolute_error(yv,ep))
            mape=float(np.mean(np.abs((yv-ep)/(yv+1e-6)))*100)
            r2=float(1-np.sum((yv-ep)**2)/np.sum((yv-np.mean(yv))**2))
        else:
            rmse=0;mae=0;mape=0;r2=1
        self.models[node_id]=(rf,gbm)
        self.metrics[node_id]={"rmse":round(rmse,2),"mae":round(mae,2),"mape":round(mape,2),"r2":round(r2,3)}
        hv=list(y); fp=[]
        for step in range(horizon):
            i=len(hv); lags=[1,7,14,30]
            feats=[i,i%7,i%30,i//30]
            for lag in lags: feats.append(float(hv[-lag]) if lag<=len(hv) else float(np.mean(hv[-min(lag,len(hv)):])))
            feats.append(float(np.mean(hv[-7:]))); feats.append(float(np.mean(hv[-30:])))
            arr=np.array(feats,dtype=float).reshape(1,-1)
            pred=max(0, 0.5*rf.predict(arr)[0]+0.5*gbm.predict(arr)[0])
            fp.append(pred); hv.append(pred)
        ld=df["date"].iloc[-1]; fd=[ld+timedelta(days=i+1) for i in range(horizon)]
        fdf=pd.DataFrame({"date":fd,"forecast":fp,"node_id":node_id,"node_name":df["node_name"].iloc[0]})
        fdf["upper"]=fdf["forecast"]+1.5*rmse; fdf["lower"]=np.maximum(0,fdf["forecast"]-1.5*rmse)
        self.forecasts[node_id]=fdf; return fdf

    def aggregate_to_warehouses(self, sc, horizon=30):
        wh_fc={}
        for wid, wh in sc.nodes.items():
            if wh.node_type!="warehouse": continue
            G=sc.to_nx(); served=[]
            for did, d in sc.nodes.items():
                if d.node_type=="demand":
                    try: nx.shortest_path(G,wid,did); served.append(did)
                    except: pass
            total=np.zeros(horizon)
            for did in served:
                if did in self.forecasts:
                    vals=self.forecasts[did]["forecast"].values[:horizon]; total[:len(vals)]+=vals
            wh_fc[wid]={"name":wh.name,"forecast":total.tolist(),"served_demands":served}
        return wh_fc

    def get_plant_requirements(self, sc, wh_fc, horizon=30):
        pr={}
        for pid, p in sc.nodes.items():
            if p.node_type!="plant": continue
            G=sc.to_nx(); served=[]
            for wid in wh_fc:
                try: nx.shortest_path(G,pid,wid); served.append(wid)
                except: pass
            total=np.zeros(horizon)
            for wid in served: total+=np.array(wh_fc[wid]["forecast"][:horizon])
            pr[pid]={"name":p.name,"required":total.tolist(),"capacity":p.capacity,"served_wh":served}
        return pr


# ═══════════════════════════════════════════════════════════════
# REPORT GENERATOR
# ═══════════════════════════════════════════════════════════════
def generate_excel_report(sc, inv, forecaster, ranking, start_date, end_date):
    """Comprehensive Excel report with color-coded data and embedded charts"""
    wb = openpyxl.Workbook()

    # Styles
    fill_green  = PatternFill("solid", fgColor="DCFCE7")
    fill_yellow = PatternFill("solid", fgColor="FEF3C7")
    fill_red    = PatternFill("solid", fgColor="FEE2E2")
    fill_blue   = PatternFill("solid", fgColor="DBEAFE")
    fill_hdr    = PatternFill("solid", fgColor="1E3A5F")
    fill_sub    = PatternFill("solid", fgColor="EFF6FF")
    fill_alt    = PatternFill("solid", fgColor="F8FAFC")

    fnt_hdr  = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
    fnt_title= Font(bold=True, size=15, color="1E3A5F", name="Calibri")
    fnt_sub  = Font(bold=True, size=12, color="334155", name="Calibri")
    fnt_bold = Font(bold=True, size=10, name="Calibri")
    fnt_norm = Font(size=10, name="Calibri")
    fnt_sm   = Font(size=9, color="64748B", name="Calibri")
    aln_c    = Alignment(horizontal="center", vertical="center", wrap_text=True)
    aln_l    = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    thin     = Side(style="thin", color="E2E8F0")
    bdr      = Border(left=thin, right=thin, top=thin, bottom=thin)

    def hdr_row(ws, row, cols):
        for c in range(1, cols+1):
            cell=ws.cell(row=row,column=c)
            cell.fill=fill_hdr; cell.font=fnt_hdr; cell.alignment=aln_c; cell.border=bdr

    def data_row(ws, row, cols, fill=None, alt=False):
        f = fill if fill else (fill_alt if alt else None)
        for c in range(1, cols+1):
            cell=ws.cell(row=row,column=c)
            if f: cell.fill=f
            cell.font=fnt_norm; cell.border=bdr; cell.alignment=aln_l

    # ═══ SHEET 1 — EXECUTIVE SUMMARY ═══════════════
    ws = wb.active; ws.title="Executive Summary"
    ws.sheet_view.showGridLines=False
    for col,w in zip("ABCDEFG",[28,16,16,16,16,16,14]):
        ws.column_dimensions[col].width=w

    ws.merge_cells("A1:G1")
    ws["A1"]="SUPPLY CHAIN & OPERATIONS — EXECUTIVE REPORT"
    ws["A1"].font=Font(bold=True,size=16,color="1E3A5F",name="Calibri")
    ws["A1"].alignment=aln_c
    ws["A1"].fill=PatternFill("solid",fgColor="EFF6FF")
    ws.row_dimensions[1].height=36

    ws.merge_cells("A2:G2")
    ws["A2"]=f"Period: {start_date} to {end_date}  |  Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}  |  SR Supply Chain & Operations Helper"
    ws["A2"].font=fnt_sm; ws["A2"].alignment=aln_c

    plants=[n for n in sc.nodes.values() if n.node_type=="plant"]
    whs_  =[n for n in sc.nodes.values() if n.node_type=="warehouse"]
    dems_ =[n for n in sc.nodes.values() if n.node_type=="demand"]
    total_cap=sum(n.capacity for n in plants); total_dem=sum(n.capacity for n in dems_)
    cov=round(total_cap/max(total_dem,1)*100,1)

    # KPI row
    kpis=[("Plants",len(plants),"1E3A5F"),("Warehouses",len(whs_),"1E3A5F"),
          ("Demand Points",len(dems_),"1E3A5F"),("Connections",len(sc.edges),"1E3A5F"),
          ("Plant Capacity",f"{total_cap:,.0f}","14532D"),
          ("Total Demand",f"{total_dem:,.0f}","1E3A5F"),
          ("Coverage",f"{cov}%","14532D" if cov>=100 else "91241C")]
    for c,(lbl,val,col) in enumerate(kpis,start=1):
        ws.cell(row=4,column=c,value=lbl).font=Font(size=9,bold=True,color="94A3B8",name="Calibri")
        ws.cell(row=4,column=c).alignment=aln_c
        ws.cell(row=5,column=c,value=val).font=Font(size=13,bold=True,color=col,name="Calibri")
        ws.cell(row=5,column=c).alignment=aln_c
        ws.cell(row=5,column=c).fill=fill_sub
    ws.row_dimensions[5].height=28

    # Demand fulfillment table
    ws.cell(row=7,column=1,value="DEMAND FULFILLMENT OVERVIEW").font=fnt_sub
    ff=sc.demand_fulfillment()
    hff=["Demand Point","Required","Fulfilled","Fulfillment %","Status"]
    for c,h in enumerate(hff,1): ws.cell(row=8,column=c,value=h)
    hdr_row(ws,8,len(hff))
    r=9
    for did,info in ff.items():
        pct=info["pct"]; status="On Target" if pct>=90 else "At Risk" if pct>=60 else "Shortage"
        fill=fill_green if pct>=90 else fill_yellow if pct>=60 else fill_red
        for c,v in enumerate([sc.nodes[did].name,info["required"],info["fulfilled"],f"{pct}%",status],1):
            ws.cell(row=r,column=c,value=v)
        data_row(ws,r,len(hff),fill); r+=1

    # Chart: Demand fulfillment
    chart1=BarChart(); chart1.type="col"
    chart1.title="Demand Fulfillment (%)"; chart1.style=10
    chart1.y_axis.title="Fulfillment (%)"; chart1.y_axis.scaling.max=110
    d_ref=Reference(ws,min_col=4,min_row=8,max_row=r-1)
    c_ref=Reference(ws,min_col=1,min_row=9,max_row=r-1)
    chart1.add_data(d_ref,titles_from_data=True); chart1.set_categories(c_ref)
    chart1.height=12; chart1.width=20; ws.add_chart(chart1,f"A{r+1}")

    # ═══ SHEET 2 — INVENTORY ══════════════════════
    ws2=wb.create_sheet("Inventory Status")
    ws2.sheet_view.showGridLines=False
    for col,w in zip("ABCDEFGHIJK",[22,13,16,9,13,13,13,11,13,16,11]):
        ws2.column_dimensions[col].width=w

    ws2.merge_cells("A1:K1"); ws2["A1"]="INVENTORY STATUS REPORT"
    ws2["A1"].font=fnt_title; ws2["A1"].alignment=aln_c
    ws2["A1"].fill=PatternFill("solid",fgColor="EFF6FF"); ws2.row_dimensions[1].height=30

    inv_df=inv.to_df(sc.nodes)
    hdr2=["Node","Type","Item","Unit","Current","Safety","Reorder","Daily Demand","Coverage Days","Available","Status"]
    for c,h in enumerate(hdr2,1): ws2.cell(row=3,column=c,value=h)
    hdr_row(ws2,3,len(hdr2))
    for alt,(idx,rd) in enumerate(inv_df.iterrows()):
        r2=alt+4; st2=rd["Status"]
        fill=fill_green if st2=="Normal" else fill_blue if st2=="Low" else fill_red
        vals=[rd["Node"],rd["Type"],rd["Item"],rd["Unit"],rd["Current Stock"],rd["Safety Stock"],
              rd["Reorder Point"],rd["Daily Demand"],rd["Coverage Days"],rd["Available (above safety)"],st2]
        for c,v in enumerate(vals,1): ws2.cell(row=r2,column=c,value=v)
        data_row(ws2,r2,len(hdr2),fill,alt%2==0)

    n_inv=len(inv_df)
    # Chart: Stock levels
    chart2=BarChart(); chart2.type="col"
    chart2.title="Current vs Safety Stock by Item"; chart2.style=10
    chart2.y_axis.title="Units"
    cur_r=Reference(ws2,min_col=5,min_row=3,max_row=3+n_inv)
    saf_r=Reference(ws2,min_col=6,min_row=3,max_row=3+n_inv)
    cat_r=Reference(ws2,min_col=3,min_row=4,max_row=3+n_inv)
    chart2.add_data(cur_r,titles_from_data=True); chart2.add_data(saf_r,titles_from_data=True)
    chart2.set_categories(cat_r); chart2.height=13; chart2.width=24
    ws2.add_chart(chart2,f"A{n_inv+6}")

    # Chart: Coverage days
    chart3=LineChart(); chart3.title="Coverage Days per Item"; chart3.style=10
    chart3.y_axis.title="Days"
    cov_r=Reference(ws2,min_col=9,min_row=3,max_row=3+n_inv)
    chart3.add_data(cov_r,titles_from_data=True); chart3.set_categories(cat_r)
    chart3.height=12; chart3.width=22; ws2.add_chart(chart3,f"A{n_inv+26}")

    # ═══ SHEET 3 — RISK ANALYSIS ══════════════════
    ws3=wb.create_sheet("Risk Analysis")
    ws3.sheet_view.showGridLines=False
    for col,w in zip("ABCDE",[30,20,18,14,30]):
        ws3.column_dimensions[col].width=w

    ws3.merge_cells("A1:E1"); ws3["A1"]="RISK ANALYSIS — CRITICAL CONNECTIONS"
    ws3["A1"].font=fnt_title; ws3["A1"].alignment=aln_c
    ws3["A1"].fill=PatternFill("solid",fgColor="FEF2F2"); ws3.row_dimensions[1].height=30

    rmap={"critical":"URGENT: Add redundant path immediately","high":"HIGH: Plan backup within 30 days",
          "medium":"MEDIUM: Monitor quarterly","low":"LOW: Well-redundant, monitor"}
    hdr3=["Connection","Avg Drop","Resilience","Severity","Recommendation"]
    for c,h in enumerate(hdr3,1): ws3.cell(row=3,column=c,value=h)
    hdr_row(ws3,3,len(hdr3))
    for alt2,rec in enumerate(ranking or [],start=4):
        sev=rec["severity"]
        fill=fill_red if sev=="critical" else fill_yellow if sev=="high" else fill_blue if sev=="medium" else fill_green
        for c,v in enumerate([rec["label"],f"{rec['avg_fulfillment_drop']}%",f"{rec['resilience_score']}%",sev.upper(),rmap.get(sev,"")],1):
            ws3.cell(row=alt2,column=c,value=v)
        data_row(ws3,alt2,len(hdr3),fill)

    if ranking:
        nr=len(ranking)
        chart4=BarChart(); chart4.type="bar"; chart4.title="Connection Criticality (Avg Drop %)"; chart4.style=10
        chart4.x_axis.title="Drop (%)"
        dr=Reference(ws3,min_col=2,min_row=3,max_row=3+nr)
        cr=Reference(ws3,min_col=1,min_row=4,max_row=3+nr)
        chart4.add_data(dr,titles_from_data=True); chart4.set_categories(cr)
        chart4.height=max(12,nr*1.8); chart4.width=22; ws3.add_chart(chart4,f"A{nr+6}")

    # ═══ SHEET 4 — FORECAST ═══════════════════════
    ws4=wb.create_sheet("Demand Forecast")
    ws4.sheet_view.showGridLines=False
    for col,w in zip("ABCDEF",[18,13,13,13,13,14]):
        ws4.column_dimensions[col].width=w

    ws4.merge_cells("A1:F1"); ws4["A1"]="DEMAND FORECASTING REPORT"
    ws4["A1"].font=fnt_title; ws4["A1"].alignment=aln_c
    ws4["A1"].fill=PatternFill("solid",fgColor="F0FDF4"); ws4.row_dimensions[1].height=30

    row4=3
    for node_id,fdf in (forecaster.forecasts or {}).items():
        ws4.merge_cells(f"A{row4}:F{row4}")
        ws4.cell(row=row4,column=1,value=f"Forecast: {fdf['node_name'].iloc[0]}").font=fnt_sub
        ws4.row_dimensions[row4].height=22; row4+=1
        if node_id in forecaster.metrics:
            m=forecaster.metrics[node_id]
            ws4.merge_cells(f"A{row4}:F{row4}")
            ws4.cell(row=row4,column=1,value=f"RMSE: {m['rmse']}  MAE: {m['mae']}  MAPE: {m['mape']}%  R2: {m['r2']}").font=fnt_sm
            row4+=1
        hdr4=["Date","Forecast","Lower CI","Upper CI"]
        for c,h in enumerate(hdr4,1): ws4.cell(row=row4,column=c,value=h)
        hdr_row(ws4,row4,4); ds=row4+1; row4+=1
        for alt3,(idx3,fr) in enumerate(fdf.iterrows()):
            for c,v in enumerate([str(fr["date"])[:10],round(fr["forecast"],1),round(fr["lower"],1),round(fr["upper"],1)],1):
                ws4.cell(row=row4,column=c,value=v)
            data_row(ws4,row4,4,fill_blue if alt3%2==0 else None); row4+=1
        nf=len(fdf); chart5=LineChart(); chart5.title=f"Forecast: {fdf['node_name'].iloc[0]}"; chart5.style=10
        chart5.y_axis.title="Units"
        fr_ref=Reference(ws4,min_col=2,min_row=ds-1,max_row=ds+nf-1)
        lo_ref=Reference(ws4,min_col=3,min_row=ds-1,max_row=ds+nf-1)
        hi_ref=Reference(ws4,min_col=4,min_row=ds-1,max_row=ds+nf-1)
        chart5.add_data(fr_ref,titles_from_data=True); chart5.add_data(lo_ref,titles_from_data=True)
        chart5.add_data(hi_ref,titles_from_data=True)
        chart5.height=12; chart5.width=22; ws4.add_chart(chart5,f"A{row4+1}"); row4+=20

    # ═══ SHEET 5 — SCORECARD ══════════════════════
    ws5=wb.create_sheet("Supplier Scorecard")
    ws5.sheet_view.showGridLines=False
    for col,w in zip("ABCDEFG",[22,13,13,13,15,14,9]):
        ws5.column_dimensions[col].width=w
    ws5.merge_cells("A1:G1"); ws5["A1"]="SUPPLIER & NODE PERFORMANCE SCORECARD"
    ws5["A1"].font=fnt_title; ws5["A1"].alignment=aln_c
    ws5["A1"].fill=PatternFill("solid",fgColor="FFFBEB"); ws5.row_dimensions[1].height=30
    hdr5=["Node","Type","Reliability","Lead Time","Quality","Cost Efficiency","Grade"]
    for c,h in enumerate(hdr5,1): ws5.cell(row=3,column=c,value=h)
    hdr_row(ws5,3,len(hdr5))
    scores_ss=st.session_state.get("scores",{})
    row5=4
    for nid,node in sc.nodes.items():
        if node.node_type not in ("plant","warehouse") or nid not in scores_ss: continue
        s=scores_ss[nid]; ov=round((s["reliability"]+s["lead_time"]+s["quality"]+s["cost_efficiency"])/4,1)
        grade="A" if ov>=90 else "B" if ov>=75 else "C" if ov>=60 else "D"
        fill=fill_green if grade=="A" else fill_blue if grade=="B" else fill_yellow if grade=="C" else fill_red
        for c,v in enumerate([node.name,node.node_type.capitalize(),s["reliability"],s["lead_time"],s["quality"],s["cost_efficiency"],grade],1):
            ws5.cell(row=row5,column=c,value=v)
        data_row(ws5,row5,len(hdr5),fill); row5+=1
    if row5>4:
        ns=row5-4; chart6=BarChart(); chart6.type="col"; chart6.title="Supplier Performance"; chart6.style=10
        chart6.y_axis.title="Score"; chart6.y_axis.scaling.max=100
        cats5=Reference(ws5,min_col=1,min_row=4,max_row=row5-1)
        for ci in range(3,7):
            r6=Reference(ws5,min_col=ci,min_row=3,max_row=row5-1)
            chart6.add_data(r6,titles_from_data=True)
        chart6.set_categories(cats5); chart6.height=13; chart6.width=24
        ws5.add_chart(chart6,f"A{row5+2}")

    # ═══ SHEET 6 — ANALYTICS (CVaR + SPC) ═════════
    ws6=wb.create_sheet("Analytics")
    ws6.sheet_view.showGridLines=False
    for col in "ABCDEFGH": ws6.column_dimensions[col].width=16
    ws6.column_dimensions["A"].width=22
    ws6.merge_cells("A1:H1"); ws6["A1"]="SUPPLY CHAIN ANALYTICS — RISK & CONTROL"
    ws6["A1"].font=fnt_title; ws6["A1"].alignment=aln_c
    ws6["A1"].fill=PatternFill("solid",fgColor="EFF6FF"); ws6.row_dimensions[1].height=30

    ws6.cell(row=3,column=1,value="CVaR (Conditional Value at Risk) Analysis").font=fnt_sub
    hdr6a=["Demand Point","Mean Demand","Std Dev","VaR 95%","CVaR 95%","Risk Level"]
    for c,h in enumerate(hdr6a,1): ws6.cell(row=4,column=c,value=h)
    hdr_row(ws6,4,len(hdr6a))
    ff2 = get_fulfillment_cached(sc); r6=5
    for did,info in ff2.items():
        node=sc.nodes[did]; base=node.capacity; std_d=base*0.15
        var95=round(base+1.645*std_d,1); cvar95=round(base+2.0*std_d,1)
        sf=max(0,cvar95-info["fulfilled"]); rl="High" if sf>base*0.3 else "Medium" if sf>base*0.1 else "Low"
        fill=fill_red if rl=="High" else fill_yellow if rl=="Medium" else fill_green
        for c,v in enumerate([node.name,round(base,1),round(std_d,1),var95,cvar95,rl],1):
            ws6.cell(row=r6,column=c,value=v)
        data_row(ws6,r6,len(hdr6a),fill); r6+=1

    chart7=BarChart(); chart7.type="col"; chart7.title="CVaR Risk by Demand Point"; chart7.style=10
    chart7.y_axis.title="Units"
    cv_ref=Reference(ws6,min_col=5,min_row=4,max_row=r6-1)
    cc_ref=Reference(ws6,min_col=1,min_row=5,max_row=r6-1)
    chart7.add_data(cv_ref,titles_from_data=True); chart7.set_categories(cc_ref)
    chart7.height=12; chart7.width=20; ws6.add_chart(chart7,f"A{r6+2}")

    # SPC Control Chart
    r6c=r6+18
    ws6.cell(row=r6c,column=1,value="SPC Control Chart — Demand Monitoring").font=fnt_sub; r6c+=1
    hdr6b=["Period","Demand","UCL","Mean","LCL","In Control"]
    for c,h in enumerate(hdr6b,1): ws6.cell(row=r6c,column=c,value=h)
    hdr_row(ws6,r6c,len(hdr6b)); ds6=r6c+1; r6c+=1
    first_dem=next((n for n in sc.nodes.values() if n.node_type=="demand"),None)
    if first_dem:
        np.random.seed(42)
        dvals=first_dem.capacity+np.random.normal(0,first_dem.capacity*0.08,20)
        md=np.mean(dvals); sd=np.std(dvals); ucl=md+3*sd; lcl=max(0,md-3*sd)
        for i,d in enumerate(dvals,1):
            ic="Yes" if lcl<=d<=ucl else "No"
            fill=fill_green if ic=="Yes" else fill_red
            for c,v in enumerate([f"P{i}",round(d,1),round(ucl,1),round(md,1),round(lcl,1),ic],1):
                ws6.cell(row=r6c,column=c,value=v)
            data_row(ws6,r6c,len(hdr6b),fill); r6c+=1
        chart8=LineChart(); chart8.title=f"SPC Control Chart — {first_dem.name}"; chart8.style=10
        chart8.y_axis.title="Demand Units"
        for ci in range(2,6):
            cr8=Reference(ws6,min_col=ci,min_row=ds6-1,max_row=r6c-1)
            chart8.add_data(cr8,titles_from_data=True)
        chart8.height=14; chart8.width=28; ws6.add_chart(chart8,f"A{r6c+2}")

    buf=io.BytesIO(); wb.save(buf); buf.seek(0); return buf.getvalue()


def load_demo_data():
    sc=SupplyChainGraph()
    sc.add_node(Node("P1","Plant Mumbai",  "plant",    500,"Mumbai, Maharashtra",  72.88,19.08))
    sc.add_node(Node("P2","Plant Chennai", "plant",    400,"Chennai, Tamil Nadu",  80.27,13.08))
    sc.add_node(Node("P3","Plant Pune",    "plant",    350,"Pune, Maharashtra",    73.86,18.52))
    sc.add_node(Node("W1","WH North",  "warehouse",400,"Delhi, Delhi",          77.21,28.61))
    sc.add_node(Node("W2","WH West",   "warehouse",350,"Ahmedabad, Gujarat",    72.57,23.02))
    sc.add_node(Node("W3","WH East",   "warehouse",300,"Kolkata, West Bengal",  88.36,22.57))
    sc.add_node(Node("W4","WH South",  "warehouse",280,"Bangalore, Karnataka",  77.59,12.97))
    sc.add_node(Node("D1","Delhi Market",    "demand",200,"Delhi, Delhi",          77.21,28.61))
    sc.add_node(Node("D2","Jaipur Market",   "demand",120,"Jaipur, Rajasthan",    75.79,26.91))
    sc.add_node(Node("D3","Surat Market",    "demand",180,"Surat, Gujarat",       72.83,21.17))
    sc.add_node(Node("D4","Bhubaneswar Mkt", "demand",100,"Bhubaneswar, Odisha", 85.82,20.30))
    sc.add_node(Node("D5","Hyderabad Market","demand",160,"Hyderabad, Telangana", 78.49,17.39))
    sc.add_node(Node("D6","Kochi Market",    "demand",140,"Kochi, Kerala",        76.27, 9.93))
    for src,tgt,cap,cost in [
        ("P1","W1",300,2.0),("P1","W2",250,1.5),("P2","W3",200,1.8),
        ("P2","W4",220,2.2),("P3","W2",180,1.2),("P3","W4",200,1.6),
        ("P1","W3",150,3.0),("P2","W1",100,3.5),("W1","D1",220,1.0),
        ("W1","D2",150,1.5),("W2","D2",130,1.2),("W2","D3",200,1.0),
        ("W3","D4",120,1.0),("W3","D5",130,1.8),("W4","D5",180,1.0),
        ("W4","D6",160,1.2),("W1","D4",80,2.5),("W2","D6",90,2.0)]:
        sc.add_edge(Edge(src,tgt,cap,cost))
    return sc

def load_demo_inventory():
    inv=InventoryManager()
    for iid,name,unit in [("SKU001","Rice","Tonnes"),("SKU002","Wheat","Tonnes"),
                           ("SKU003","Sugar","Tonnes"),("SKU004","Edible Oil","KL")]:
        inv.add_item(iid,name,unit)
    for nid,iid,cur,saf,reo,dd in [
        ("P1","SKU001",520,100,150,22),("P1","SKU002",410,80,120,16),
        ("P2","SKU002",360,70,110,14),("P2","SKU003",310,60,90,11),
        ("P3","SKU001",285,55,85,11),("P3","SKU004",205,40,65,9),
        ("W1","SKU001",185,50,75,9),("W1","SKU002",155,40,65,7),
        ("W2","SKU001",125,30,55,6),("W2","SKU003",42,25,45,5),
        ("W2","SKU004",82,20,35,4),("W3","SKU002",92,20,38,5),
        ("W3","SKU003",18,15,28,4),
        ("W4","SKU001",62,15,28,4),("W4","SKU004",52,12,22,3)]:
        inv.set_stock(nid,iid,cur,saf,reo,dd)
    return inv

def load_demo_scores():
    return {"P1":{"reliability":92,"lead_time":88,"quality":95,"cost_efficiency":78},
            "P2":{"reliability":85,"lead_time":91,"quality":89,"cost_efficiency":84},
            "P3":{"reliability":79,"lead_time":83,"quality":91,"cost_efficiency":90},
            "W1":{"reliability":94,"lead_time":90,"quality":87,"cost_efficiency":75},
            "W2":{"reliability":88,"lead_time":85,"quality":82,"cost_efficiency":88},
            "W3":{"reliability":76,"lead_time":79,"quality":84,"cost_efficiency":92},
            "W4":{"reliability":82,"lead_time":86,"quality":88,"cost_efficiency":85}}

def create_excel_template():
    buf=io.BytesIO()
    with pd.ExcelWriter(buf,engine="openpyxl") as w:
        pd.DataFrame({"id":["P1","W1","D1"],"name":["Plant 1","Warehouse 1","Market 1"],
            "node_type":["plant","warehouse","demand"],"capacity":[500,400,200],
            "location":["Mumbai, Maharashtra","Delhi, Delhi","Jaipur, Rajasthan"],
            "x_longitude":[72.88,77.21,75.79],"y_latitude":[19.08,28.61,26.91]
        }).to_excel(w,"Nodes",index=False)
        pd.DataFrame({"source":["P1","W1"],"target":["W1","D1"],"capacity":[300,200],"cost":[2.0,1.0]
        }).to_excel(w,"Connections",index=False)
        pd.DataFrame({"node_id":["P1","W1"],"item_id":["SKU001","SKU001"],
            "item_name":["Rice","Rice"],"unit":["Tonnes","Tonnes"],
            "current_stock":[500,180],"safety_stock":[100,50],"reorder_point":[150,70],"daily_demand":[20,8]
        }).to_excel(w,"Inventory",index=False)
        pd.DataFrame({"date":["2024-01-01","2024-01-02"],"node_id":["D1","D1"],"demand":[198,205]
        }).to_excel(w,"Historical_Demand",index=False)
    buf.seek(0); return buf.getvalue()


# ═══════════════════════════════════════════════════════════════
# VISUALIZATIONS
# ═══════════════════════════════════════════════════════════════
# ── Performance: graph hash for cache invalidation ──────────────────
def _graph_hash(sc):
    """Cheap fingerprint: node count + edge count + sum of capacities"""
    return hash((len(sc.nodes), len(sc.edges),
                 sum(e.capacity for e in sc.edges),
                 sum(n.capacity for n in sc.nodes.values())))

def get_fulfillment_cached(sc):
    """Return cached demand fulfillment; recompute only when graph changes"""
    h = _graph_hash(sc)
    if (st.session_state.get("_ff_hash") == h and
            st.session_state.get("_ff_cache") is not None):
        return st.session_state["_ff_cache"]
    result = sc.demand_fulfillment()
    st.session_state["_ff_cache"] = result
    st.session_state["_ff_hash"]  = h
    return result

NC={"plant":"#0F6E56","warehouse":"#185FA5","demand":"#993C1D"}
NS={"plant":"square","warehouse":"diamond","demand":"circle"}
SEV={"critical":"#E74C3C","high":"#E67E22","medium":"#3498DB","low":"#27AE60","none":"#95A5A6"}

# Green → Yellow → Red heatmap (proper, visible)
HEATMAP_COLORSCALE=[
    [0.0,  "#27AE60"],   # bright green — safe
    [0.25, "#82E0AA"],   # light green
    [0.50, "#F9E79F"],   # yellow — moderate
    [0.65, "#F0B27A"],   # orange
    [0.80, "#E74C3C"],   # red — high risk
    [1.0,  "#922B21"],   # dark red — critical
]

def _auto_layout(sc):
    layers={"plant":[],"warehouse":[],"demand":[]}
    for nid,n in sc.nodes.items(): layers.get(n.node_type,layers["warehouse"]).append(nid)
    pos={}
    for lname,nids in layers.items():
        x={"plant":0.0,"warehouse":1.0,"demand":2.0}[lname]; n=len(nids)
        for i,nid in enumerate(nids): pos[nid]=(x,(i-(n-1)/2)*1.5)
    return pos

def draw_network(sc, highlight_path=None, disrupted_edge=None, show_cap=True, in_transit=None):
    pos=_auto_layout(sc); hp=highlight_path or []; it=in_transit or []
    hp_set=set(zip(hp,hp[1:])) if len(hp)>1 else set()
    it_edges={(d.get("from_id",""),d.get("to_id","")) for d in it if d.get("status")=="In Transit"}
    large_graph = len(sc.edges) > 60  # skip arrows and cap labels for big graphs
    traces=[]
    for e in sc.edges:
        x0,y0=pos[e.source]; x1,y1=pos[e.target]
        is_dis=disrupted_edge and e.source==disrupted_edge[0] and e.target==disrupted_edge[1]
        is_hi=(e.source,e.target) in hp_set; is_it=(e.source,e.target) in it_edges
        col="#E74C3C" if is_dis else "#E67E22" if is_hi else "#2980B9" if is_it else "#BDC3C7"
        w=3 if is_hi else 2.5 if is_dis else 2.5 if is_it else 1.5
        dash="dot" if(not e.active or is_dis) else "solid"
        ht=f"<b>{sc.nodes[e.source].name} → {sc.nodes[e.target].name}</b><br>Capacity: {e.capacity} | Cost: {e.cost}"
        traces.append(go.Scatter(x=[x0,x1,None],y=[y0,y1,None],mode="lines",
            line=dict(color=col,width=w,dash=dash),hovertext=ht,hoverinfo="text",showlegend=False))
        if show_cap and not large_graph:
            mx,my=(x0+x1)/2,(y0+y1)/2
            traces.append(go.Scatter(x=[mx],y=[my],mode="text",text=[f"{int(e.capacity)}"],
                textfont=dict(size=9,color=col),showlegend=False,hoverinfo="skip"))
        if not large_graph:
            dx,dy=x1-x0,y1-y0; L=math.hypot(dx,dy)
            if L>0:
                ux,uy=dx/L,dy/L
                traces.append(go.Scatter(x=[x1-ux*0.07],y=[y1-uy*0.07],mode="markers",
                    marker=dict(symbol="arrow",size=10,color=col,angle=math.degrees(math.atan2(-dy,dx))+90),
                    showlegend=False,hoverinfo="skip"))
    for ntype in ["plant","warehouse","demand"]:
        nids=[n for n,nd in sc.nodes.items() if nd.node_type==ntype]
        if not nids: continue
        in_p=[n in hp for n in nids]
        traces.append(go.Scatter(
            x=[pos[n][0] for n in nids],y=[pos[n][1] for n in nids],
            mode="markers+text",name=ntype.capitalize()+"s",
            text=[sc.nodes[n].name for n in nids],
            textposition="middle left" if ntype=="plant" else "top center" if ntype=="warehouse" else "middle right",
            textfont=dict(size=11,color="#2C3E50"),
            hovertext=[f"<b>{sc.nodes[n].name}</b><br>{ntype}<br>Cap:{sc.nodes[n].capacity}<br>{sc.nodes[n].location}" for n in nids],
            hoverinfo="text",
            marker=dict(symbol=NS[ntype],size=[22 if ip else 15 for ip in in_p],
                color=["#E67E22" if ip else NC[ntype] for ip in in_p],
                line=dict(width=2,color="white"))))
    fig=go.Figure(data=traces)
    fig.update_layout(paper_bgcolor="#FFFFFF",plot_bgcolor="#FFFFFF",
        margin=dict(l=10,r=10,t=10,b=10),height=460,hovermode="closest",
        xaxis=dict(showgrid=False,zeroline=False,showticklabels=False),
        yaxis=dict(showgrid=False,zeroline=False,showticklabels=False),
        legend=dict(orientation="h",yanchor="bottom",y=1.01,xanchor="right",x=1,
            font=dict(size=11,color="#2C3E50"),bgcolor="rgba(0,0,0,0)"))
    return fig

def draw_gauge_charts(ff, nodes):
    demands=list(ff.items()); cols=min(3,len(demands)); rows=math.ceil(len(demands)/cols)
    specs=[[{"type":"indicator"} for _ in range(cols)] for _ in range(rows)]
    fig=make_subplots(rows=rows,cols=cols,specs=specs)
    for i,(d_id,info) in enumerate(demands):
        c=(i%cols)+1; r=(i//cols)+1; pct=info["pct"]
        col="#E74C3C" if pct<50 else "#E67E22" if pct<80 else "#27AE60"
        fig.add_trace(go.Indicator(mode="gauge+number",value=pct,
            title={"text":nodes[d_id].name,"font":{"size":10,"color":"#2C3E50"}},
            number={"suffix":"%","font":{"size":16,"color":col}},
            gauge={"axis":{"range":[0,100],"tickwidth":1},
                   "bar":{"color":col},
                   "steps":[{"range":[0,50],"color":"#FADBD8"},{"range":[50,80],"color":"#FDEBD0"},{"range":[80,100],"color":"#D5F5E3"}]}),
            row=r,col=c)
    fig.update_layout(height=210*rows,paper_bgcolor="#FFFFFF",margin=dict(l=5,r=5,t=20,b=5))
    return fig

def draw_impact_chart(impact):
    names=[v["demand_name"] for v in impact.values()]
    before=[v["before_pct"] for v in impact.values()]; after=[v["after_pct"] for v in impact.values()]
    fig=go.Figure()
    fig.add_trace(go.Bar(name="Before",x=names,y=before,marker_color="#27AE60",
        text=[f"{v}%" for v in before],textposition="outside",textfont=dict(size=10,color="#2C3E50")))
    fig.add_trace(go.Bar(name="After",x=names,y=after,marker_color="#E74C3C",
        text=[f"{v}%" for v in after],textposition="outside",textfont=dict(size=10,color="#2C3E50")))
    fig.update_layout(barmode="group",yaxis=dict(title="Fulfillment (%)",range=[0,118],gridcolor="#F1F5F9"),
        xaxis=dict(tickfont=dict(color="#64748B",size=11)),paper_bgcolor="#FFFFFF",plot_bgcolor="#FFFFFF",height=300,
        legend=dict(orientation="h",yanchor="bottom",y=1.01,xanchor="right",x=1),margin=dict(l=10,r=10,t=10,b=10))
    fig.add_hline(y=100,line_dash="dot",line_color="#AAB7B8")
    return fig

def draw_heatmap(sc, ranking):
    """Green-Red heatmap — safe=green, danger=red"""
    if not ranking: return go.Figure()
    node_ids=list(sc.nodes.keys()); node_names=[sc.nodes[n].name for n in node_ids]
    z=[]
    for nid in node_ids:
        outs=[r for r in ranking if r["source"]==nid]; ins=[r for r in ranking if r["target"]==nid]
        max_out=max((r["avg_fulfillment_drop"] for r in outs),default=0)
        max_in =max((r["avg_fulfillment_drop"] for r in ins),default=0)
        nc_out=sum(1 for r in outs if r["severity"]=="critical")*25
        nc_in =sum(1 for r in ins if r["severity"]=="critical")*25
        z.append([max_out,max_in,nc_out,nc_in])
    fig=go.Figure(go.Heatmap(
        z=z, x=["Outbound Risk","Inbound Risk","Critical Out Links","Critical In Links"],
        y=node_names,
        colorscale=HEATMAP_COLORSCALE,
        text=[[f"<b>{v:.0f}</b>" for v in row] for row in z],
        texttemplate="%{text}",textfont={"size":12,"color":"#2C3E50"},
        hoverongaps=False,showscale=True,zmin=0,zmax=100,
        colorbar=dict(title=dict(text="Risk Level",font=dict(color="#0F172A",size=11)),
                     tickfont=dict(color="#64748B",size=11),
                     tickvals=[0,25,50,75,100],ticktext=["0 — Safe","25","50 — Med","75","100 — Critical"])))
    fig.update_layout(
        height=max(350,len(node_ids)*48),
        paper_bgcolor="#FFFFFF",plot_bgcolor="#FFFFFF",
        margin=dict(l=10,r=10,t=50,b=10),
        title=dict(text="<b>Node Risk Exposure Matrix</b>  (Green = Safe → Red = Critical)",
                   font=dict(color="#0F172A",size=13),x=0.5),
        xaxis=dict(side="top",tickfont=dict(size=11,color="#2C3E50")),
        yaxis=dict(tickfont=dict(size=11,color="#2C3E50"),autorange="reversed"))
    return fig

def draw_criticality_chart(ranking):
    if not ranking: return go.Figure()
    labels=[r["label"] for r in ranking]; drops=[r["avg_fulfillment_drop"] for r in ranking]
    colors=[SEV[r["severity"]] for r in ranking]
    fig=go.Figure(go.Bar(x=drops,y=labels,orientation="h",marker_color=colors,
        text=[f"  {d}%" for d in drops],textposition="outside",textfont=dict(size=11,color="#2C3E50")))
    fig.update_layout(
        xaxis=dict(title="<b>Avg. Fulfillment Drop (%)</b>",range=[0,max(drops or [10])*1.35],gridcolor="#F1F5F9"),
        yaxis=dict(autorange="reversed",tickfont=dict(size=11,color="#2C3E50")),
        paper_bgcolor="#FFFFFF",plot_bgcolor="#FFFFFF",
        height=max(320,len(ranking)*44),margin=dict(l=10,r=80,t=10,b=10))
    return fig

def draw_resilience_gauge(score):
    col="#E74C3C" if score<40 else "#E67E22" if score<70 else "#27AE60"
    fig=go.Figure(go.Indicator(mode="gauge+number",value=score,
        number={"suffix":"%","font":{"size":36,"color":col}},
        gauge={"axis":{"range":[0,100]},"bar":{"color":col},
               "steps":[{"range":[0,40],"color":"#FADBD8"},{"range":[40,70],"color":"#FDEBD0"},{"range":[70,100],"color":"#D5F5E3"}],
               "threshold":{"line":{"color":col,"width":3},"thickness":0.75,"value":score}}))
    fig.update_layout(height=200,margin=dict(l=20,r=20,t=10,b=10),paper_bgcolor="#FFFFFF")
    return fig

def draw_forecast_chart(fc, node_id, sc_nodes):
    if node_id not in fc.history or node_id not in fc.forecasts: return go.Figure()
    hist=fc.history[node_id]; fore=fc.forecasts[node_id]
    fig=go.Figure()
    fig.add_trace(go.Scatter(x=hist["date"],y=hist["demand"],mode="lines",
        name="Historical",line=dict(color="#2980B9",width=1.5),opacity=0.8))
    fig.add_trace(go.Scatter(x=fore["date"],y=fore["forecast"],mode="lines",
        name="Forecast",line=dict(color="#E67E22",width=2.5,dash="dash")))
    fig.add_trace(go.Scatter(
        x=list(fore["date"])+list(fore["date"][::-1]),
        y=list(fore["upper"])+list(fore["lower"][::-1]),
        fill="toself",fillcolor="rgba(230,126,34,0.12)",
        line=dict(color="rgba(0,0,0,0)"),name="CI Band"))
    node_name=sc_nodes[node_id].name if node_id in sc_nodes else node_id
    fig.update_layout(
        title=dict(text=f"<b>Demand Forecast — {node_name}</b>",font=dict(color="#0F172A",size=13),x=0.5),
        xaxis=dict(gridcolor="#F1F5F9"),yaxis=dict(title="Units",gridcolor="#F1F5F9"),
        paper_bgcolor="#FFFFFF",plot_bgcolor="#FFFFFF",height=360,
        legend=dict(font=dict(color="#0F172A"),bgcolor="rgba(0,0,0,0)"),
        margin=dict(l=10,r=10,t=40,b=10))
    return fig

def draw_scorecard_radar(name, scores):
    cats=["Reliability","Lead Time","Quality","Cost Efficiency","Reliability"]
    vals=[scores["reliability"],scores["lead_time"],scores["quality"],scores["cost_efficiency"],scores["reliability"]]
    fig=go.Figure(go.Scatterpolar(r=vals,theta=cats,fill="toself",
        line=dict(color="#1B4F72",width=2),fillcolor="rgba(27,79,114,0.15)"))
    fig.update_layout(polar=dict(radialaxis=dict(visible=True,range=[0,100],gridcolor="#E5E8E8")),
        showlegend=False,height=280,paper_bgcolor="#FFFFFF",
        margin=dict(l=30,r=30,t=30,b=10),
        title=dict(text=f"<b>{name}</b>",font=dict(size=12,color="#2C3E50"),x=0.5))
    return fig

def draw_geo_map(sc, scope="india"):
    fig=go.Figure()
    for e in sc.edges:
        s=sc.nodes[e.source]; t=sc.nodes[e.target]
        if s.x and s.y and t.x and t.y:
            fig.add_trace(go.Scattergeo(lon=[s.x,t.x,None],lat=[s.y,t.y,None],
                mode="lines",line=dict(width=1.5,color="#AAB7B8"),showlegend=False,hoverinfo="skip"))
    for ntype,nlist,col,sym,sz in [
        ("plant",[n for n in sc.nodes.values() if n.node_type=="plant"],"#0F6E56","square",14),
        ("warehouse",[n for n in sc.nodes.values() if n.node_type=="warehouse"],"#185FA5","diamond",12),
        ("demand",[n for n in sc.nodes.values() if n.node_type=="demand"],"#993C1D","circle",10)]:
        if nlist:
            fig.add_trace(go.Scattergeo(lon=[n.x for n in nlist],lat=[n.y for n in nlist],
                mode="markers+text",name=ntype.capitalize()+"s",
                text=[n.name for n in nlist],textposition="top center",textfont=dict(size=9,color="#2C3E50"),
                hovertext=[f"<b>{n.name}</b><br>{ntype}<br>{n.location}" for n in nlist],hoverinfo="text",
                marker=dict(size=sz,color=col,symbol=sym,line=dict(width=1.5,color="white"))))
    sc2="asia" if scope=="india" else "world"
    geo=dict(scope=sc2,showland=True,landcolor="#F2F3F4",showocean=True,oceancolor="#EBF5FB",
             showcountries=True,countrycolor="#BDC3C7",showcoastlines=True,coastlinecolor="#85929E",bgcolor="#FDFEFE")
    if scope=="india": geo.update(center=dict(lon=80,lat=22),projection_scale=4.5)
    fig.update_layout(geo=geo,height=500,paper_bgcolor="#FFFFFF",
        margin=dict(l=0,r=0,t=0,b=0),
        legend=dict(orientation="h",yanchor="bottom",y=1.01,xanchor="right",x=1,font=dict(size=11,color="#2C3E50")))
    return fig


# ═══════════════════════════════════════════════════════════════
# FREE AI (Groq — free tier with llama3)
# ═══════════════════════════════════════════════════════════════
FREE_AI_MODELS = {
    "Groq — Llama 3.1 8B (Fast, Free)": {
        "provider":"groq","model":"llama-3.1-8b-instant","url":"https://api.groq.com/openai/v1/chat/completions"
    },
    "Groq — Llama 3.3 70B (Smart, Free)": {
        "provider":"groq","model":"llama-3.3-70b-versatile","url":"https://api.groq.com/openai/v1/chat/completions"
    },
    "Groq — Gemma 2 9B (Free)": {
        "provider":"groq","model":"gemma2-9b-it","url":"https://api.groq.com/openai/v1/chat/completions"
    },
}

GROQ_SIGNUP = "https://console.groq.com"  # Free signup, no credit card needed

SC_SYSTEM_PROMPT = """You are an expert AI assistant for SR's Supply Chain & Operations Helper platform.
You help with supply chain management, inventory optimization, demand forecasting, risk analysis, and logistics.

When users describe actions (e.g. "I added 100 units of Rice"), extract:
- Intent (update_stock, find_path, check_status, etc.)
- Entities (node name, item, quantity)
- Then include an action block if appropriate:

ACTION_JSON_START
{"action":"update_stock","node":"Plant Mumbai","item":"Rice","quantity":100,"operation":"add"}
ACTION_JSON_END

Available actions: update_stock, find_path, run_disruption, get_status, check_atp, run_forecast
Always ask for confirmation before modifying data. Be concise and professional."""

def call_free_ai(messages, api_key, model_config):
    if not api_key or not api_key.strip():
        return None, "no_key"
    try:
        r=requests.post(
            model_config["url"],
            headers={"Authorization":f"Bearer {api_key.strip()}","Content-Type":"application/json"},
            json={"model":model_config["model"],"messages":[{"role":"system","content":SC_SYSTEM_PROMPT}]+messages,"max_tokens":1200,"temperature":0.7},
            timeout=30)
        if r.status_code==200:
            data=r.json()
            if "choices" in data and data["choices"]:
                return data["choices"][0]["message"]["content"], "ok"
            return None,"empty"
        elif r.status_code==401: return None,"invalid_key"
        elif r.status_code==429: return None,"rate_limit"
        else: return None,f"error_{r.status_code}"
    except requests.exceptions.Timeout: return None,"timeout"
    except Exception as ex: return None,f"exception_{str(ex)}"

def parse_action(text):
    match=re.search(r"ACTION_JSON_START\s*(.*?)\s*ACTION_JSON_END",text,re.DOTALL)
    if match:
        try: return json.loads(match.group(1).strip())
        except: return None
    return None

def clean_response(text):
    return re.sub(r"ACTION_JSON_START.*?ACTION_JSON_END","",text,flags=re.DOTALL).strip()

def execute_action(action, sc, inv):
    a=action.get("action","")
    try:
        if a=="update_stock":
            nn=action.get("node",""); iname=action.get("item","")
            qty=float(action.get("quantity",0)); op=action.get("operation","add")
            nid=next((n.id for n in sc.nodes.values() if n.name.lower()==nn.lower()),None)
            iid=next((id for id,iv in inv.items.items() if iv["name"].lower()==iname.lower()),None)
            if not nid: return False,f"Node '{nn}' not found"
            if not iid: return False,f"Item '{iname}' not found"
            delta=qty if op=="add" else -qty if op=="remove" else None
            if delta is not None:
                inv.update_stock(nid,iid,delta)
                return True,f"{'Added' if op=='add' else 'Removed'} {qty:.0f} units of {iname} {'to' if op=='add' else 'from'} {nn}"
        elif a=="find_path":
            frm=action.get("from",""); to=action.get("to","")
            fid=next((n.id for n in sc.nodes.values() if n.name.lower()==frm.lower()),None)
            tid=next((n.id for n in sc.nodes.values() if n.name.lower()==to.lower()),None)
            if not fid or not tid: return False,"One or both nodes not found"
            res=sc.all_shortest_paths(fid,tid,k=3)
            if res:
                st.session_state.highlight_path=res[0]["path"]
                pnames=" → ".join(sc.nodes[n].name for n in res[0]["path"])
                return True,f"Path highlighted: {pnames}"
            return False,"No path found"
        elif a=="get_status":
            plants=[n for n in sc.nodes.values() if n.node_type=="plant"]
            al=inv.get_alerts(sc.nodes)
            return True,(f"Network: {len(plants)} plants, {len(sc.edges)} connections. "
                        f"Alerts: {len([x for x in al if x['level']=='critical'])} critical.")
    except Exception as ex: return False,f"Action failed: {str(ex)}"
    return False,"Unknown action"


# ═══════════════════════════════════════════════════════════════
# CSS — Clean Light Professional Theme
# ═══════════════════════════════════════════════════════════════
APP_CSS = """
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800&display=swap');

*{box-sizing:border-box;}
html,body,[class*="css"]{font-family:'Inter',-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;font-size:14px;}

/* ─── App background ─────────────────────────────── */
.main .block-container{background:#F9FAFB;padding-top:1.2rem;padding-bottom:2rem;}

/* ─── Header ─────────────────────────────────────── */
.app-header{
  background:#0F172A;
  border-radius:12px;padding:18px 26px;
  display:flex;align-items:center;gap:20px;
  margin-bottom:24px;
  border:1px solid rgba(255,255,255,0.06);
  box-shadow:0 8px 32px rgba(0,0,0,0.24);
}
.app-header h1{color:#F8FAFC;font-size:17px;font-weight:700;margin:0;letter-spacing:-0.2px;}
.app-header .tagline{color:#64748B;font-size:10px;margin:3px 0 0;text-transform:uppercase;letter-spacing:2px;font-weight:500;}
.app-header .ver-badge{
  margin-left:auto;background:rgba(255,255,255,0.06);
  border:1px solid rgba(255,255,255,0.10);
  color:#94A3B8;font-size:10px;font-weight:600;
  padding:4px 12px;border-radius:20px;letter-spacing:1px;text-transform:uppercase;
  white-space:nowrap;
}

/* ─── Section Header ─────────────────────────────── */
.sh{
  font-size:9.5px;font-weight:800;color:#94A3B8;
  text-transform:uppercase;letter-spacing:2px;
  padding-bottom:9px;margin-bottom:16px;margin-top:6px;
  border-bottom:1px solid #E2E8F0;
}

/* ─── KPI Cards ──────────────────────────────────── */
.kpi{
  background:#FFFFFF;border:1px solid #E2E8F0;border-radius:10px;
  padding:16px 20px;border-top:3px solid #1E40AF;
  box-shadow:0 1px 3px rgba(0,0,0,0.06);
  transition:transform .15s,box-shadow .15s;
}
.kpi:hover{transform:translateY(-2px);box-shadow:0 6px 20px rgba(0,0,0,0.09);}
.kpi-lbl{font-size:9.5px;color:#94A3B8;font-weight:700;text-transform:uppercase;letter-spacing:1px;margin-bottom:6px;}
.kpi-val{font-size:24px;font-weight:800;color:#0F172A;line-height:1;letter-spacing:-1px;}
.kpi-sub{font-size:10px;color:#94A3B8;margin-top:5px;font-weight:400;}

/* ─── Cards ──────────────────────────────────────── */
.card{background:#FFFFFF;border:1px solid #E2E8F0;border-radius:8px;
  padding:14px 18px;margin:5px 0;font-size:13px;color:#1E293B;
  box-shadow:0 1px 2px rgba(0,0,0,0.04);}
.section-card{background:#FFFFFF;border:1px solid #E2E8F0;border-radius:12px;
  padding:20px 24px;margin-bottom:18px;box-shadow:0 1px 4px rgba(0,0,0,0.05);}

/* ─── Status Alerts ──────────────────────────────── */
.al-r{background:#FEF2F2;border-left:3px solid #DC2626;color:#1E293B;
  padding:10px 16px;border-radius:0 8px 8px 0;margin:6px 0;font-size:13px;}
.al-a{background:#FFFBEB;border-left:3px solid #D97706;color:#1E293B;
  padding:10px 16px;border-radius:0 8px 8px 0;margin:6px 0;font-size:13px;}
.al-g{background:#F0FDF4;border-left:3px solid #16A34A;color:#1E293B;
  padding:10px 16px;border-radius:0 8px 8px 0;margin:6px 0;font-size:13px;}
.al-b{background:#EFF6FF;border-left:3px solid #2563EB;color:#1E293B;
  padding:10px 16px;border-radius:0 8px 8px 0;margin:6px 0;font-size:13px;}

/* ─── Inline Badges ──────────────────────────────── */
.b-r{background:#FEE2E2;color:#991B1B;padding:2px 9px;border-radius:4px;font-size:11px;font-weight:700;}
.b-a{background:#FEF3C7;color:#92400E;padding:2px 9px;border-radius:4px;font-size:11px;font-weight:700;}
.b-g{background:#DCFCE7;color:#14532D;padding:2px 9px;border-radius:4px;font-size:11px;font-weight:700;}
.b-b{background:#DBEAFE;color:#1E40AF;padding:2px 9px;border-radius:4px;font-size:11px;font-weight:700;}

/* ─── Transport Result ───────────────────────────── */
.transport-card{background:#FFFFFF;border:1px solid #E2E8F0;border-radius:10px;
  padding:16px 20px;margin:8px 0;box-shadow:0 1px 3px rgba(0,0,0,0.04);}
.transport-mode{font-size:14px;font-weight:800;color:#0F172A;letter-spacing:-0.2px;}
.transport-detail{font-size:12px;color:#64748B;margin-top:6px;line-height:1.7;}
.transport-reason{font-size:11px;color:#94A3B8;margin-top:4px;font-style:italic;}
.transport-route{font-size:13px;font-weight:600;color:#1E293B;}

/* ─── Buttons ────────────────────────────────────── */
.stButton>button{
  background:#0F172A!important;color:#FFFFFF!important;
  border:none!important;border-radius:7px!important;
  font-weight:600!important;font-size:13px!important;
  padding:0.45rem 1.1rem!important;letter-spacing:0.1px!important;
  box-shadow:0 1px 3px rgba(0,0,0,0.15)!important;
  transition:background .15s,box-shadow .15s,transform .1s!important;
}
.stButton>button:hover{background:#1E293B!important;box-shadow:0 4px 12px rgba(0,0,0,0.2)!important;}
.stButton>button:active{transform:scale(0.98)!important;}
.stDownloadButton>button{background:#0F172A!important;color:#FFFFFF!important;border:none!important;border-radius:7px!important;font-weight:600!important;}

/* ─── Inputs ─────────────────────────────────────── */
.stTextInput>div>div>input,
.stNumberInput>div>div>input,
.stTextArea>div>textarea{
  border:1px solid #D1D5DB!important;border-radius:7px!important;
  font-size:13px!important;background:#FFFFFF!important;color:#0F172A!important;
}
.stTextInput>div>div>input:focus,
.stNumberInput>div>div>input:focus{
  border-color:#1E40AF!important;box-shadow:0 0 0 3px rgba(30,64,175,0.1)!important;
}
.stSelectbox>div>div{
  border:1px solid #D1D5DB!important;border-radius:7px!important;
  background:#FFFFFF!important;font-size:13px!important;
}

/* ─── Tabs ───────────────────────────────────────── */
.stTabs [data-baseweb="tab-list"]{
  background:transparent!important;
  border-bottom:2px solid #E2E8F0!important;gap:0!important;
}
.stTabs [data-baseweb="tab"]{
  font-size:11.5px!important;font-weight:700!important;
  color:#64748B!important;text-transform:uppercase!important;
  letter-spacing:1px!important;padding:10px 18px!important;
  border-radius:0!important;background:transparent!important;
  border-bottom:2px solid transparent!important;margin-bottom:-2px!important;
}
.stTabs [aria-selected="true"]{
  color:#0F172A!important;border-bottom:2px solid #0F172A!important;
}

/* ─── Expander ───────────────────────────────────── */
div[data-testid="stExpander"]{
  border:1px solid #E2E8F0!important;border-radius:8px!important;
  background:#FFFFFF!important;box-shadow:0 1px 3px rgba(0,0,0,0.04)!important;
}

/* ─── Sidebar ────────────────────────────────────── */
section[data-testid="stSidebar"]{background:#FFFFFF!important;border-right:1px solid #E2E8F0!important;}
.sb-sec{
  font-size:9px;font-weight:800;color:#94A3B8;
  text-transform:uppercase;letter-spacing:2px;
  margin:14px 0 7px;padding-top:12px;
  border-top:1px solid #F1F5F9;
}
.sb-sec:first-child{border-top:none;padding-top:0;}

/* ─── Dataframe ──────────────────────────────────── */
[data-testid="stDataFrame"]{border-radius:8px!important;overflow:hidden;border:1px solid #E2E8F0;}
.dataframe{font-size:12px!important;}

/* ─── Chat ───────────────────────────────────────── */
.user-bubble{
  background:#0F172A;color:#F8FAFC;
  padding:10px 16px;border-radius:16px 16px 3px 16px;
  max-width:74%;font-size:13px;line-height:1.5;
  margin:5px 0;box-shadow:0 2px 8px rgba(15,23,42,0.2);
}
.ai-bubble{
  background:#FFFFFF;border:1px solid #E2E8F0;color:#1E293B;
  padding:10px 16px;border-radius:16px 16px 16px 3px;
  max-width:80%;font-size:13px;line-height:1.6;
  margin:5px 0;box-shadow:0 1px 3px rgba(0,0,0,0.04);
}
.chat-wrap{display:flex;margin:3px 0;}
.chat-right{justify-content:flex-end;}.chat-left{justify-content:flex-start;}

/* ─── Metrics ────────────────────────────────────── */
[data-testid="metric-container"]{
  background:#FFFFFF;border:1px solid #E2E8F0;border-radius:9px;padding:14px 16px;
}

/* ─── Scrollbar ──────────────────────────────────── */
::-webkit-scrollbar{width:5px;height:5px;}
::-webkit-scrollbar-track{background:#F8FAFC;}
::-webkit-scrollbar-thumb{background:#CBD5E1;border-radius:4px;}
::-webkit-scrollbar-thumb:hover{background:#94A3B8;}

/* ─── Slider ─────────────────────────────────────── */
.stSlider [data-baseweb="slider"] div{background:#0F172A!important;}
.stProgress>div>div{background:#0F172A!important;border-radius:4px!important;}
</style>
"""

# ═══════════════════════════════════════════════════════════════
# IMPORTS
# ═══════════════════════════════════════════════════════════════
import streamlit as st
import streamlit.components.v1 as components
import pandas as pd
import numpy as np
import networkx as nx
import plotly.graph_objects as go
import plotly.express as px
from plotly.subplots import make_subplots
from dataclasses import dataclass
from datetime import datetime, timedelta, date
from sklearn.ensemble import RandomForestRegressor, GradientBoostingRegressor
from sklearn.metrics import mean_squared_error, mean_absolute_error
import json, math, io, re, requests
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.utils import get_column_letter
import warnings
warnings.filterwarnings("ignore")

# ═══════════════════════════════════════════════════════════════
# CUSTOM SVG LOGO (SR + Supply Chain Network)
# ═══════════════════════════════════════════════════════════════
LOGO_SVG = """<svg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 120 120' width='60' height='60'>
  <defs>
    <linearGradient id='bg' x1='0%' y1='0%' x2='100%' y2='100%'>
      <stop offset='0%' style='stop-color:#1B4F72'/>
      <stop offset='100%' style='stop-color:#0F6E56'/>
    </linearGradient>
  </defs>
  <circle cx='60' cy='60' r='58' fill='url(#bg)' stroke='#FFFFFF' stroke-width='2.5'/>
  <circle cx='30' cy='38' r='8' fill='#F39C12' stroke='white' stroke-width='1.5'/>
  <circle cx='60' cy='22' r='8' fill='#27AE60' stroke='white' stroke-width='1.5'/>
  <circle cx='90' cy='38' r='8' fill='#E74C3C' stroke='white' stroke-width='1.5'/>
  <circle cx='38' cy='68' r='8' fill='#9B59B6' stroke='white' stroke-width='1.5'/>
  <circle cx='82' cy='68' r='8' fill='#2980B9' stroke='white' stroke-width='1.5'/>
  <circle cx='60' cy='88' r='8' fill='#F39C12' stroke='white' stroke-width='1.5'/>
  <line x1='30' y1='38' x2='60' y2='22' stroke='#F8C471' stroke-width='2.5'/>
  <line x1='60' y1='22' x2='90' y2='38' stroke='#7DCEA0' stroke-width='2.5'/>
  <line x1='30' y1='38' x2='38' y2='68' stroke='#F8C471' stroke-width='2.5'/>
  <line x1='90' y1='38' x2='82' y2='68' stroke='#7DCEA0' stroke-width='2.5'/>
  <line x1='38' y1='68' x2='60' y2='88' stroke='#AED6F1' stroke-width='2.5'/>
  <line x1='82' y1='68' x2='60' y2='88' stroke='#AED6F1' stroke-width='2.5'/>
  <line x1='38' y1='68' x2='82' y2='68' stroke='white' stroke-width='1.5' opacity='0.5' stroke-dasharray='3,2'/>
  <text x='60' y='110' text-anchor='middle' font-family='Arial,sans-serif' font-size='13' font-weight='900' fill='white' letter-spacing='3'>SR</text>
</svg>"""

LOGO_SIDEBAR = """<svg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 120 120' width='42' height='42'>
  <defs><linearGradient id='bg2' x1='0%' y1='0%' x2='100%' y2='100%'><stop offset='0%' style='stop-color:#1B4F72'/><stop offset='100%' style='stop-color:#0F6E56'/></linearGradient></defs>
  <circle cx='60' cy='60' r='58' fill='url(#bg2)' stroke='#FFFFFF' stroke-width='2'/>
  <circle cx='30' cy='38' r='7' fill='#F39C12' stroke='white' stroke-width='1.5'/>
  <circle cx='60' cy='22' r='7' fill='#27AE60' stroke='white' stroke-width='1.5'/>
  <circle cx='90' cy='38' r='7' fill='#E74C3C' stroke='white' stroke-width='1.5'/>
  <circle cx='38' cy='68' r='7' fill='#9B59B6' stroke='white' stroke-width='1.5'/>
  <circle cx='82' cy='68' r='7' fill='#2980B9' stroke='white' stroke-width='1.5'/>
  <circle cx='60' cy='88' r='7' fill='#F39C12' stroke='white' stroke-width='1.5'/>
  <line x1='30' y1='38' x2='60' y2='22' stroke='#F8C471' stroke-width='2'/>
  <line x1='60' y1='22' x2='90' y2='38' stroke='#7DCEA0' stroke-width='2'/>
  <line x1='30' y1='38' x2='38' y2='68' stroke='#F8C471' stroke-width='2'/>
  <line x1='90' y1='38' x2='82' y2='68' stroke='#7DCEA0' stroke-width='2'/>
  <line x1='38' y1='68' x2='60' y2='88' stroke='#AED6F1' stroke-width='2'/>
  <line x1='82' y1='68' x2='60' y2='88' stroke='#AED6F1' stroke-width='2'/>
  <text x='60' y='110' text-anchor='middle' font-family='Arial,sans-serif' font-size='12' font-weight='900' fill='white' letter-spacing='2'>SR</text>
</svg>"""

# ═══════════════════════════════════════════════════════════════
# DATA MODELS
# ═══════════════════════════════════════════════════════════════
@dataclass
class Node:
    id: str; name: str; node_type: str; capacity: float
    location: str = ""; x: float = 0.0; y: float = 0.0

@dataclass
class Edge:
    source: str; target: str; capacity: float
    cost: float = 1.0; active: bool = True

# ═══════════════════════════════════════════════════════════════
# TRANSPORT LOGIC
# ═══════════════════════════════════════════════════════════════
TRANSPORT_MODES = {
    "road":  {"speed_kmh": 60,  "cost_per_km": 2.5,  "label": "Road (Truck)",  "icon": "Truck"},
    "rail":  {"speed_kmh": 80,  "cost_per_km": 1.2,  "label": "Rail",          "icon": "Rail"},
    "air":   {"speed_kmh": 800, "cost_per_km": 12.0,  "label": "Air",           "icon": "Air"},
    "sea":   {"speed_kmh": 35,  "cost_per_km": 0.8,  "label": "Sea/Inland",    "icon": "Sea"},
}

def haversine_km(lat1, lon1, lat2, lon2):
    """Distance between two lat/lon points in km"""
    R = 6371
    dlat = math.radians(lat2 - lat1)
    dlon = math.radians(lon2 - lon1)
    a = math.sin(dlat/2)**2 + math.cos(math.radians(lat1))*math.cos(math.radians(lat2))*math.sin(dlon/2)**2
    return R * 2 * math.asin(math.sqrt(a))

def get_state(location_str):
    """Extract state from 'City, State' format"""
    parts = location_str.split(",")
    return parts[-1].strip().lower() if len(parts) > 1 else location_str.strip().lower()

def recommend_transport(src_node, tgt_node, load_units, urgent=False):
    """
    Smart transport mode recommendation based on:
    - Same city/state → road
    - Load size → truck / rail / air
    - Urgent → air
    - Distance → affects mode
    """
    src_loc = src_node.location; tgt_loc = tgt_node.location
    src_state = get_state(src_loc); tgt_state = get_state(tgt_loc)
    same_state = (src_state == tgt_state)

    # Compute distance if coordinates available
    dist_km = 0
    if src_node.x and src_node.y and tgt_node.x and tgt_node.y:
        dist_km = haversine_km(src_node.y, src_node.x, tgt_node.y, tgt_node.x)

    recommendations = []

    if urgent:
        mode = "air"
        reason = "Urgent delivery — air freight selected"
    elif same_state or dist_km < 150:
        mode = "road"
        reason = "Same region — road transport (cheapest)"
    elif load_units >= 500:
        mode = "rail"
        reason = f"High load ({load_units} units) — rail preferred"
    elif load_units >= 150:
        # Offer both
        mode = "rail"
        reason = f"Medium load ({load_units} units) — rail recommended"
    else:
        mode = "road"
        reason = f"Small load ({load_units} units) — road (truck)"

    m = TRANSPORT_MODES[mode]
    if dist_km > 0:
        cost = dist_km * m["cost_per_km"]
        hours = dist_km / m["speed_kmh"]
        days = hours / 24
    else:
        cost = None; hours = None; days = None

    return {
        "mode": mode, "label": m["label"], "icon": m["icon"],
        "reason": reason, "dist_km": round(dist_km, 1) if dist_km else None,
        "cost_estimate": round(cost, 0) if cost else None,
        "hours": round(hours, 1) if hours else None,
        "days": round(days, 2) if days else None,
        "same_state": same_state,
    }

# ═══════════════════════════════════════════════════════════════
# GRAPH ENGINE
# ═══════════════════════════════════════════════════════════════
class SupplyChainGraph:
    def __init__(self):
        self.nodes: dict = {}
        self.edges: list = []

    def add_node(self, node): self.nodes[node.id] = node

    def add_edge(self, edge):
        if edge.source not in self.nodes or edge.target not in self.nodes:
            raise ValueError(f"Node {edge.source} or {edge.target} not found")
        for e in self.edges:
            if e.source == edge.source and e.target == edge.target: return
        self.edges.append(edge)

    def remove_edge(self, src, tgt):
        self.edges = [e for e in self.edges if not (e.source==src and e.target==tgt)]

    def toggle_edge(self, src, tgt, active):
        for e in self.edges:
            if e.source==src and e.target==tgt: e.active = active

    def to_nx(self):
        G = nx.DiGraph()
        for nid, n in self.nodes.items(): G.add_node(nid, **vars(n))
        for e in self.edges:
            if e.active:
                G.add_edge(e.source, e.target, capacity=e.capacity, weight=e.cost)
        return G

    def geo_shortest_path(self, src, tgt):
        """Shortest path using real lat/lon distances as weights"""
        G = nx.DiGraph()
        for nid, n in self.nodes.items(): G.add_node(nid, **vars(n))
        for e in self.edges:
            if not e.active: continue
            sn = self.nodes[e.source]; tn = self.nodes[e.target]
            if sn.x and sn.y and tn.x and tn.y:
                dist = haversine_km(sn.y, sn.x, tn.y, tn.x)
                w = dist if dist > 0 else e.cost
            else:
                w = e.cost
            G.add_edge(e.source, e.target, capacity=e.capacity, weight=w, cost=e.cost)
        try:
            path = nx.dijkstra_path(G, src, tgt, weight="weight")
            total_dist = sum(
                haversine_km(self.nodes[path[i]].y, self.nodes[path[i]].x,
                             self.nodes[path[i+1]].y, self.nodes[path[i+1]].x)
                if (self.nodes[path[i]].x and self.nodes[path[i+1]].x) else 0
                for i in range(len(path)-1)
            )
            total_cost = sum(G[path[i]][path[i+1]].get("cost", 1) for i in range(len(path)-1))
            return {"found": True, "path": path, "dist_km": round(total_dist, 1), "cost": round(total_cost, 2)}
        except:
            return {"found": False, "path": [], "dist_km": None, "cost": None}

    def shortest_path(self, src, tgt):
        return self.geo_shortest_path(src, tgt)

    def all_shortest_paths(self, src, tgt, k=3):
        G = self.to_nx(); results = []
        try:
            for path in nx.shortest_simple_paths(G, src, tgt, weight="weight"):
                cost = sum(G[path[i]][path[i+1]]["weight"] for i in range(len(path)-1))
                results.append({"path": path, "cost": round(cost, 2)})
                if len(results) >= k: break
        except: pass
        return results

    def demand_fulfillment(self):
        G = self.to_nx()
        plants  = [n for n, d in self.nodes.items() if d.node_type=="plant"]
        demands = [n for n, d in self.nodes.items() if d.node_type=="demand"]
        results = {}
        for d_id in demands:
            req = self.nodes[d_id].capacity; total, reach = 0.0, []
            for p in plants:
                try:
                    fv, _ = nx.maximum_flow(G, p, d_id, capacity="capacity")
                    if fv > 0: total += fv; reach.append(p)
                except: pass
            ful = min(total, req)
            results[d_id] = {"required": req, "fulfilled": round(ful, 2),
                             "pct": round(ful/req*100 if req>0 else 100, 1),
                             "reachable_from": reach}
        return results

    def simulate_disruption(self, src, tgt):
        baseline = self.demand_fulfillment()
        self.toggle_edge(src, tgt, False)
        disrupted = self.demand_fulfillment()
        self.toggle_edge(src, tgt, True)
        impact = {}
        for d_id, base in baseline.items():
            dis = disrupted[d_id]; drop = base["pct"] - dis["pct"]
            impact[d_id] = {
                "demand_name": self.nodes[d_id].name,
                "before_pct": base["pct"], "after_pct": dis["pct"],
                "drop_pct": round(drop,1), "before_fulfilled": base["fulfilled"],
                "after_fulfilled": dis["fulfilled"],
                "lost_units": round(base["fulfilled"]-dis["fulfilled"],2),
                "severity": "critical" if drop>=50 else "high" if drop>=25 else "medium" if drop>0 else "none"
            }
        alt_paths = {}
        for d_id, imp in impact.items():
            if imp["drop_pct"] > 0:
                self.toggle_edge(src, tgt, False)
                paths = []
                for p in [n for n, nd in self.nodes.items() if nd.node_type=="plant"]:
                    r = self.shortest_path(p, d_id)
                    if r["found"]:
                        paths.append({"from_plant": self.nodes[p].name,
                                      "path": [self.nodes[n].name for n in r["path"]],
                                      "path_ids": r["path"], "cost": r["cost"]})
                paths.sort(key=lambda x: x["cost"])
                alt_paths[d_id] = paths[:3]
                self.toggle_edge(src, tgt, True)
        avg_drop = sum(v["drop_pct"] for v in impact.values())/len(impact) if impact else 0
        return {"removed_edge": f"{self.nodes[src].name} → {self.nodes[tgt].name}",
                "removed_src": src, "removed_tgt": tgt,
                "resilience_score": round(max(0, 100-avg_drop), 1),
                "impact": impact, "alt_paths": alt_paths}

    def rank_critical_edges(self):
        ranking = []
        for e in self.edges:
            if not e.active: continue
            r = self.simulate_disruption(e.source, e.target)
            avg = sum(v["drop_pct"] for v in r["impact"].values())/len(r["impact"]) if r["impact"] else 0
            ranking.append({"source":e.source,"target":e.target,"label":r["removed_edge"],
                "avg_fulfillment_drop":round(avg,1),"resilience_score":r["resilience_score"],
                "severity":"critical" if avg>=50 else "high" if avg>=25 else "medium" if avg>=5 else "low"})
        return sorted(ranking, key=lambda x: x["avg_fulfillment_drop"], reverse=True)

    def to_dict(self):
        return {"nodes":[vars(n) for n in self.nodes.values()],"edges":[vars(e) for e in self.edges]}


# ═══════════════════════════════════════════════════════════════
# INVENTORY MANAGER
# ═══════════════════════════════════════════════════════════════
class InventoryManager:
    def __init__(self):
        self.items = {}; self.stock = {}

    def add_item(self, iid, name, unit="units"):
        self.items[iid] = {"name": name, "unit": unit}

    def set_stock(self, node_id, iid, current, safety, reorder, daily_demand=1.0):
        if node_id not in self.stock: self.stock[node_id] = {}
        self.stock[node_id][iid] = {"current":float(current),"safety":float(safety),
                                     "reorder":float(reorder),"daily_demand":float(daily_demand)}

    def update_stock(self, node_id, iid, delta):
        if node_id in self.stock and iid in self.stock[node_id]:
            self.stock[node_id][iid]["current"] = max(0.0, self.stock[node_id][iid]["current"]+delta)
            return True
        return False

    def coverage_days(self, node_id, iid):
        if node_id in self.stock and iid in self.stock[node_id]:
            s = self.stock[node_id][iid]; dd = s.get("daily_demand",1)
            return round(s["current"]/dd,1) if dd>0 else 999
        return None

    def get_alerts(self, sc_nodes=None):
        alerts = []
        for node_id, items in self.stock.items():
            nn = sc_nodes[node_id].name if sc_nodes and node_id in sc_nodes else node_id
            for iid, s in items.items():
                iname = self.items.get(iid,{}).get("name",iid)
                if s["current"] <= s["reorder"]:
                    alerts.append({"node_id":node_id,"node_name":nn,"item_id":iid,"item_name":iname,
                        "level":"critical" if s["current"]<=s["safety"] else "warning",
                        "current":s["current"],"safety":s["safety"],"reorder":s["reorder"],
                        "coverage":self.coverage_days(node_id,iid)})
        return sorted(alerts, key=lambda x:(x["level"]!="critical", x["coverage"] or 999))

    def find_alternatives(self, demand_id, iid, required, sc, disrupted=None):
        alts = []
        for node_id, items in self.stock.items():
            if node_id==demand_id or iid not in items: continue
            s=items[iid]; avail=max(0.0, s["current"]-s["safety"])
            if avail<=0: continue
            if disrupted: sc.toggle_edge(disrupted[0], disrupted[1], False)
            r = sc.shortest_path(node_id, demand_id)
            if disrupted: sc.toggle_edge(disrupted[0], disrupted[1], True)
            if r["found"]:
                alts.append({"node_id":node_id,
                    "node_name":sc.nodes[node_id].name if node_id in sc.nodes else node_id,
                    "available":avail,"can_cover":avail>=required,
                    "coverage_pct":min(100, round(avail/max(required,1)*100,1)),
                    "path":[sc.nodes[n].name for n in r["path"] if n in sc.nodes],
                    "path_ids":r["path"],"route_cost":r["cost"],
                    "coverage_days":self.coverage_days(node_id,iid)})
        return sorted(alts, key=lambda x:(-x["coverage_pct"], x["route_cost"]))[:5]

    def to_df(self, sc_nodes=None):
        rows = []
        for node_id, items in self.stock.items():
            nn = sc_nodes[node_id].name if sc_nodes and node_id in sc_nodes else node_id
            nt = sc_nodes[node_id].node_type if sc_nodes and node_id in sc_nodes else ""
            for iid, s in items.items():
                iname=self.items.get(iid,{}).get("name",iid); unit=self.items.get(iid,{}).get("unit","units")
                dd=s.get("daily_demand",1); cov=round(s["current"]/dd,1) if dd>0 else 999
                status="Critical" if s["current"]<=s["safety"] else "Low" if s["current"]<=s["reorder"] else "Normal"
                rows.append({"Node":nn,"Type":nt.capitalize(),"Item":iname,"Unit":unit,
                    "Current Stock":s["current"],"Safety Stock":s["safety"],"Reorder Point":s["reorder"],
                    "Daily Demand":dd,"Coverage Days":cov,
                    "Available (above safety)":max(0,s["current"]-s["safety"]),"Status":status})
        return pd.DataFrame(rows)


# ═══════════════════════════════════════════════════════════════
# DEMAND FORECASTER
# ═══════════════════════════════════════════════════════════════
class DemandForecaster:
    def __init__(self):
        self.models={}; self.history={}; self.forecasts={}; self.metrics={}

    def generate_synthetic_history(self, node_id, node_name, base_demand, days=365, seed=None):
        if seed is None: seed=hash(node_id)%1000
        np.random.seed(seed); t=np.arange(days)
        demand=np.maximum(0, base_demand+0.05*t+base_demand*0.15*np.sin(2*np.pi*t/7)+
                          base_demand*0.20*np.sin(2*np.pi*t/365)+np.random.normal(0,base_demand*0.08,days))
        dates=[datetime.today()-timedelta(days=days-i) for i in range(days)]
        df=pd.DataFrame({"date":dates,"demand":demand,"node_id":node_id,"node_name":node_name})
        self.history[node_id]=df; return df

    def _make_features(self, y, lags=[1,7,14,30]):
        ml=max(lags); X,yo=[],[]
        for i in range(ml, len(y)):
            feats=[i,i%7,i%30,i//30]
            for lag in lags: feats.append(float(y[i-lag]))
            feats.append(float(np.mean(y[max(0,i-7):i]))); feats.append(float(np.mean(y[max(0,i-30):i])))
            X.append(feats); yo.append(float(y[i]))
        return np.array(X,dtype=float), np.array(yo,dtype=float)

    def train(self, node_id, horizon=30):
        if node_id not in self.history: return None
        df=self.history[node_id]; y=df["demand"].values
        X,yt=self._make_features(y); split=max(int(len(X)*0.80),1)
        rf=RandomForestRegressor(n_estimators=100,max_depth=6,random_state=42,n_jobs=-1)
        gbm=GradientBoostingRegressor(n_estimators=80,max_depth=4,learning_rate=0.1,random_state=42)
        rf.fit(X[:split],yt[:split]); gbm.fit(X[:split],yt[:split])
        if split<len(X):
            ep=0.5*rf.predict(X[split:])+0.5*gbm.predict(X[split:])
            yv=yt[split:]
            rmse=float(np.sqrt(mean_squared_error(yv,ep))); mae=float(mean_absolute_error(yv,ep))
            mape=float(np.mean(np.abs((yv-ep)/(yv+1e-6)))*100)
            r2=float(1-np.sum((yv-ep)**2)/np.sum((yv-np.mean(yv))**2))
        else:
            rmse=0;mae=0;mape=0;r2=1
        self.models[node_id]=(rf,gbm)
        self.metrics[node_id]={"rmse":round(rmse,2),"mae":round(mae,2),"mape":round(mape,2),"r2":round(r2,3)}
        hv=list(y); fp=[]
        for step in range(horizon):
            i=len(hv); lags=[1,7,14,30]
            feats=[i,i%7,i%30,i//30]
            for lag in lags: feats.append(float(hv[-lag]) if lag<=len(hv) else float(np.mean(hv[-min(lag,len(hv)):])))
            feats.append(float(np.mean(hv[-7:]))); feats.append(float(np.mean(hv[-30:])))
            arr=np.array(feats,dtype=float).reshape(1,-1)
            pred=max(0, 0.5*rf.predict(arr)[0]+0.5*gbm.predict(arr)[0])
            fp.append(pred); hv.append(pred)
        ld=df["date"].iloc[-1]; fd=[ld+timedelta(days=i+1) for i in range(horizon)]
        fdf=pd.DataFrame({"date":fd,"forecast":fp,"node_id":node_id,"node_name":df["node_name"].iloc[0]})
        fdf["upper"]=fdf["forecast"]+1.5*rmse; fdf["lower"]=np.maximum(0,fdf["forecast"]-1.5*rmse)
        self.forecasts[node_id]=fdf; return fdf

    def aggregate_to_warehouses(self, sc, horizon=30):
        wh_fc={}
        for wid, wh in sc.nodes.items():
            if wh.node_type!="warehouse": continue
            G=sc.to_nx(); served=[]
            for did, d in sc.nodes.items():
                if d.node_type=="demand":
                    try: nx.shortest_path(G,wid,did); served.append(did)
                    except: pass
            total=np.zeros(horizon)
            for did in served:
                if did in self.forecasts:
                    vals=self.forecasts[did]["forecast"].values[:horizon]; total[:len(vals)]+=vals
            wh_fc[wid]={"name":wh.name,"forecast":total.tolist(),"served_demands":served}
        return wh_fc

    def get_plant_requirements(self, sc, wh_fc, horizon=30):
        pr={}
        for pid, p in sc.nodes.items():
            if p.node_type!="plant": continue
            G=sc.to_nx(); served=[]
            for wid in wh_fc:
                try: nx.shortest_path(G,pid,wid); served.append(wid)
                except: pass
            total=np.zeros(horizon)
            for wid in served: total+=np.array(wh_fc[wid]["forecast"][:horizon])
            pr[pid]={"name":p.name,"required":total.tolist(),"capacity":p.capacity,"served_wh":served}
        return pr


# ═══════════════════════════════════════════════════════════════
# REPORT GENERATOR
# ═══════════════════════════════════════════════════════════════
def generate_excel_report(sc, inv, forecaster, ranking, start_date, end_date):
    """Comprehensive Excel report with color-coded data and embedded charts"""
    wb = openpyxl.Workbook()

    # Styles
    fill_green  = PatternFill("solid", fgColor="DCFCE7")
    fill_yellow = PatternFill("solid", fgColor="FEF3C7")
    fill_red    = PatternFill("solid", fgColor="FEE2E2")
    fill_blue   = PatternFill("solid", fgColor="DBEAFE")
    fill_hdr    = PatternFill("solid", fgColor="1E3A5F")
    fill_sub    = PatternFill("solid", fgColor="EFF6FF")
    fill_alt    = PatternFill("solid", fgColor="F8FAFC")

    fnt_hdr  = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
    fnt_title= Font(bold=True, size=15, color="1E3A5F", name="Calibri")
    fnt_sub  = Font(bold=True, size=12, color="334155", name="Calibri")
    fnt_bold = Font(bold=True, size=10, name="Calibri")
    fnt_norm = Font(size=10, name="Calibri")
    fnt_sm   = Font(size=9, color="64748B", name="Calibri")
    aln_c    = Alignment(horizontal="center", vertical="center", wrap_text=True)
    aln_l    = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    thin     = Side(style="thin", color="E2E8F0")
    bdr      = Border(left=thin, right=thin, top=thin, bottom=thin)

    def hdr_row(ws, row, cols):
        for c in range(1, cols+1):
            cell=ws.cell(row=row,column=c)
            cell.fill=fill_hdr; cell.font=fnt_hdr; cell.alignment=aln_c; cell.border=bdr

    def data_row(ws, row, cols, fill=None, alt=False):
        f = fill if fill else (fill_alt if alt else None)
        for c in range(1, cols+1):
            cell=ws.cell(row=row,column=c)
            if f: cell.fill=f
            cell.font=fnt_norm; cell.border=bdr; cell.alignment=aln_l

    # ═══ SHEET 1 — EXECUTIVE SUMMARY ═══════════════
    ws = wb.active; ws.title="Executive Summary"
    ws.sheet_view.showGridLines=False
    for col,w in zip("ABCDEFG",[28,16,16,16,16,16,14]):
        ws.column_dimensions[col].width=w

    ws.merge_cells("A1:G1")
    ws["A1"]="SUPPLY CHAIN & OPERATIONS — EXECUTIVE REPORT"
    ws["A1"].font=Font(bold=True,size=16,color="1E3A5F",name="Calibri")
    ws["A1"].alignment=aln_c
    ws["A1"].fill=PatternFill("solid",fgColor="EFF6FF")
    ws.row_dimensions[1].height=36

    ws.merge_cells("A2:G2")
    ws["A2"]=f"Period: {start_date} to {end_date}  |  Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}  |  SR Supply Chain & Operations Helper"
    ws["A2"].font=fnt_sm; ws["A2"].alignment=aln_c

    plants=[n for n in sc.nodes.values() if n.node_type=="plant"]
    whs_  =[n for n in sc.nodes.values() if n.node_type=="warehouse"]
    dems_ =[n for n in sc.nodes.values() if n.node_type=="demand"]
    total_cap=sum(n.capacity for n in plants); total_dem=sum(n.capacity for n in dems_)
    cov=round(total_cap/max(total_dem,1)*100,1)

    # KPI row
    kpis=[("Plants",len(plants),"1E3A5F"),("Warehouses",len(whs_),"1E3A5F"),
          ("Demand Points",len(dems_),"1E3A5F"),("Connections",len(sc.edges),"1E3A5F"),
          ("Plant Capacity",f"{total_cap:,.0f}","14532D"),
          ("Total Demand",f"{total_dem:,.0f}","1E3A5F"),
          ("Coverage",f"{cov}%","14532D" if cov>=100 else "91241C")]
    for c,(lbl,val,col) in enumerate(kpis,start=1):
        ws.cell(row=4,column=c,value=lbl).font=Font(size=9,bold=True,color="94A3B8",name="Calibri")
        ws.cell(row=4,column=c).alignment=aln_c
        ws.cell(row=5,column=c,value=val).font=Font(size=13,bold=True,color=col,name="Calibri")
        ws.cell(row=5,column=c).alignment=aln_c
        ws.cell(row=5,column=c).fill=fill_sub
    ws.row_dimensions[5].height=28

    # Demand fulfillment table
    ws.cell(row=7,column=1,value="DEMAND FULFILLMENT OVERVIEW").font=fnt_sub
    ff=sc.demand_fulfillment()
    hff=["Demand Point","Required","Fulfilled","Fulfillment %","Status"]
    for c,h in enumerate(hff,1): ws.cell(row=8,column=c,value=h)
    hdr_row(ws,8,len(hff))
    r=9
    for did,info in ff.items():
        pct=info["pct"]; status="On Target" if pct>=90 else "At Risk" if pct>=60 else "Shortage"
        fill=fill_green if pct>=90 else fill_yellow if pct>=60 else fill_red
        for c,v in enumerate([sc.nodes[did].name,info["required"],info["fulfilled"],f"{pct}%",status],1):
            ws.cell(row=r,column=c,value=v)
        data_row(ws,r,len(hff),fill); r+=1

    # Chart: Demand fulfillment
    chart1=BarChart(); chart1.type="col"
    chart1.title="Demand Fulfillment (%)"; chart1.style=10
    chart1.y_axis.title="Fulfillment (%)"; chart1.y_axis.scaling.max=110
    d_ref=Reference(ws,min_col=4,min_row=8,max_row=r-1)
    c_ref=Reference(ws,min_col=1,min_row=9,max_row=r-1)
    chart1.add_data(d_ref,titles_from_data=True); chart1.set_categories(c_ref)
    chart1.height=12; chart1.width=20; ws.add_chart(chart1,f"A{r+1}")

    # ═══ SHEET 2 — INVENTORY ══════════════════════
    ws2=wb.create_sheet("Inventory Status")
    ws2.sheet_view.showGridLines=False
    for col,w in zip("ABCDEFGHIJK",[22,13,16,9,13,13,13,11,13,16,11]):
        ws2.column_dimensions[col].width=w

    ws2.merge_cells("A1:K1"); ws2["A1"]="INVENTORY STATUS REPORT"
    ws2["A1"].font=fnt_title; ws2["A1"].alignment=aln_c
    ws2["A1"].fill=PatternFill("solid",fgColor="EFF6FF"); ws2.row_dimensions[1].height=30

    inv_df=inv.to_df(sc.nodes)
    hdr2=["Node","Type","Item","Unit","Current","Safety","Reorder","Daily Demand","Coverage Days","Available","Status"]
    for c,h in enumerate(hdr2,1): ws2.cell(row=3,column=c,value=h)
    hdr_row(ws2,3,len(hdr2))
    for alt,(idx,rd) in enumerate(inv_df.iterrows()):
        r2=alt+4; st2=rd["Status"]
        fill=fill_green if st2=="Normal" else fill_blue if st2=="Low" else fill_red
        vals=[rd["Node"],rd["Type"],rd["Item"],rd["Unit"],rd["Current Stock"],rd["Safety Stock"],
              rd["Reorder Point"],rd["Daily Demand"],rd["Coverage Days"],rd["Available (above safety)"],st2]
        for c,v in enumerate(vals,1): ws2.cell(row=r2,column=c,value=v)
        data_row(ws2,r2,len(hdr2),fill,alt%2==0)

    n_inv=len(inv_df)
    # Chart: Stock levels
    chart2=BarChart(); chart2.type="col"
    chart2.title="Current vs Safety Stock by Item"; chart2.style=10
    chart2.y_axis.title="Units"
    cur_r=Reference(ws2,min_col=5,min_row=3,max_row=3+n_inv)
    saf_r=Reference(ws2,min_col=6,min_row=3,max_row=3+n_inv)
    cat_r=Reference(ws2,min_col=3,min_row=4,max_row=3+n_inv)
    chart2.add_data(cur_r,titles_from_data=True); chart2.add_data(saf_r,titles_from_data=True)
    chart2.set_categories(cat_r); chart2.height=13; chart2.width=24
    ws2.add_chart(chart2,f"A{n_inv+6}")

    # Chart: Coverage days
    chart3=LineChart(); chart3.title="Coverage Days per Item"; chart3.style=10
    chart3.y_axis.title="Days"
    cov_r=Reference(ws2,min_col=9,min_row=3,max_row=3+n_inv)
    chart3.add_data(cov_r,titles_from_data=True); chart3.set_categories(cat_r)
    chart3.height=12; chart3.width=22; ws2.add_chart(chart3,f"A{n_inv+26}")

    # ═══ SHEET 3 — RISK ANALYSIS ══════════════════
    ws3=wb.create_sheet("Risk Analysis")
    ws3.sheet_view.showGridLines=False
    for col,w in zip("ABCDE",[30,20,18,14,30]):
        ws3.column_dimensions[col].width=w

    ws3.merge_cells("A1:E1"); ws3["A1"]="RISK ANALYSIS — CRITICAL CONNECTIONS"
    ws3["A1"].font=fnt_title; ws3["A1"].alignment=aln_c
    ws3["A1"].fill=PatternFill("solid",fgColor="FEF2F2"); ws3.row_dimensions[1].height=30

    rmap={"critical":"URGENT: Add redundant path immediately","high":"HIGH: Plan backup within 30 days",
          "medium":"MEDIUM: Monitor quarterly","low":"LOW: Well-redundant, monitor"}
    hdr3=["Connection","Avg Drop","Resilience","Severity","Recommendation"]
    for c,h in enumerate(hdr3,1): ws3.cell(row=3,column=c,value=h)
    hdr_row(ws3,3,len(hdr3))
    for alt2,rec in enumerate(ranking or [],start=4):
        sev=rec["severity"]
        fill=fill_red if sev=="critical" else fill_yellow if sev=="high" else fill_blue if sev=="medium" else fill_green
        for c,v in enumerate([rec["label"],f"{rec['avg_fulfillment_drop']}%",f"{rec['resilience_score']}%",sev.upper(),rmap.get(sev,"")],1):
            ws3.cell(row=alt2,column=c,value=v)
        data_row(ws3,alt2,len(hdr3),fill)

    if ranking:
        nr=len(ranking)
        chart4=BarChart(); chart4.type="bar"; chart4.title="Connection Criticality (Avg Drop %)"; chart4.style=10
        chart4.x_axis.title="Drop (%)"
        dr=Reference(ws3,min_col=2,min_row=3,max_row=3+nr)
        cr=Reference(ws3,min_col=1,min_row=4,max_row=3+nr)
        chart4.add_data(dr,titles_from_data=True); chart4.set_categories(cr)
        chart4.height=max(12,nr*1.8); chart4.width=22; ws3.add_chart(chart4,f"A{nr+6}")

    # ═══ SHEET 4 — FORECAST ═══════════════════════
    ws4=wb.create_sheet("Demand Forecast")
    ws4.sheet_view.showGridLines=False
    for col,w in zip("ABCDEF",[18,13,13,13,13,14]):
        ws4.column_dimensions[col].width=w

    ws4.merge_cells("A1:F1"); ws4["A1"]="DEMAND FORECASTING REPORT"
    ws4["A1"].font=fnt_title; ws4["A1"].alignment=aln_c
    ws4["A1"].fill=PatternFill("solid",fgColor="F0FDF4"); ws4.row_dimensions[1].height=30

    row4=3
    for node_id,fdf in (forecaster.forecasts or {}).items():
        ws4.merge_cells(f"A{row4}:F{row4}")
        ws4.cell(row=row4,column=1,value=f"Forecast: {fdf['node_name'].iloc[0]}").font=fnt_sub
        ws4.row_dimensions[row4].height=22; row4+=1
        if node_id in forecaster.metrics:
            m=forecaster.metrics[node_id]
            ws4.merge_cells(f"A{row4}:F{row4}")
            ws4.cell(row=row4,column=1,value=f"RMSE: {m['rmse']}  MAE: {m['mae']}  MAPE: {m['mape']}%  R2: {m['r2']}").font=fnt_sm
            row4+=1
        hdr4=["Date","Forecast","Lower CI","Upper CI"]
        for c,h in enumerate(hdr4,1): ws4.cell(row=row4,column=c,value=h)
        hdr_row(ws4,row4,4); ds=row4+1; row4+=1
        for alt3,(idx3,fr) in enumerate(fdf.iterrows()):
            for c,v in enumerate([str(fr["date"])[:10],round(fr["forecast"],1),round(fr["lower"],1),round(fr["upper"],1)],1):
                ws4.cell(row=row4,column=c,value=v)
            data_row(ws4,row4,4,fill_blue if alt3%2==0 else None); row4+=1
        nf=len(fdf); chart5=LineChart(); chart5.title=f"Forecast: {fdf['node_name'].iloc[0]}"; chart5.style=10
        chart5.y_axis.title="Units"
        fr_ref=Reference(ws4,min_col=2,min_row=ds-1,max_row=ds+nf-1)
        lo_ref=Reference(ws4,min_col=3,min_row=ds-1,max_row=ds+nf-1)
        hi_ref=Reference(ws4,min_col=4,min_row=ds-1,max_row=ds+nf-1)
        chart5.add_data(fr_ref,titles_from_data=True); chart5.add_data(lo_ref,titles_from_data=True)
        chart5.add_data(hi_ref,titles_from_data=True)
        chart5.height=12; chart5.width=22; ws4.add_chart(chart5,f"A{row4+1}"); row4+=20

    # ═══ SHEET 5 — SCORECARD ══════════════════════
    ws5=wb.create_sheet("Supplier Scorecard")
    ws5.sheet_view.showGridLines=False
    for col,w in zip("ABCDEFG",[22,13,13,13,15,14,9]):
        ws5.column_dimensions[col].width=w
    ws5.merge_cells("A1:G1"); ws5["A1"]="SUPPLIER & NODE PERFORMANCE SCORECARD"
    ws5["A1"].font=fnt_title; ws5["A1"].alignment=aln_c
    ws5["A1"].fill=PatternFill("solid",fgColor="FFFBEB"); ws5.row_dimensions[1].height=30
    hdr5=["Node","Type","Reliability","Lead Time","Quality","Cost Efficiency","Grade"]
    for c,h in enumerate(hdr5,1): ws5.cell(row=3,column=c,value=h)
    hdr_row(ws5,3,len(hdr5))
    scores_ss=st.session_state.get("scores",{})
    row5=4
    for nid,node in sc.nodes.items():
        if node.node_type not in ("plant","warehouse") or nid not in scores_ss: continue
        s=scores_ss[nid]; ov=round((s["reliability"]+s["lead_time"]+s["quality"]+s["cost_efficiency"])/4,1)
        grade="A" if ov>=90 else "B" if ov>=75 else "C" if ov>=60 else "D"
        fill=fill_green if grade=="A" else fill_blue if grade=="B" else fill_yellow if grade=="C" else fill_red
        for c,v in enumerate([node.name,node.node_type.capitalize(),s["reliability"],s["lead_time"],s["quality"],s["cost_efficiency"],grade],1):
            ws5.cell(row=row5,column=c,value=v)
        data_row(ws5,row5,len(hdr5),fill); row5+=1
    if row5>4:
        ns=row5-4; chart6=BarChart(); chart6.type="col"; chart6.title="Supplier Performance"; chart6.style=10
        chart6.y_axis.title="Score"; chart6.y_axis.scaling.max=100
        cats5=Reference(ws5,min_col=1,min_row=4,max_row=row5-1)
        for ci in range(3,7):
            r6=Reference(ws5,min_col=ci,min_row=3,max_row=row5-1)
            chart6.add_data(r6,titles_from_data=True)
        chart6.set_categories(cats5); chart6.height=13; chart6.width=24
        ws5.add_chart(chart6,f"A{row5+2}")

    # ═══ SHEET 6 — ANALYTICS (CVaR + SPC) ═════════
    ws6=wb.create_sheet("Analytics")
    ws6.sheet_view.showGridLines=False
    for col in "ABCDEFGH": ws6.column_dimensions[col].width=16
    ws6.column_dimensions["A"].width=22
    ws6.merge_cells("A1:H1"); ws6["A1"]="SUPPLY CHAIN ANALYTICS — RISK & CONTROL"
    ws6["A1"].font=fnt_title; ws6["A1"].alignment=aln_c
    ws6["A1"].fill=PatternFill("solid",fgColor="EFF6FF"); ws6.row_dimensions[1].height=30

    ws6.cell(row=3,column=1,value="CVaR (Conditional Value at Risk) Analysis").font=fnt_sub
    hdr6a=["Demand Point","Mean Demand","Std Dev","VaR 95%","CVaR 95%","Risk Level"]
    for c,h in enumerate(hdr6a,1): ws6.cell(row=4,column=c,value=h)
    hdr_row(ws6,4,len(hdr6a))
    ff2 = get_fulfillment_cached(sc); r6=5
    for did,info in ff2.items():
        node=sc.nodes[did]; base=node.capacity; std_d=base*0.15
        var95=round(base+1.645*std_d,1); cvar95=round(base+2.0*std_d,1)
        sf=max(0,cvar95-info["fulfilled"]); rl="High" if sf>base*0.3 else "Medium" if sf>base*0.1 else "Low"
        fill=fill_red if rl=="High" else fill_yellow if rl=="Medium" else fill_green
        for c,v in enumerate([node.name,round(base,1),round(std_d,1),var95,cvar95,rl],1):
            ws6.cell(row=r6,column=c,value=v)
        data_row(ws6,r6,len(hdr6a),fill); r6+=1

    chart7=BarChart(); chart7.type="col"; chart7.title="CVaR Risk by Demand Point"; chart7.style=10
    chart7.y_axis.title="Units"
    cv_ref=Reference(ws6,min_col=5,min_row=4,max_row=r6-1)
    cc_ref=Reference(ws6,min_col=1,min_row=5,max_row=r6-1)
    chart7.add_data(cv_ref,titles_from_data=True); chart7.set_categories(cc_ref)
    chart7.height=12; chart7.width=20; ws6.add_chart(chart7,f"A{r6+2}")

    # SPC Control Chart
    r6c=r6+18
    ws6.cell(row=r6c,column=1,value="SPC Control Chart — Demand Monitoring").font=fnt_sub; r6c+=1
    hdr6b=["Period","Demand","UCL","Mean","LCL","In Control"]
    for c,h in enumerate(hdr6b,1): ws6.cell(row=r6c,column=c,value=h)
    hdr_row(ws6,r6c,len(hdr6b)); ds6=r6c+1; r6c+=1
    first_dem=next((n for n in sc.nodes.values() if n.node_type=="demand"),None)
    if first_dem:
        np.random.seed(42)
        dvals=first_dem.capacity+np.random.normal(0,first_dem.capacity*0.08,20)
        md=np.mean(dvals); sd=np.std(dvals); ucl=md+3*sd; lcl=max(0,md-3*sd)
        for i,d in enumerate(dvals,1):
            ic="Yes" if lcl<=d<=ucl else "No"
            fill=fill_green if ic=="Yes" else fill_red
            for c,v in enumerate([f"P{i}",round(d,1),round(ucl,1),round(md,1),round(lcl,1),ic],1):
                ws6.cell(row=r6c,column=c,value=v)
            data_row(ws6,r6c,len(hdr6b),fill); r6c+=1
        chart8=LineChart(); chart8.title=f"SPC Control Chart — {first_dem.name}"; chart8.style=10
        chart8.y_axis.title="Demand Units"
        for ci in range(2,6):
            cr8=Reference(ws6,min_col=ci,min_row=ds6-1,max_row=r6c-1)
            chart8.add_data(cr8,titles_from_data=True)
        chart8.height=14; chart8.width=28; ws6.add_chart(chart8,f"A{r6c+2}")

    buf=io.BytesIO(); wb.save(buf); buf.seek(0); return buf.getvalue()


def load_demo_data():
    sc=SupplyChainGraph()
    sc.add_node(Node("P1","Plant Mumbai",  "plant",    500,"Mumbai, Maharashtra",  72.88,19.08))
    sc.add_node(Node("P2","Plant Chennai", "plant",    400,"Chennai, Tamil Nadu",  80.27,13.08))
    sc.add_node(Node("P3","Plant Pune",    "plant",    350,"Pune, Maharashtra",    73.86,18.52))
    sc.add_node(Node("W1","WH North",  "warehouse",400,"Delhi, Delhi",          77.21,28.61))
    sc.add_node(Node("W2","WH West",   "warehouse",350,"Ahmedabad, Gujarat",    72.57,23.02))
    sc.add_node(Node("W3","WH East",   "warehouse",300,"Kolkata, West Bengal",  88.36,22.57))
    sc.add_node(Node("W4","WH South",  "warehouse",280,"Bangalore, Karnataka",  77.59,12.97))
    sc.add_node(Node("D1","Delhi Market",    "demand",200,"Delhi, Delhi",          77.21,28.61))
    sc.add_node(Node("D2","Jaipur Market",   "demand",120,"Jaipur, Rajasthan",    75.79,26.91))
    sc.add_node(Node("D3","Surat Market",    "demand",180,"Surat, Gujarat",       72.83,21.17))
    sc.add_node(Node("D4","Bhubaneswar Mkt", "demand",100,"Bhubaneswar, Odisha", 85.82,20.30))
    sc.add_node(Node("D5","Hyderabad Market","demand",160,"Hyderabad, Telangana", 78.49,17.39))
    sc.add_node(Node("D6","Kochi Market",    "demand",140,"Kochi, Kerala",        76.27, 9.93))
    for src,tgt,cap,cost in [
        ("P1","W1",300,2.0),("P1","W2",250,1.5),("P2","W3",200,1.8),
        ("P2","W4",220,2.2),("P3","W2",180,1.2),("P3","W4",200,1.6),
        ("P1","W3",150,3.0),("P2","W1",100,3.5),("W1","D1",220,1.0),
        ("W1","D2",150,1.5),("W2","D2",130,1.2),("W2","D3",200,1.0),
        ("W3","D4",120,1.0),("W3","D5",130,1.8),("W4","D5",180,1.0),
        ("W4","D6",160,1.2),("W1","D4",80,2.5),("W2","D6",90,2.0)]:
        sc.add_edge(Edge(src,tgt,cap,cost))
    return sc

def load_demo_inventory():
    inv=InventoryManager()
    for iid,name,unit in [("SKU001","Rice","Tonnes"),("SKU002","Wheat","Tonnes"),
                           ("SKU003","Sugar","Tonnes"),("SKU004","Edible Oil","KL")]:
        inv.add_item(iid,name,unit)
    for nid,iid,cur,saf,reo,dd in [
        ("P1","SKU001",520,100,150,22),("P1","SKU002",410,80,120,16),
        ("P2","SKU002",360,70,110,14),("P2","SKU003",310,60,90,11),
        ("P3","SKU001",285,55,85,11),("P3","SKU004",205,40,65,9),
        ("W1","SKU001",185,50,75,9),("W1","SKU002",155,40,65,7),
        ("W2","SKU001",125,30,55,6),("W2","SKU003",42,25,45,5),
        ("W2","SKU004",82,20,35,4),("W3","SKU002",92,20,38,5),
        ("W3","SKU003",18,15,28,4),
        ("W4","SKU001",62,15,28,4),("W4","SKU004",52,12,22,3)]:
        inv.set_stock(nid,iid,cur,saf,reo,dd)
    return inv

def load_demo_scores():
    return {"P1":{"reliability":92,"lead_time":88,"quality":95,"cost_efficiency":78},
            "P2":{"reliability":85,"lead_time":91,"quality":89,"cost_efficiency":84},
            "P3":{"reliability":79,"lead_time":83,"quality":91,"cost_efficiency":90},
            "W1":{"reliability":94,"lead_time":90,"quality":87,"cost_efficiency":75},
            "W2":{"reliability":88,"lead_time":85,"quality":82,"cost_efficiency":88},
            "W3":{"reliability":76,"lead_time":79,"quality":84,"cost_efficiency":92},
            "W4":{"reliability":82,"lead_time":86,"quality":88,"cost_efficiency":85}}

def create_excel_template():
    buf=io.BytesIO()
    with pd.ExcelWriter(buf,engine="openpyxl") as w:
        pd.DataFrame({"id":["P1","W1","D1"],"name":["Plant 1","Warehouse 1","Market 1"],
            "node_type":["plant","warehouse","demand"],"capacity":[500,400,200],
            "location":["Mumbai, Maharashtra","Delhi, Delhi","Jaipur, Rajasthan"],
            "x_longitude":[72.88,77.21,75.79],"y_latitude":[19.08,28.61,26.91]
        }).to_excel(w,"Nodes",index=False)
        pd.DataFrame({"source":["P1","W1"],"target":["W1","D1"],"capacity":[300,200],"cost":[2.0,1.0]
        }).to_excel(w,"Connections",index=False)
        pd.DataFrame({"node_id":["P1","W1"],"item_id":["SKU001","SKU001"],
            "item_name":["Rice","Rice"],"unit":["Tonnes","Tonnes"],
            "current_stock":[500,180],"safety_stock":[100,50],"reorder_point":[150,70],"daily_demand":[20,8]
        }).to_excel(w,"Inventory",index=False)
        pd.DataFrame({"date":["2024-01-01","2024-01-02"],"node_id":["D1","D1"],"demand":[198,205]
        }).to_excel(w,"Historical_Demand",index=False)
    buf.seek(0); return buf.getvalue()


# ═══════════════════════════════════════════════════════════════
# VISUALIZATIONS
# ═══════════════════════════════════════════════════════════════
# ── Performance: graph hash for cache invalidation ──────────────────
def _graph_hash(sc):
    """Cheap fingerprint: node count + edge count + sum of capacities"""
    return hash((len(sc.nodes), len(sc.edges),
                 sum(e.capacity for e in sc.edges),
                 sum(n.capacity for n in sc.nodes.values())))

def get_fulfillment_cached(sc):
    """Return cached demand fulfillment; recompute only when graph changes"""
    h = _graph_hash(sc)
    if (st.session_state.get("_ff_hash") == h and
            st.session_state.get("_ff_cache") is not None):
        return st.session_state["_ff_cache"]
    result = sc.demand_fulfillment()
    st.session_state["_ff_cache"] = result
    st.session_state["_ff_hash"]  = h
    return result

NC={"plant":"#0F6E56","warehouse":"#185FA5","demand":"#993C1D"}
NS={"plant":"square","warehouse":"diamond","demand":"circle"}
SEV={"critical":"#E74C3C","high":"#E67E22","medium":"#3498DB","low":"#27AE60","none":"#95A5A6"}

# Green → Yellow → Red heatmap (proper, visible)
HEATMAP_COLORSCALE=[
    [0.0,  "#27AE60"],   # bright green — safe
    [0.25, "#82E0AA"],   # light green
    [0.50, "#F9E79F"],   # yellow — moderate
    [0.65, "#F0B27A"],   # orange
    [0.80, "#E74C3C"],   # red — high risk
    [1.0,  "#922B21"],   # dark red — critical
]

def _auto_layout(sc):
    layers={"plant":[],"warehouse":[],"demand":[]}
    for nid,n in sc.nodes.items(): layers.get(n.node_type,layers["warehouse"]).append(nid)
    pos={}
    for lname,nids in layers.items():
        x={"plant":0.0,"warehouse":1.0,"demand":2.0}[lname]; n=len(nids)
        for i,nid in enumerate(nids): pos[nid]=(x,(i-(n-1)/2)*1.5)
    return pos

def draw_network(sc, highlight_path=None, disrupted_edge=None, show_cap=True, in_transit=None):
    pos=_auto_layout(sc); hp=highlight_path or []; it=in_transit or []
    hp_set=set(zip(hp,hp[1:])) if len(hp)>1 else set()
    it_edges={(d.get("from_id",""),d.get("to_id","")) for d in it if d.get("status")=="In Transit"}
    large_graph = len(sc.edges) > 60  # skip arrows and cap labels for big graphs
    traces=[]
    for e in sc.edges:
        x0,y0=pos[e.source]; x1,y1=pos[e.target]
        is_dis=disrupted_edge and e.source==disrupted_edge[0] and e.target==disrupted_edge[1]
        is_hi=(e.source,e.target) in hp_set; is_it=(e.source,e.target) in it_edges
        col="#E74C3C" if is_dis else "#E67E22" if is_hi else "#2980B9" if is_it else "#BDC3C7"
        w=3 if is_hi else 2.5 if is_dis else 2.5 if is_it else 1.5
        dash="dot" if(not e.active or is_dis) else "solid"
        ht=f"<b>{sc.nodes[e.source].name} → {sc.nodes[e.target].name}</b><br>Capacity: {e.capacity} | Cost: {e.cost}"
        traces.append(go.Scatter(x=[x0,x1,None],y=[y0,y1,None],mode="lines",
            line=dict(color=col,width=w,dash=dash),hovertext=ht,hoverinfo="text",showlegend=False))
        if show_cap and not large_graph:
            mx,my=(x0+x1)/2,(y0+y1)/2
            traces.append(go.Scatter(x=[mx],y=[my],mode="text",text=[f"{int(e.capacity)}"],
                textfont=dict(size=9,color=col),showlegend=False,hoverinfo="skip"))
        if not large_graph:
            dx,dy=x1-x0,y1-y0; L=math.hypot(dx,dy)
            if L>0:
                ux,uy=dx/L,dy/L
                traces.append(go.Scatter(x=[x1-ux*0.07],y=[y1-uy*0.07],mode="markers",
                    marker=dict(symbol="arrow",size=10,color=col,angle=math.degrees(math.atan2(-dy,dx))+90),
                    showlegend=False,hoverinfo="skip"))
    for ntype in ["plant","warehouse","demand"]:
        nids=[n for n,nd in sc.nodes.items() if nd.node_type==ntype]
        if not nids: continue
        in_p=[n in hp for n in nids]
        traces.append(go.Scatter(
            x=[pos[n][0] for n in nids],y=[pos[n][1] for n in nids],
            mode="markers+text",name=ntype.capitalize()+"s",
            text=[sc.nodes[n].name for n in nids],
            textposition="middle left" if ntype=="plant" else "top center" if ntype=="warehouse" else "middle right",
            textfont=dict(size=11,color="#2C3E50"),
            hovertext=[f"<b>{sc.nodes[n].name}</b><br>{ntype}<br>Cap:{sc.nodes[n].capacity}<br>{sc.nodes[n].location}" for n in nids],
            hoverinfo="text",
            marker=dict(symbol=NS[ntype],size=[22 if ip else 15 for ip in in_p],
                color=["#E67E22" if ip else NC[ntype] for ip in in_p],
                line=dict(width=2,color="white"))))
    fig=go.Figure(data=traces)
    fig.update_layout(paper_bgcolor="#FFFFFF",plot_bgcolor="#FFFFFF",
        margin=dict(l=10,r=10,t=10,b=10),height=460,hovermode="closest",
        xaxis=dict(showgrid=False,zeroline=False,showticklabels=False),
        yaxis=dict(showgrid=False,zeroline=False,showticklabels=False),
        legend=dict(orientation="h",yanchor="bottom",y=1.01,xanchor="right",x=1,
            font=dict(size=11,color="#2C3E50"),bgcolor="rgba(0,0,0,0)"))
    return fig

def draw_gauge_charts(ff, nodes):
    demands=list(ff.items()); cols=min(3,len(demands)); rows=math.ceil(len(demands)/cols)
    specs=[[{"type":"indicator"} for _ in range(cols)] for _ in range(rows)]
    fig=make_subplots(rows=rows,cols=cols,specs=specs)
    for i,(d_id,info) in enumerate(demands):
        c=(i%cols)+1; r=(i//cols)+1; pct=info["pct"]
        col="#E74C3C" if pct<50 else "#E67E22" if pct<80 else "#27AE60"
        fig.add_trace(go.Indicator(mode="gauge+number",value=pct,
            title={"text":nodes[d_id].name,"font":{"size":10,"color":"#2C3E50"}},
            number={"suffix":"%","font":{"size":16,"color":col}},
            gauge={"axis":{"range":[0,100],"tickwidth":1},
                   "bar":{"color":col},
                   "steps":[{"range":[0,50],"color":"#FADBD8"},{"range":[50,80],"color":"#FDEBD0"},{"range":[80,100],"color":"#D5F5E3"}]}),
            row=r,col=c)
    fig.update_layout(height=210*rows,paper_bgcolor="#FFFFFF",margin=dict(l=5,r=5,t=20,b=5))
    return fig

def draw_impact_chart(impact):
    names=[v["demand_name"] for v in impact.values()]
    before=[v["before_pct"] for v in impact.values()]; after=[v["after_pct"] for v in impact.values()]
    fig=go.Figure()
    fig.add_trace(go.Bar(name="Before",x=names,y=before,marker_color="#27AE60",
        text=[f"{v}%" for v in before],textposition="outside",textfont=dict(size=10,color="#2C3E50")))
    fig.add_trace(go.Bar(name="After",x=names,y=after,marker_color="#E74C3C",
        text=[f"{v}%" for v in after],textposition="outside",textfont=dict(size=10,color="#2C3E50")))
    fig.update_layout(barmode="group",yaxis=dict(title="Fulfillment (%)",range=[0,118],gridcolor="#F1F5F9"),
        xaxis=dict(tickfont=dict(color="#64748B",size=11)),paper_bgcolor="#FFFFFF",plot_bgcolor="#FFFFFF",height=300,
        legend=dict(orientation="h",yanchor="bottom",y=1.01,xanchor="right",x=1),margin=dict(l=10,r=10,t=10,b=10))
    fig.add_hline(y=100,line_dash="dot",line_color="#AAB7B8")
    return fig

def draw_heatmap(sc, ranking):
    """Green-Red heatmap — safe=green, danger=red"""
    if not ranking: return go.Figure()
    node_ids=list(sc.nodes.keys()); node_names=[sc.nodes[n].name for n in node_ids]
    z=[]
    for nid in node_ids:
        outs=[r for r in ranking if r["source"]==nid]; ins=[r for r in ranking if r["target"]==nid]
        max_out=max((r["avg_fulfillment_drop"] for r in outs),default=0)
        max_in =max((r["avg_fulfillment_drop"] for r in ins),default=0)
        nc_out=sum(1 for r in outs if r["severity"]=="critical")*25
        nc_in =sum(1 for r in ins if r["severity"]=="critical")*25
        z.append([max_out,max_in,nc_out,nc_in])
    fig=go.Figure(go.Heatmap(
        z=z, x=["Outbound Risk","Inbound Risk","Critical Out Links","Critical In Links"],
        y=node_names,
        colorscale=HEATMAP_COLORSCALE,
        text=[[f"<b>{v:.0f}</b>" for v in row] for row in z],
        texttemplate="%{text}",textfont={"size":12,"color":"#2C3E50"},
        hoverongaps=False,showscale=True,zmin=0,zmax=100,
        colorbar=dict(title=dict(text="Risk Level",font=dict(color="#0F172A",size=11)),
                     tickfont=dict(color="#64748B",size=11),
                     tickvals=[0,25,50,75,100],ticktext=["0 — Safe","25","50 — Med","75","100 — Critical"])))
    fig.update_layout(
        height=max(350,len(node_ids)*48),
        paper_bgcolor="#FFFFFF",plot_bgcolor="#FFFFFF",
        margin=dict(l=10,r=10,t=50,b=10),
        title=dict(text="<b>Node Risk Exposure Matrix</b>  (Green = Safe → Red = Critical)",
                   font=dict(color="#0F172A",size=13),x=0.5),
        xaxis=dict(side="top",tickfont=dict(size=11,color="#2C3E50")),
        yaxis=dict(tickfont=dict(size=11,color="#2C3E50"),autorange="reversed"))
    return fig

def draw_criticality_chart(ranking):
    if not ranking: return go.Figure()
    labels=[r["label"] for r in ranking]; drops=[r["avg_fulfillment_drop"] for r in ranking]
    colors=[SEV[r["severity"]] for r in ranking]
    fig=go.Figure(go.Bar(x=drops,y=labels,orientation="h",marker_color=colors,
        text=[f"  {d}%" for d in drops],textposition="outside",textfont=dict(size=11,color="#2C3E50")))
    fig.update_layout(
        xaxis=dict(title="<b>Avg. Fulfillment Drop (%)</b>",range=[0,max(drops or [10])*1.35],gridcolor="#F1F5F9"),
        yaxis=dict(autorange="reversed",tickfont=dict(size=11,color="#2C3E50")),
        paper_bgcolor="#FFFFFF",plot_bgcolor="#FFFFFF",
        height=max(320,len(ranking)*44),margin=dict(l=10,r=80,t=10,b=10))
    return fig

def draw_resilience_gauge(score):
    col="#E74C3C" if score<40 else "#E67E22" if score<70 else "#27AE60"
    fig=go.Figure(go.Indicator(mode="gauge+number",value=score,
        number={"suffix":"%","font":{"size":36,"color":col}},
        gauge={"axis":{"range":[0,100]},"bar":{"color":col},
               "steps":[{"range":[0,40],"color":"#FADBD8"},{"range":[40,70],"color":"#FDEBD0"},{"range":[70,100],"color":"#D5F5E3"}],
               "threshold":{"line":{"color":col,"width":3},"thickness":0.75,"value":score}}))
    fig.update_layout(height=200,margin=dict(l=20,r=20,t=10,b=10),paper_bgcolor="#FFFFFF")
    return fig

def draw_forecast_chart(fc, node_id, sc_nodes):
    if node_id not in fc.history or node_id not in fc.forecasts: return go.Figure()
    hist=fc.history[node_id]; fore=fc.forecasts[node_id]
    fig=go.Figure()
    fig.add_trace(go.Scatter(x=hist["date"],y=hist["demand"],mode="lines",
        name="Historical",line=dict(color="#2980B9",width=1.5),opacity=0.8))
    fig.add_trace(go.Scatter(x=fore["date"],y=fore["forecast"],mode="lines",
        name="Forecast",line=dict(color="#E67E22",width=2.5,dash="dash")))
    fig.add_trace(go.Scatter(
        x=list(fore["date"])+list(fore["date"][::-1]),
        y=list(fore["upper"])+list(fore["lower"][::-1]),
        fill="toself",fillcolor="rgba(230,126,34,0.12)",
        line=dict(color="rgba(0,0,0,0)"),name="CI Band"))
    node_name=sc_nodes[node_id].name if node_id in sc_nodes else node_id
    fig.update_layout(
        title=dict(text=f"<b>Demand Forecast — {node_name}</b>",font=dict(color="#0F172A",size=13),x=0.5),
        xaxis=dict(gridcolor="#F1F5F9"),yaxis=dict(title="Units",gridcolor="#F1F5F9"),
        paper_bgcolor="#FFFFFF",plot_bgcolor="#FFFFFF",height=360,
        legend=dict(font=dict(color="#0F172A"),bgcolor="rgba(0,0,0,0)"),
        margin=dict(l=10,r=10,t=40,b=10))
    return fig

def draw_scorecard_radar(name, scores):
    cats=["Reliability","Lead Time","Quality","Cost Efficiency","Reliability"]
    vals=[scores["reliability"],scores["lead_time"],scores["quality"],scores["cost_efficiency"],scores["reliability"]]
    fig=go.Figure(go.Scatterpolar(r=vals,theta=cats,fill="toself",
        line=dict(color="#1B4F72",width=2),fillcolor="rgba(27,79,114,0.15)"))
    fig.update_layout(polar=dict(radialaxis=dict(visible=True,range=[0,100],gridcolor="#E5E8E8")),
        showlegend=False,height=280,paper_bgcolor="#FFFFFF",
        margin=dict(l=30,r=30,t=30,b=10),
        title=dict(text=f"<b>{name}</b>",font=dict(size=12,color="#2C3E50"),x=0.5))
    return fig

def draw_geo_map(sc, scope="india"):
    fig=go.Figure()
    for e in sc.edges:
        s=sc.nodes[e.source]; t=sc.nodes[e.target]
        if s.x and s.y and t.x and t.y:
            fig.add_trace(go.Scattergeo(lon=[s.x,t.x,None],lat=[s.y,t.y,None],
                mode="lines",line=dict(width=1.5,color="#AAB7B8"),showlegend=False,hoverinfo="skip"))
    for ntype,nlist,col,sym,sz in [
        ("plant",[n for n in sc.nodes.values() if n.node_type=="plant"],"#0F6E56","square",14),
        ("warehouse",[n for n in sc.nodes.values() if n.node_type=="warehouse"],"#185FA5","diamond",12),
        ("demand",[n for n in sc.nodes.values() if n.node_type=="demand"],"#993C1D","circle",10)]:
        if nlist:
            fig.add_trace(go.Scattergeo(lon=[n.x for n in nlist],lat=[n.y for n in nlist],
                mode="markers+text",name=ntype.capitalize()+"s",
                text=[n.name for n in nlist],textposition="top center",textfont=dict(size=9,color="#2C3E50"),
                hovertext=[f"<b>{n.name}</b><br>{ntype}<br>{n.location}" for n in nlist],hoverinfo="text",
                marker=dict(size=sz,color=col,symbol=sym,line=dict(width=1.5,color="white"))))
    sc2="asia" if scope=="india" else "world"
    geo=dict(scope=sc2,showland=True,landcolor="#F2F3F4",showocean=True,oceancolor="#EBF5FB",
             showcountries=True,countrycolor="#BDC3C7",showcoastlines=True,coastlinecolor="#85929E",bgcolor="#FDFEFE")
    if scope=="india": geo.update(center=dict(lon=80,lat=22),projection_scale=4.5)
    fig.update_layout(geo=geo,height=500,paper_bgcolor="#FFFFFF",
        margin=dict(l=0,r=0,t=0,b=0),
        legend=dict(orientation="h",yanchor="bottom",y=1.01,xanchor="right",x=1,font=dict(size=11,color="#2C3E50")))
    return fig


# ═══════════════════════════════════════════════════════════════
# FREE AI (Groq — free tier with llama3)
# ═══════════════════════════════════════════════════════════════
FREE_AI_MODELS = {
    "Groq — Llama 3.1 8B (Fast, Free)": {
        "provider":"groq","model":"llama-3.1-8b-instant","url":"https://api.groq.com/openai/v1/chat/completions"
    },
    "Groq — Llama 3.3 70B (Smart, Free)": {
        "provider":"groq","model":"llama-3.3-70b-versatile","url":"https://api.groq.com/openai/v1/chat/completions"
    },
    "Groq — Gemma 2 9B (Free)": {
        "provider":"groq","model":"gemma2-9b-it","url":"https://api.groq.com/openai/v1/chat/completions"
    },
}

GROQ_SIGNUP = "https://console.groq.com"  # Free signup, no credit card needed

SC_SYSTEM_PROMPT = """You are an expert AI assistant for SR's Supply Chain & Operations Helper platform.
You help with supply chain management, inventory optimization, demand forecasting, risk analysis, and logistics.

When users describe actions (e.g. "I added 100 units of Rice"), extract:
- Intent (update_stock, find_path, check_status, etc.)
- Entities (node name, item, quantity)
- Then include an action block if appropriate:

ACTION_JSON_START
{"action":"update_stock","node":"Plant Mumbai","item":"Rice","quantity":100,"operation":"add"}
ACTION_JSON_END

Available actions: update_stock, find_path, run_disruption, get_status, check_atp, run_forecast
Always ask for confirmation before modifying data. Be concise and professional."""

def call_free_ai(messages, api_key, model_config):
    if not api_key or not api_key.strip():
        return None, "no_key"
    try:
        r=requests.post(
            model_config["url"],
            headers={"Authorization":f"Bearer {api_key.strip()}","Content-Type":"application/json"},
            json={"model":model_config["model"],"messages":[{"role":"system","content":SC_SYSTEM_PROMPT}]+messages,"max_tokens":1200,"temperature":0.7},
            timeout=30)
        if r.status_code==200:
            data=r.json()
            if "choices" in data and data["choices"]:
                return data["choices"][0]["message"]["content"], "ok"
            return None,"empty"
        elif r.status_code==401: return None,"invalid_key"
        elif r.status_code==429: return None,"rate_limit"
        else: return None,f"error_{r.status_code}"
    except requests.exceptions.Timeout: return None,"timeout"
    except Exception as ex: return None,f"exception_{str(ex)}"

def parse_action(text):
    match=re.search(r"ACTION_JSON_START\s*(.*?)\s*ACTION_JSON_END",text,re.DOTALL)
    if match:
        try: return json.loads(match.group(1).strip())
        except: return None
    return None

def clean_response(text):
    return re.sub(r"ACTION_JSON_START.*?ACTION_JSON_END","",text,flags=re.DOTALL).strip()

def execute_action(action, sc, inv):
    a=action.get("action","")
    try:
        if a=="update_stock":
            nn=action.get("node",""); iname=action.get("item","")
            qty=float(action.get("quantity",0)); op=action.get("operation","add")
            nid=next((n.id for n in sc.nodes.values() if n.name.lower()==nn.lower()),None)
            iid=next((id for id,iv in inv.items.items() if iv["name"].lower()==iname.lower()),None)
            if not nid: return False,f"Node '{nn}' not found"
            if not iid: return False,f"Item '{iname}' not found"
            delta=qty if op=="add" else -qty if op=="remove" else None
            if delta is not None:
                inv.update_stock(nid,iid,delta)
                return True,f"{'Added' if op=='add' else 'Removed'} {qty:.0f} units of {iname} {'to' if op=='add' else 'from'} {nn}"
        elif a=="find_path":
            frm=action.get("from",""); to=action.get("to","")
            fid=next((n.id for n in sc.nodes.values() if n.name.lower()==frm.lower()),None)
            tid=next((n.id for n in sc.nodes.values() if n.name.lower()==to.lower()),None)
            if not fid or not tid: return False,"One or both nodes not found"
            res=sc.all_shortest_paths(fid,tid,k=3)
            if res:
                st.session_state.highlight_path=res[0]["path"]
                pnames=" → ".join(sc.nodes[n].name for n in res[0]["path"])
                return True,f"Path highlighted: {pnames}"
            return False,"No path found"
        elif a=="get_status":
            plants=[n for n in sc.nodes.values() if n.node_type=="plant"]
            al=inv.get_alerts(sc.nodes)
            return True,(f"Network: {len(plants)} plants, {len(sc.edges)} connections. "
                        f"Alerts: {len([x for x in al if x['level']=='critical'])} critical.")
    except Exception as ex: return False,f"Action failed: {str(ex)}"
    return False,"Unknown action"


# ═══════════════════════════════════════════════════════════════
# CSS — Clean Light Professional Theme
# ═══════════════════════════════════════════════════════════════
APP_CSS = """
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap');
html,body,[class*="css"]{font-family:'Inter',sans-serif;}

.hdr{background:linear-gradient(135deg,#1B2631 0%,#1B4F72 100%);
  padding:16px 24px;border-radius:10px;display:flex;align-items:center;gap:16px;
  margin-bottom:20px;box-shadow:0 2px 12px rgba(0,0,0,0.12);}
.hdr h1{color:#fff;font-size:19px;font-weight:700;margin:0;}
.hdr p{color:#AEB6BF;font-size:11px;margin:2px 0 0;text-transform:uppercase;letter-spacing:0.8px;}
.hdr span{color:#F39C12;font-weight:600;font-size:12px;}

.kpi{background:#fff;border:1px solid #E5E8E8;border-radius:8px;padding:14px 18px;
  border-left:4px solid #1B4F72;transition:box-shadow 0.2s;}
.kpi:hover{box-shadow:0 2px 8px rgba(0,0,0,0.08);}
.kpi-lbl{font-size:10px;color:#7F8C8D;font-weight:600;text-transform:uppercase;letter-spacing:0.8px;margin-bottom:3px;}
.kpi-val{font-size:24px;font-weight:700;color:#1B2631;line-height:1;}
.kpi-sub{font-size:10px;color:#5D6D7E;margin-top:3px;}

.sh{font-size:11px;font-weight:700;color:#2C3E50;text-transform:uppercase;
  letter-spacing:1.2px;border-bottom:2px solid #E5E8E8;padding-bottom:6px;margin-bottom:14px;}

.section-card{background:#FFFFFF;border:1px solid #E5E8E8;border-radius:10px;
  padding:16px 20px;margin-bottom:16px;box-shadow:0 1px 4px rgba(0,0,0,0.05);}

.al-r{background:#FDEDEC;border-left:4px solid #E74C3C;padding:10px 14px;border-radius:0 6px 6px 0;margin:6px 0;font-size:13px;color:#2C3E50;}
.al-a{background:#FEF9E7;border-left:4px solid #E67E22;padding:10px 14px;border-radius:0 6px 6px 0;margin:6px 0;font-size:13px;color:#2C3E50;}
.al-g{background:#EAFAF1;border-left:4px solid #27AE60;padding:10px 14px;border-radius:0 6px 6px 0;margin:6px 0;font-size:13px;color:#2C3E50;}
.al-b{background:#EBF5FB;border-left:4px solid #2980B9;padding:10px 14px;border-radius:0 6px 6px 0;margin:6px 0;font-size:13px;color:#2C3E50;}

.transport-card{background:#F8F9FA;border:1px solid #E5E8E8;border-radius:8px;padding:14px 18px;margin:8px 0;}
.transport-mode{font-size:18px;font-weight:700;color:#1B4F72;}
.transport-detail{font-size:12px;color:#5D6D7E;margin-top:4px;}

.card{background:#F8F9FA;border:1px solid #E5E8E8;border-radius:6px;padding:12px 16px;margin:5px 0;font-size:13px;color:#2C3E50;}

.b-r{background:#FADBD8;color:#C0392B;padding:2px 8px;border-radius:4px;font-size:11px;font-weight:600;}
.b-a{background:#FDEBD0;color:#E67E22;padding:2px 8px;border-radius:4px;font-size:11px;font-weight:600;}
.b-g{background:#D5F5E3;color:#1A8A4A;padding:2px 8px;border-radius:4px;font-size:11px;font-weight:600;}
.b-b{background:#D6EAF8;color:#1B4F72;padding:2px 8px;border-radius:4px;font-size:11px;font-weight:600;}

.stButton>button{background:#1B4F72!important;color:#fff!important;border:none!important;
  border-radius:6px!important;font-weight:600!important;font-size:13px!important;}
.stButton>button:hover{background:#154360!important;}

div[data-testid="stExpander"]{border:1px solid #E5E8E8!important;border-radius:8px!important;}

.stTabs [data-baseweb="tab"]{font-size:13px;font-weight:500;color:#5D6D7E;}
.stTabs [aria-selected="true"]{color:#1B4F72!important;font-weight:700!important;}

.user-bubble{background:#1B4F72;color:#fff;padding:10px 14px;border-radius:14px 14px 2px 14px;
  max-width:75%;font-size:13px;line-height:1.5;margin:6px 0;}
.ai-bubble{background:#F8F9FA;border:1px solid #E5E8E8;color:#2C3E50;padding:10px 14px;
  border-radius:14px 14px 14px 2px;max-width:80%;font-size:13px;line-height:1.6;margin:6px 0;}
.chat-wrap{display:flex;margin:6px 0;}
.chat-right{justify-content:flex-end;}.chat-left{justify-content:flex-start;}

.sb-sec{font-size:10px;font-weight:700;color:#7F8C8D;text-transform:uppercase;letter-spacing:1px;margin:12px 0 6px;}
b,.bold-label{font-weight:700!important;}
</style>
"""

# ═══════════════════════════════════════════════════════════════
# VOICE COMPONENT
# ═══════════════════════════════════════════════════════════════
def voice_component():
    html="""
    <div style="background:#F8F9FA;border:1px solid #E5E8E8;border-radius:10px;padding:16px;font-family:Inter,sans-serif">
      <div style="display:flex;gap:10px;align-items:center;margin-bottom:10px">
        <button id="mic-btn" onclick="toggleMic()" style="background:#1B4F72;color:#fff;border:none;border-radius:6px;padding:8px 16px;font-size:13px;font-weight:600;cursor:pointer">Start Listening</button>
        <button id="tts-btn" onclick="toggleTTS()" style="background:#0F6E56;color:#fff;border:none;border-radius:6px;padding:8px 14px;font-size:13px;font-weight:600;cursor:pointer">Voice On</button>
        <span id="status" style="font-size:11px;color:#5D6D7E;margin-left:8px"></span>
      </div>
      <div id="transcript" style="min-height:44px;background:#fff;border:1px solid #D5D8DC;border-radius:6px;padding:10px;font-size:13px;color:#2C3E50;margin-bottom:10px">Your speech will appear here...</div>
      <div style="display:flex;gap:8px">
        <button id="copy-btn" onclick="copyText()" style="display:none;background:#1B4F72;color:#fff;border:none;border-radius:6px;padding:6px 14px;font-size:12px;font-weight:600;cursor:pointer">Copy to Chat</button>
        <button id="clear-btn" onclick="clearText()" style="display:none;background:#E74C3C;color:#fff;border:none;border-radius:6px;padding:6px 14px;font-size:12px;font-weight:600;cursor:pointer">Clear</button>
      </div>
      <div style="margin-top:10px;font-size:11px;color:#7F8C8D"> Click <b>Start Listening</b> → speak → <b>Copy to Chat</b> → paste below</div>
    </div>
    <script>
    let isListening=false,ttsEnabled=true,recognition=null,finalText='';
    if('SpeechRecognition' in window||'webkitSpeechRecognition' in window){
      recognition=new(window.SpeechRecognition||window.webkitSpeechRecognition)();
      recognition.continuous=true;recognition.interimResults=true;recognition.lang='en-US';
      recognition.onresult=function(e){let int='';finalText='';
        for(let i=0;i<e.results.length;i++){if(e.results[i].isFinal)finalText+=e.results[i][0].transcript;else int+=e.results[i][0].transcript;}
        document.getElementById('transcript').textContent=finalText+(int?' ['+int+']':'');
        if(finalText){document.getElementById('copy-btn').style.display='inline-block';document.getElementById('clear-btn').style.display='inline-block';}};
      recognition.onerror=function(e){document.getElementById('status').textContent='Error: '+e.error;isListening=false;document.getElementById('mic-btn').textContent='Start Listening';document.getElementById('mic-btn').style.background='#1B4F72';};
      recognition.onend=function(){if(isListening)recognition.start();};
    }else{document.getElementById('mic-btn').disabled=true;document.getElementById('mic-btn').textContent='Not Supported';}
    function toggleMic(){if(!recognition)return;if(isListening){recognition.stop();isListening=false;document.getElementById('mic-btn').textContent='Start Listening';document.getElementById('mic-btn').style.background='#1B4F72';document.getElementById('status').textContent='Stopped.';}else{finalText='';document.getElementById('transcript').textContent='Listening...';recognition.start();isListening=true;document.getElementById('mic-btn').textContent='Stop Recording';document.getElementById('mic-btn').style.background='#E74C3C';document.getElementById('status').textContent='Recording...';}}
    function toggleTTS(){ttsEnabled=!ttsEnabled;document.getElementById('tts-btn').textContent=ttsEnabled?'Voice On':'Voice Off';document.getElementById('tts-btn').style.background=ttsEnabled?'#0F6E56':'#95A5A6';}
    function copyText(){if(!finalText)return;navigator.clipboard.writeText(finalText).then(()=>{document.getElementById('status').textContent='Copied to clipboard';setTimeout(()=>document.getElementById('status').textContent='',3000);});}
    function clearText(){finalText='';document.getElementById('transcript').textContent='Your speech will appear here...';document.getElementById('copy-btn').style.display='none';document.getElementById('clear-btn').style.display='none';}
    window.speakText=function(text){if(!ttsEnabled||!('speechSynthesis' in window))return;speechSynthesis.cancel();const u=new SpeechSynthesisUtterance(text.replace(/<[^>]*>/g,'').substring(0,400));u.rate=1.0;speechSynthesis.speak(u);};
    </script>"""
    components.html(html, height=210)


# ═══════════════════════════════════════════════════════════════
# STREAMLIT APP
# ═══════════════════════════════════════════════════════════════
st.set_page_config(page_title="SR — Supply Chain & Operations", page_icon=None, layout="wide", initial_sidebar_state="expanded")
st.markdown(APP_CSS, unsafe_allow_html=True)

# Session state
for k,v in [("sc",None),("inv",None),("scores",None),("dispatch_log",[]),
            ("disruption_result",None),("highlight_path",[]),("disrupted_edge",None),
            ("ranking",None),("forecaster",None),("forecast_trained",set()),
            ("chat_history",[]),("ai_key",""),("ai_model",list(FREE_AI_MODELS.keys())[0]),
            ("last_ai_text",""),("wh_forecasts",None),("plant_req",None),
            ("user_data_loaded", False), ("ls_restore_attempted", False),
            ("_ff_cache", None), ("_ff_hash", None),
            ("_pending_save", False), ("_save_payload", None)]:
    if k not in st.session_state: st.session_state[k]=v

# ── Restore from localStorage on first load ───────────────────
if not st.session_state.ls_restore_attempted:
    st.session_state.ls_restore_attempted = True
    # Try query-param based restore (set by JS on previous session)
    qp = st.query_params.get("_ls", None)
    if qp:
        try:
            import urllib.parse as _ul
            payload_str = _ul.unquote(qp)
            result = _restore_from_payload(payload_str)
            if result:
                sc_r, inv_r, scores_r, dispatch_r = result
                st.session_state.sc = sc_r
                st.session_state.inv = inv_r
                st.session_state.scores = scores_r
                st.session_state.dispatch_log = dispatch_r
                st.session_state.user_data_loaded = True
                st.session_state.forecaster = DemandForecaster()
                # Clear the query param after restore
                st.query_params.clear()
        except: pass

# ── Fall back to demo if nothing loaded ───────────────────────
if not st.session_state.user_data_loaded:
    if st.session_state.sc is None:     st.session_state.sc=load_demo_data()
    if st.session_state.inv is None:    st.session_state.inv=load_demo_inventory()
    if st.session_state.scores is None: st.session_state.scores=load_demo_scores()
if st.session_state.forecaster is None: st.session_state.forecaster=DemandForecaster()

sc=st.session_state.sc; inv=st.session_state.inv

# ── Auto-save to localStorage on every render ─────────────────
_save_json = _build_save_payload(
    sc, inv,
    st.session_state.scores,
    st.session_state.dispatch_log
)
if _save_json:
    # Write save script (runs in browser, invisible)
    _save_html = make_save_component(_save_json)
    components.html(_save_html, height=0, scrolling=False)

# ── Active dataset banner ────────────────────────────────
if st.session_state.get("user_data_loaded", False):
    _n=len(sc.nodes); _e=len(sc.edges)
    _fc_nodes=len(st.session_state.forecaster.history) if st.session_state.forecaster else 0
    _disp_c=len(st.session_state.dispatch_log or [])
    st.markdown(
        f'<div style="background:#F0FDF4;border:1px solid #BBF7D0;border-radius:8px;'
        f'padding:9px 18px;margin-bottom:14px;font-size:12px;color:#14532D;'
        f'display:flex;align-items:center;justify-content:space-between;">'
        f'<span><b>Your dataset is active</b> &nbsp;·&nbsp; {_n} nodes &nbsp;·&nbsp; {_e} connections'
        f'{f" &nbsp;·&nbsp; {_fc_nodes} demand histories" if _fc_nodes else ""}'
        f'{f" &nbsp;·&nbsp; {_disp_c} dispatches" if _disp_c else ""}</span>'
        f'<span style="color:#15803D;font-size:11px;font-weight:600;">Saved in browser</span>'
        f'</div>',
        unsafe_allow_html=True)
else:
    st.markdown(
        '<div style="background:#FFFBEB;border:1px solid #FDE68A;border-radius:8px;'
        'padding:9px 18px;margin-bottom:14px;font-size:12px;color:#92400E;">'
        '<b>Demo data active.</b> Upload your own dataset via Sidebar &rarr; Data &rarr; Import Data. '
        'Once uploaded, your data is saved and reloads automatically.</div>',
        unsafe_allow_html=True)

# ── HEADER ──────────────────────────────────────────────────────
st.markdown(f"""
<div class="app-header">
  {LOGO_SVG}
  <div>
    <h1>Supply Chain &amp; Operations Helper</h1>
    <div class="tagline">SR Platform &nbsp;&middot;&nbsp; Network Intelligence &nbsp;&middot;&nbsp; v5.0</div>
  </div>
  <div class="ver-badge">Enterprise</div>
</div>""", unsafe_allow_html=True)

# ═══════════════════════════════════════════════════════════════
# SIDEBAR
# ═══════════════════════════════════════════════════════════════
with st.sidebar:
    st.markdown(f"""
    <div style="display:flex;align-items:center;gap:10px;padding:10px 0 14px;border-bottom:1px solid #E5E8E8">
      {LOGO_SIDEBAR}
      <div>
        <div style="font-size:13px;font-weight:700;color:#1B2631">SR Network Builder</div>
        <div style="font-size:10px;color:#94A3B8;letter-spacing:0.5px;margin-top:2px">Supply Chain Intelligence</div>
      </div>
    </div>""", unsafe_allow_html=True)

    st_a,st_b,st_c=st.tabs(["Nodes","Connect","Data"])

    with st_a:
        st.markdown('<div class="sb-sec">Add New Node</div>', unsafe_allow_html=True)
        nn=st.text_input("**Node Name**",placeholder="e.g. Plant Delhi",key="nn",label_visibility="collapsed")
        c1,c2=st.columns(2)
        nt=c1.selectbox("**Type**",["plant","warehouse","demand"],key="nt")
        nc=c2.number_input("**Capacity**",min_value=1,value=200,key="nc")
        nl=st.text_input("**Location**",placeholder="City, State",key="nl",label_visibility="collapsed")
        if st.button("+ Add Node",use_container_width=True):
            if nn.strip():
                nid=nt[0].upper()+str(len([n for n in sc.nodes.values() if n.node_type==nt])+1)
                sc.add_node(Node(nid,nn.strip(),nt,nc,nl)); st.success(f"Node added."); st.rerun()
            else: st.error("Enter a name")
        st.markdown('<div class="sb-sec">Current Nodes</div>', unsafe_allow_html=True)
        for ntype,label in [("plant"," Plants"),("warehouse"," Warehouses"),("demand"," Demand Points")]:
            nlist=[n for n in sc.nodes.values() if n.node_type==ntype]
            if nlist:
                with st.expander(f"**{label}** ({len(nlist)})"):
                    for n in nlist:
                        c1,c2=st.columns([4,1])
                        c1.markdown(f"**{n.name}** <span style='color:#7F8C8D;font-size:11px'>{int(n.capacity)}</span>",unsafe_allow_html=True)
                        if c2.button("X",key=f"dn_{n.id}"):
                            del sc.nodes[n.id]; sc.edges=[e for e in sc.edges if e.source!=n.id and e.target!=n.id]; st.rerun()

    with st_b:
        st.markdown('<div class="sb-sec">Add Connection</div>', unsafe_allow_html=True)
        no={n.name:n.id for n in sc.nodes.values()}
        if len(sc.nodes)>=2:
            sl=st.selectbox("**From Node**",list(no.keys()),key="es")
            tl=st.selectbox("**To Node**",  list(no.keys()),key="et")
            c1,c2=st.columns(2)
            ecap=c1.number_input("**Capacity**",min_value=1,value=100,key="ec")
            ecos=c2.number_input("**Cost**",min_value=0.1,value=1.0,step=0.1,key="ecs")
            if st.button(" Add Connection",use_container_width=True):
                s,t=no[sl],no[tl]
                if s==t: st.error("Source and target must differ")
                else:
                    try: sc.add_edge(Edge(s,t,ecap,ecos)); st.success("Connection added."); st.rerun()
                    except ValueError as ex: st.error(str(ex))
        else: st.info("Add at least 2 nodes first")
        if sc.edges:
            st.markdown('<div class="sb-sec">Connections</div>', unsafe_allow_html=True)
            for e in sc.edges:
                c1,c2=st.columns([5,1])
                c1.markdown(f"<span style='font-size:11px'><b>{sc.nodes[e.source].name}</b>→{sc.nodes[e.target].name} <span style='color:#7F8C8D'>({int(e.capacity)})</span></span>",unsafe_allow_html=True)
                if c2.button("X",key=f"de_{e.source}_{e.target}"): sc.remove_edge(e.source,e.target); st.rerun()

    with st_c:
        st.markdown('<div class="sb-sec">Quick Start</div>', unsafe_allow_html=True)
        if st.button(" Load Demo Supply Chain",use_container_width=True):
            for k in ["sc","inv","scores","disruption_result","highlight_path","disrupted_edge","ranking","forecaster","forecast_trained","dispatch_log","wh_forecasts","plant_req"]:
                st.session_state[k]=(load_demo_data() if k=="sc" else load_demo_inventory() if k=="inv" else load_demo_scores() if k=="scores" else DemandForecaster() if k=="forecaster" else set() if k=="forecast_trained" else [] if k=="dispatch_log" else None)
            st.session_state.user_data_loaded=False
            st.session_state._ff_cache=None
            st.session_state._ff_hash=None
            st.rerun()

        st.markdown('<div class="sb-sec">Download Templates</div>', unsafe_allow_html=True)
        try:
            eb=create_excel_template()
            st.download_button(" Excel Template",eb,"sc_template.xlsx",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",use_container_width=True)
        except: pass

        st.markdown('<div class="sb-sec">Import Data</div>', unsafe_allow_html=True)
        uf=st.file_uploader("Upload Excel or CSV",type=["xlsx","csv","xls"],key="uf")
        if uf:
            fn=uf.name.lower()
            try:
                if fn.endswith((".xlsx",".xls")):
                    _prog = st.progress(0, text="Opening file...")
                    import io as _io_inner
                    raw_bytes = uf.read()  # read once into memory

                    try:
                        xls = pd.ExcelFile(_io_inner.BytesIO(raw_bytes))
                    except Exception as ex:
                        st.error(f"Cannot read file: {ex}")
                        _prog.empty()
                        st.stop()

                    sheets = xls.sheet_names
                    nsc = SupplyChainGraph()

                    # ── Nodes (vectorized) ────────────────────────────────
                    if "Nodes" in sheets:
                        _prog.progress(15, text=f"Loading nodes...")
                        try:
                            ndf = pd.read_excel(xls,"Nodes",
                                dtype={"id":str,"name":str,"node_type":str,
                                       "capacity":float,"location":str,
                                       "x_longitude":float,"y_latitude":float})
                            ndf["x_longitude"] = pd.to_numeric(ndf.get("x_longitude",0), errors="coerce").fillna(0)
                            ndf["y_latitude"]  = pd.to_numeric(ndf.get("y_latitude", 0), errors="coerce").fillna(0)
                            ndf["location"]    = ndf["location"].fillna("")
                            for row in ndf.itertuples(index=False):
                                try:
                                    nsc.add_node(Node(str(row.id),str(row.name),
                                        str(row.node_type),float(row.capacity),
                                        str(row.location),
                                        float(row.x_longitude),float(row.y_latitude)))
                                except: pass
                            _prog.progress(25, text=f"{len(nsc.nodes)} nodes loaded")
                        except Exception as ex:
                            st.warning(f"Nodes sheet error: {ex}")

                    # ── Connections (vectorized) ──────────────────────────
                    if "Connections" in sheets:
                        _prog.progress(30, text="Loading connections...")
                        try:
                            edf = pd.read_excel(xls,"Connections",
                                dtype={"source":str,"target":str,"capacity":float,"cost":float})
                            edf["cost"] = pd.to_numeric(edf.get("cost",1.0), errors="coerce").fillna(1.0)
                            for row in edf.itertuples(index=False):
                                try: nsc.add_edge(Edge(str(row.source),str(row.target),
                                        float(row.capacity),float(row.cost)))
                                except: pass
                            _prog.progress(45, text=f"{len(nsc.edges)} connections loaded")
                        except Exception as ex:
                            st.warning(f"Connections sheet error: {ex}")

                    # ── Inventory (vectorized) ────────────────────────────
                    ni = InventoryManager()
                    if "Inventory" in sheets:
                        _prog.progress(50, text="Loading inventory...")
                        try:
                            idf = pd.read_excel(xls,"Inventory",dtype=str)
                            num_cols=["current_stock","safety_stock","reorder_point","daily_demand"]
                            for c in num_cols:
                                if c in idf.columns:
                                    idf[c]=pd.to_numeric(idf[c],errors="coerce").fillna(0)
                            for row in idf.itertuples(index=False):
                                try:
                                    iid=str(row.item_id)
                                    iname=str(getattr(row,"item_name",iid))
                                    unit =str(getattr(row,"unit","units"))
                                    if iid not in ni.items: ni.add_item(iid,iname,unit)
                                    ni.set_stock(str(row.node_id),iid,
                                        float(getattr(row,"current_stock",0)),
                                        float(getattr(row,"safety_stock",0)),
                                        float(getattr(row,"reorder_point",0)),
                                        float(getattr(row,"daily_demand",1)))
                                except: pass
                            _prog.progress(62, text=f"{len(idf)} inventory records loaded")
                        except Exception as ex:
                            st.warning(f"Inventory sheet error: {ex}")

                    # ── Items sheet (optional) ────────────────────────────
                    if "Items" in sheets:
                        try:
                            iraw = pd.read_excel(xls,"Items",dtype=str)
                            for row in iraw.itertuples(index=False):
                                iid=str(getattr(row,"item_id",""))
                                if iid and iid not in ni.items:
                                    ni.add_item(iid,str(getattr(row,"item_name",iid)),str(getattr(row,"unit","units")))
                        except: pass

                    # ── Historical demand — fast groupby ─────────────────
                    fc = DemandForecaster()
                    if "Historical_Demand" in sheets:
                        _prog.progress(65, text="Loading demand history (large sheet)...")
                        try:
                            hdf = pd.read_excel(xls,"Historical_Demand",
                                                dtype={"node_id":str,"demand":float})
                            hdf["date"] = pd.to_datetime(hdf["date"], errors="coerce")
                            hdf = hdf.dropna(subset=["date","demand"])
                            hdf = hdf.sort_values("date").reset_index(drop=True)
                            total_nodes = hdf["node_id"].nunique()
                            _prog.progress(78, text=f"Indexing {total_nodes} demand nodes...")
                            # Single vectorized groupby — much faster than loop
                            for nid, grp in hdf.groupby("node_id", sort=False):
                                grp2 = grp.copy()
                                nn2  = nsc.nodes[str(nid)].name if str(nid) in nsc.nodes else str(nid)
                                grp2["node_name"] = nn2
                                grp2["node_id"]   = str(nid)
                                fc.history[str(nid)] = grp2.reset_index(drop=True)
                            _prog.progress(88, text=f"{len(hdf):,} demand records indexed")
                        except Exception as ex:
                            st.warning(f"Historical demand sheet error: {ex}")

                    # ── Commit to session state ───────────────────────────
                    _prog.progress(92, text="Saving to session...")
                    st.session_state.sc              = nsc
                    st.session_state.inv             = ni
                    st.session_state.forecaster      = fc
                    st.session_state.scores          = load_demo_scores()
                    st.session_state.user_data_loaded= True
                    st.session_state.ranking         = None
                    st.session_state._ff_cache       = None
                    st.session_state._ff_hash        = None
                    st.session_state.dispatch_log    = []
                    st.session_state.highlight_path  = []
                    st.session_state.disrupted_edge  = None
                    st.session_state.disruption_result = None

                    _prog.progress(100, text="Done.")
                    _prog.empty()

                    nn = len(nsc.nodes); ne = len(nsc.edges)
                    ni_r = len(idf) if "Inventory" in sheets else 0
                    nh   = sum(len(v) for v in fc.history.values())
                    st.success(
                        f"Import complete — {nn} nodes, {ne} connections, "
                        f"{ni_r} inventory records, {nh:,} demand rows. "
                        f"Data is now active and will be remembered.")
                    st.rerun()
                else:
                    df=pd.read_csv(uf)
                    if "node_type" in df.columns:
                        nsc=SupplyChainGraph()
                        for _,r in df.iterrows():
                            nsc.add_node(Node(str(r["id"]),str(r["name"]),str(r["node_type"]),float(r["capacity"]),str(r.get("location",""))))
                        st.session_state.sc=nsc; st.success("Nodes imported."); st.rerun()
                    elif "source" in df.columns:
                        for _,r in df.iterrows(): sc.add_edge(Edge(str(r["source"]),str(r["target"]),float(r["capacity"]),float(r.get("cost",1.0))))
                        st.success("Edges imported."); st.rerun()
            except Exception as ex: st.error(f"Import failed: {ex}")

        st.markdown('<div class="sb-sec">Export</div>', unsafe_allow_html=True)
        ndf=pd.DataFrame([vars(n) for n in sc.nodes.values()]); edf=pd.DataFrame([vars(e) for e in sc.edges])
        c1,c2=st.columns(2)
        if not ndf.empty: c1.download_button("Nodes",ndf.to_csv(index=False),"nodes.csv",use_container_width=True)
        if not edf.empty: c2.download_button("Edges",edf.to_csv(index=False),"edges.csv",use_container_width=True)

        st.markdown('<div class="sb-sec">Session</div>', unsafe_allow_html=True)
        _is_user_data = st.session_state.get("user_data_loaded", False)
        if _is_user_data:
            st.markdown(
                '<div style="background:#F0FDF4;border:1px solid #BBF7D0;border-radius:6px;'
                'padding:8px 12px;font-size:11px;color:#14532D;font-weight:600;">'
                'Your data is saved in this browser. '
                'It will reload automatically next visit.</div>',
                unsafe_allow_html=True)
        else:
            st.markdown(
                '<div style="background:#FFFBEB;border:1px solid #FDE68A;border-radius:6px;'
                'padding:8px 12px;font-size:11px;color:#92400E;">'
                'Demo data active. Upload your dataset above.</div>',
                unsafe_allow_html=True)
        if st.button("Clear saved data & reset", use_container_width=True, key="clear_ls"):
            for k in ["sc","inv","scores","ranking","_ff_cache","_ff_hash",
                      "disruption_result","highlight_path","disrupted_edge",
                      "wh_forecasts","plant_req","dispatch_log","forecast_trained"]:
                st.session_state[k] = (None if k not in ("dispatch_log","highlight_path","forecast_trained")
                                       else [] if k in ("dispatch_log","highlight_path") else set())
            st.session_state.user_data_loaded = False
            components.html(make_clear_component(), height=0, scrolling=False)
            st.rerun()

        st.markdown('<div class="sb-sec"> AI Assistant Setup (Free)</div>', unsafe_allow_html=True)
        sel_model=st.selectbox("**Select Free AI Model**",list(FREE_AI_MODELS.keys()),key="model_sel")
        st.session_state.ai_model=sel_model
        st.caption(f"**Free signup:** [console.groq.com]({GROQ_SIGNUP}) — No credit card needed!")
        key_in=st.text_input("**Groq API Key**",value=st.session_state.ai_key,type="password",placeholder="gsk_...",key="key_in")
        if key_in!=st.session_state.ai_key: st.session_state.ai_key=key_in
        if st.session_state.ai_key: st.success("OK API key set")
        else: st.info("Get free key at console.groq.com")


# ═══════════════════════════════════════════════════════════════
# MAIN TABS
# ═══════════════════════════════════════════════════════════════
T=st.tabs([" Network Map"," Inventory"," Disruption"," Risk Heatmap",
           " Demand Forecast"," Dispatch Log"," ATP & Scorecard",
           " Geographic View"," Reports"," AI Assistant"])
t1,t2,t3,t4,t5,t6,t7,t8,t9,t10=T

# ─────────────────────────────────────────────────────────────
# TAB 1 — NETWORK MAP
# ─────────────────────────────────────────────────────────────
with t1:
    if not sc.nodes:
        st.info("Add nodes in the sidebar or load the demo supply chain.")
    else:
        plants=[n for n in sc.nodes.values() if n.node_type=="plant"]
        whs   =[n for n in sc.nodes.values() if n.node_type=="warehouse"]
        dems  =[n for n in sc.nodes.values() if n.node_type=="demand"]
        active_dis=[d for d in st.session_state.dispatch_log if d.get("status")=="In Transit"]

        c1,c2,c3,c4,c5=st.columns(5)
        def kpi(col,label,val,sub,color="#1B4F72"):
            col.markdown(f'<div class="kpi" style="border-left-color:{color}"><div class="kpi-lbl">{label}</div><div class="kpi-val">{val}</div><div class="kpi-sub">{sub}</div></div>',unsafe_allow_html=True)
        kpi(c1," Plants",len(plants),f"Cap: {sum(n.capacity for n in plants):,.0f}")
        kpi(c2," Warehouses",len(whs),f"Cap: {sum(n.capacity for n in whs):,.0f}","#185FA5")
        kpi(c3," Demand Pts",len(dems),f"Demand: {sum(n.capacity for n in dems):,.0f}","#993C1D")
        kpi(c4," Connections",len(sc.edges),f"Active dispatches: {len(active_dis)}","#1A5276")
        cov=round(sum(n.capacity for n in plants)/max(sum(n.capacity for n in dems),1)*100,1)
        cc="#27AE60" if cov>=100 else "#E67E22" if cov>=70 else "#E74C3C"
        kpi(c5," Coverage",f"{min(cov,999):.0f}%","vs total demand",cc)

        al=inv.get_alerts(sc.nodes); ca=[a for a in al if a["level"]=="critical"]
        if ca:
            items=" | ".join(f"<b>{a['node_name']}/{a['item_name']}</b>" for a in ca[:3])
            st.markdown(f'<div class="al-r" style="margin-top:12px"> Critical stock: {items}</div>',unsafe_allow_html=True)

        st.markdown("<br>",unsafe_allow_html=True)
        _,co=st.columns([5,1]); show_cap=co.checkbox("**Edge labels**",value=True,key="ec_cb")
        fig=draw_network(sc,st.session_state.highlight_path,st.session_state.disrupted_edge,show_cap,active_dis)
        st.plotly_chart(fig,use_container_width=True)
        st.markdown('<div style="display:flex;gap:20px;font-size:11px;color:#5D6D7E">'+
            '<span>■ <b>Plant</b></span><span>◆ <b>Warehouse</b></span><span>● <b>Demand</b></span>'+
            '<span style="color:#E67E22">■ Highlighted path</span>'+
            '<span style="color:#2980B9">■ In Transit</span>'+
            '<span style="color:#E74C3C">■ Disrupted</span></div>',unsafe_allow_html=True)

        st.markdown("<br>",unsafe_allow_html=True)
        st.markdown('<div class="sh">Shortest Path Finder with Transport Mode</div>',unsafe_allow_html=True)
        nlab={n.name:n.id for n in sc.nodes.values()}
        c1,c2,c3,c4=st.columns([2,2,1,1])
        sp1=c1.selectbox("**From**",list(nlab.keys()),key="sp1")
        sp2=c2.selectbox("**To**",list(nlab.keys()),key="sp2",index=min(2,len(nlab)-1))
        load_units=c3.number_input("**Load (units)**",min_value=1,value=100,key="load_u")
        urgent=c4.checkbox("**Urgent?**",key="urgent_cb")

        if st.button(" Find Path & Transport Mode",use_container_width=False,key="find_path_btn"):
            fid=nlab[sp1]; tid=nlab[sp2]
            # Check connection exists
            edge_exists=any((e.source==fid and e.target==tid) or
                            nx.has_path(sc.to_nx(),fid,tid) for e in sc.edges
                            if sc.to_nx().number_of_nodes()>0)
            res=sc.all_shortest_paths(fid,tid,k=3)
            if res:
                st.session_state.highlight_path=res[0]["path"]
                # Get transport recommendation for each hop
                path_ids=res[0]["path"]
                transport_recs=[]
                total_dist=0; total_cost_est=0
                for i in range(len(path_ids)-1):
                    sn=sc.nodes[path_ids[i]]; tn=sc.nodes[path_ids[i+1]]
                    tr=recommend_transport(sn,tn,load_units,urgent)
                    transport_recs.append(tr)
                    if tr["dist_km"]: total_dist+=tr["dist_km"]
                    if tr["cost_estimate"]: total_cost_est+=tr["cost_estimate"]

                st.session_state["last_path_result"]=(res,transport_recs,total_dist,total_cost_est)
                st.rerun()
            else:
                st.error(" **No path found** between these nodes. Check that connections exist.")
                st.session_state.highlight_path=[]

        if "last_path_result" in st.session_state and st.session_state.highlight_path:
            res,transport_recs,total_dist,total_cost_est=st.session_state["last_path_result"]
            path=st.session_state.highlight_path
            pnames=[sc.nodes[n].name for n in path]
            st.markdown(f'<div class="al-g">OK **Path:** {" → ".join(pnames)}</div>',unsafe_allow_html=True)

            st.markdown('<div class="sh" style="margin-top:12px">Transport Recommendations</div>',unsafe_allow_html=True)
            if transport_recs:
                for i,tr in enumerate(transport_recs):
                    src_name=sc.nodes[path[i]].name; tgt_name=sc.nodes[path[i+1]].name
                    dist_str=f"**{tr['dist_km']} km**" if tr['dist_km'] else "distance N/A"
                    cost_str=f"₹{tr['cost_estimate']:,.0f}" if tr['cost_estimate'] else "N/A"
                    time_str=f"{tr['hours']:.1f} hrs ({tr['days']:.1f} days)" if tr['hours'] else "N/A"
                    st.markdown(f"""
                    <div class="transport-card">
                      <div style="display:flex;justify-content:space-between;align-items:center">
                        <div><b>{src_name} → {tgt_name}</b></div>
                        <div class="transport-mode">{tr['icon']} {tr['label']}</div>
                      </div>
                      <div class="transport-detail">
                         Distance: {dist_str} &nbsp;|&nbsp;
                         Est. Cost: <b>{cost_str}</b> &nbsp;|&nbsp;
                        ⏱ Est. Time: <b>{time_str}</b>
                      </div>
                      <div style="font-size:11px;color:#5D6D7E;margin-top:4px"> {tr['reason']}</div>
                    </div>""",unsafe_allow_html=True)
                if total_dist>0 or total_cost_est>0:
                    st.markdown(f'<div class="al-b"> <b>Total Route:</b> {total_dist:.0f} km total distance | Est. Cost: ₹{total_cost_est:,.0f}</div>',unsafe_allow_html=True)

            if st.button("Clear path",key="clear_path"): st.session_state.highlight_path=[]; del st.session_state["last_path_result"]; st.rerun()

        st.markdown("<br>",unsafe_allow_html=True)
        st.markdown('<div class="sh">Demand Fulfillment</div>',unsafe_allow_html=True)
        if sc.edges:
            _ff_col1, _ff_col2 = st.columns([3,1])
            _ff_col1.caption(f"Computes max-flow from all plants to all {len([n for n in sc.nodes.values() if n.node_type=='demand'])} demand points.")
            if _ff_col2.button("Compute Fulfillment", key="ff_btn", use_container_width=True):
                with st.spinner("Computing..."):
                    ff=sc.demand_fulfillment()
                    st.session_state._ff_cache=ff
                    st.session_state._ff_hash=_graph_hash(sc)
            ff = st.session_state.get("_ff_cache")
            if ff:
                st.plotly_chart(draw_gauge_charts(ff,sc.nodes),use_container_width=True)
            else:
                st.info("Click **Compute Fulfillment** to analyse supply vs demand across your network.")
            if ff:
              with st.expander("View Detailed Table"):
                rows=[{"Demand Point":sc.nodes[d].name,"Required":info["required"],"Fulfilled":info["fulfilled"],
                       "Fulfillment %":info["pct"],"Sources":", ".join(sc.nodes[p].name for p in info["reachable_from"]) or "None"}
                      for d,info in ff.items()]
                df_ff=pd.DataFrame(rows)
                def cpct(v): return "background-color:#DCFCE7" if v>=90 else "background-color:#FEF3C7" if v>=50 else "background-color:#FEE2E2"
                st.dataframe(df_ff.style.map(cpct,subset=["Fulfillment %"]),use_container_width=True,hide_index=True)
        if not sc.edges: st.info("Add connections to compute demand fulfillment.")


# ─────────────────────────────────────────────────────────────
# TAB 2 — INVENTORY
# ─────────────────────────────────────────────────────────────
with t2:
    st.markdown('<div class="sh">Inventory Overview</div>',unsafe_allow_html=True)
    al=inv.get_alerts(sc.nodes)
    if al:
        for a in [x for x in al if x["level"]=="critical"]:
            st.markdown(f'<div class="al-r"><span class="b-r">CRITICAL</span> <b>{a["node_name"]}</b> — {a["item_name"]}: {a["current"]:.0f} / Safety: {a["safety"]:.0f} · <b>{a["coverage"]} days</b> left</div>',unsafe_allow_html=True)
        for a in [x for x in al if x["level"]=="warning"]:
            st.markdown(f'<div class="al-a"><span class="b-a">LOW</span> <b>{a["node_name"]}</b> — {a["item_name"]}: {a["current"]:.0f} · <b>{a["coverage"]} days</b></div>',unsafe_allow_html=True)
    else:
        st.markdown('<div class="al-g">All stock levels are within normal range.</div>',unsafe_allow_html=True)

    st.markdown("<br>",unsafe_allow_html=True)
    st.markdown('<div class="sh">Stock Levels by Node</div>',unsafe_allow_html=True)
    inv_df=inv.to_df(sc.nodes)
    if not inv_df.empty:
        STC={"Normal":"#27AE60","Low":"#E67E22","Critical":"#E74C3C"}
        def cs(v): return f"color:{STC.get(v,'#2C3E50')};font-weight:700"
        def cc2(v): return "background-color:#FADBD8" if v<=3 else "background-color:#FDEBD0" if v<=7 else "background-color:#D5F5E3"
        st.dataframe(inv_df.style.map(cs,subset=["Status"]).map(cc2,subset=["Coverage Days"]),use_container_width=True,hide_index=True)

    # Inventory overview charts
    st.markdown("<br>",unsafe_allow_html=True)
    if not inv_df.empty:
        st.markdown('<div class="sh">Inventory Analytics</div>',unsafe_allow_html=True)
        ch_c1, ch_c2 = st.columns(2)
        with ch_c1:
            status_counts = inv_df["Status"].value_counts()
            color_map = {"Normal":"#16A34A","Low":"#D97706","Critical":"#DC2626"}
            pie_colors = [color_map.get(s,"#94A3B8") for s in status_counts.index]
            fig_pie = go.Figure(go.Pie(
                labels=status_counts.index, values=status_counts.values,
                marker=dict(colors=pie_colors, line=dict(color="#FFFFFF", width=2)),
                textfont=dict(size=12), hole=0.45))
            fig_pie.update_layout(
                title=dict(text="Stock Status Distribution", font=dict(size=12,color="#0F172A"), x=0.5),
                paper_bgcolor="#FFFFFF", height=260, margin=dict(l=10,r=10,t=36,b=10),
                legend=dict(font=dict(size=11,color="#64748B"), orientation="h", y=-0.1))
            st.plotly_chart(fig_pie, use_container_width=True)
        with ch_c2:
            if "Coverage Days" in inv_df.columns:
                cov_df = inv_df.sort_values("Coverage Days")
                cov_labels = (cov_df["Node"] + " / " + cov_df["Item"]).tolist()
                cov_vals   = cov_df["Coverage Days"].tolist()
                cov_colors = ["#DC2626" if v<=3 else "#D97706" if v<=7 else "#16A34A" for v in cov_vals]
                fig_cov = go.Figure(go.Bar(
                    x=cov_vals, y=cov_labels, orientation="h",
                    marker_color=cov_colors, text=[f"{v:.0f}d" for v in cov_vals],
                    textposition="outside", textfont=dict(size=10,color="#64748B")))
                fig_cov.update_layout(
                    title=dict(text="Coverage Days by Item", font=dict(size=12,color="#0F172A"), x=0.5),
                    paper_bgcolor="#FFFFFF", plot_bgcolor="#FFFFFF",
                    xaxis=dict(title="Days", gridcolor="#F1F5F9"),
                    yaxis=dict(tickfont=dict(size=9)),
                    height=260, margin=dict(l=10,r=50,t=36,b=10))
                st.plotly_chart(fig_cov, use_container_width=True)

    st.markdown("<br>",unsafe_allow_html=True)
    st.markdown('<div class="sh">Stock Chart</div>',unsafe_allow_html=True)
    nws=[nid for nid in sc.nodes if nid in inv.stock and inv.stock[nid]]
    if nws:
        sel_n=st.selectbox("**Select Node**",[sc.nodes[n].name for n in nws],key="inv_ns")
        sel_id=[n for n in nws if sc.nodes[n].name==sel_n]
        if sel_id:
            nid=sel_id[0]; items_data=inv.stock[nid]
            inames=[inv.items.get(iid,{}).get("name",iid) for iid in items_data]
            curs=[s["current"] for s in items_data.values()]
            safs=[s["safety"] for s in items_data.values()]
            reos=[s["reorder"] for s in items_data.values()]
            cbars=["#E74C3C" if c<=s else "#E67E22" if c<=r else "#27AE60" for c,s,r in zip(curs,safs,reos)]
            fig_bar=go.Figure()
            fig_bar.add_trace(go.Bar(x=inames,y=curs,marker_color=cbars,name="Current",text=[f"{c:.0f}" for c in curs],textposition="outside"))
            fig_bar.add_trace(go.Scatter(x=inames,y=safs,mode="markers+lines",marker=dict(symbol="line-ew",size=18,color="#E74C3C",line=dict(width=2,color="#E74C3C")),line=dict(dash="dot",color="#E74C3C"),name="Safety Stock"))
            fig_bar.add_trace(go.Scatter(x=inames,y=reos,mode="markers+lines",marker=dict(symbol="line-ew",size=18,color="#E67E22",line=dict(width=2,color="#E67E22")),line=dict(dash="dash",color="#E67E22"),name="Reorder Point"))
            fig_bar.update_layout(height=280,paper_bgcolor="#FFFFFF",plot_bgcolor="#FFFFFF",
                legend=dict(font=dict(color="#0F172A")),margin=dict(l=10,r=10,t=10,b=10))
            st.plotly_chart(fig_bar,use_container_width=True)

    st.markdown("<br>",unsafe_allow_html=True)
    st.markdown('<div class="sh">Live Stock Update</div>',unsafe_allow_html=True)
    with st.form("stock_upd"):
        c1,c2,c3,c4=st.columns([2,2,1,2])
        un=[sc.nodes[n].name for n in sc.nodes if n in inv.stock]
        upd_n=c1.selectbox("**Node**",un,key="upd_n") if un else c1.text_input("Node")
        uid=[n for n in sc.nodes if sc.nodes[n].name==upd_n and n in inv.stock]
        iopts={inv.items.get(iid,{}).get("name",iid):iid for iid in (inv.stock.get(uid[0],{}) if uid else {})}
        upd_i=c2.selectbox("**Item**",list(iopts.keys()),key="upd_i") if iopts else c2.text_input("Item ID")
        upd_q=c3.number_input("**Qty (±)**",value=0.0,step=1.0,key="upd_q")
        upd_nt=c4.text_input("**Note**",placeholder="Optional",key="upd_nt")
        if st.form_submit_button("Update Stock",use_container_width=True) and uid:
            iid=iopts.get(upd_i,upd_i)
            if inv.update_stock(uid[0],iid,upd_q):
                st.session_state.dispatch_log.append({"timestamp":datetime.now().strftime("%Y-%m-%d %H:%M"),
                    "from_node":upd_n,"to_node":"Manual Update","from_id":uid[0],"to_id":"",
                    "item":inv.items.get(iid,{}).get("name",iid),"item_id":iid,
                    "quantity":upd_q,"status":"Delivered","notes":upd_nt})
                st.success(f"Stock updated: {upd_q:+.0f} units"); st.rerun()
            else: st.error("Node/item not found")


# ─────────────────────────────────────────────────────────────
# TAB 3 — DISRUPTION SIMULATOR
# ─────────────────────────────────────────────────────────────
with t3:
    st.markdown('<div class="sh">Disruption Scenario Analysis</div>',unsafe_allow_html=True)
    st.caption("Simulate a connection failure — see **resilience score**, **alternate routes**, and **safety stock** coverage.")
    if not sc.edges: st.info("Add connections to run disruption analysis.")
    else:
        eopts={f"{sc.nodes[e.source].name} → {sc.nodes[e.target].name} (cap:{int(e.capacity)})":(e.source,e.target) for e in sc.edges}
        c1,c2=st.columns([4,1])
        chosen=c1.selectbox("**Connection to disrupt**",list(eopts.keys()),label_visibility="collapsed",key="dis_sel")
        c2.markdown("<br>",unsafe_allow_html=True)
        if c2.button("Analyse",use_container_width=True,type="primary",key="dis_btn"):
            src,tgt=eopts[chosen]
            with st.spinner("Running disruption analysis..."): result=sc.simulate_disruption(src,tgt)
            st.session_state.disruption_result=result; st.session_state.disrupted_edge=(src,tgt); st.rerun()

        if st.session_state.disruption_result:
            result=st.session_state.disruption_result; st.markdown("<br>",unsafe_allow_html=True)
            cg,cs2=st.columns([1,2])
            with cg:
                st.markdown('<div class="sh">Resilience Score</div>',unsafe_allow_html=True)
                st.plotly_chart(draw_resilience_gauge(result["resilience_score"]),use_container_width=True)
            with cs2:
                st.markdown(f'<div class="sh">Impact — {result["removed_edge"]}</div>',unsafe_allow_html=True)
                sv=result["resilience_score"]
                cls="al-g" if sv>=70 else "al-a" if sv>=40 else "al-r"
                txt="**Low Risk** — Chain retains continuity." if sv>=70 else "**Moderate Risk** — Partial disruption." if sv>=40 else "**Critical Risk** — Major disruption!"
                st.markdown(f'<div class="{cls}"><b>{sv}%</b> — {txt}</div>',unsafe_allow_html=True)
                st.markdown("<br>",unsafe_allow_html=True)
                for d_id,imp in result["impact"].items():
                    if imp["drop_pct"]>0:
                        sv2=SEV[imp["severity"]]
                        st.markdown(f"""<div style="display:flex;justify-content:space-between;align-items:center;
                          padding:7px 12px;background:#F8F9FA;border-radius:5px;margin:3px 0;border-left:3px solid {sv2}">
                          <b style="font-size:13px">{imp['demand_name']}</b>
                          <span style="font-size:12px;color:#5D6D7E">{imp['before_pct']}%→{imp['after_pct']}% | <b>−{imp['drop_pct']}%</b> | −{imp['lost_units']} units</span>
                          <span style="color:{sv2};font-size:11px;font-weight:700">{imp['severity'].upper()}</span></div>""",unsafe_allow_html=True)

            st.markdown("<br>",unsafe_allow_html=True)
            st.markdown('<div class="sh">Fulfillment Before vs After</div>',unsafe_allow_html=True)
            st.plotly_chart(draw_impact_chart(result["impact"]),use_container_width=True)

            st.markdown("<br>",unsafe_allow_html=True)
            st.markdown('<div class="sh">Alternate Network Routes</div>',unsafe_allow_html=True)
            if result["alt_paths"]:
                for d_id,paths in result["alt_paths"].items():
                    if paths:
                        st.markdown(f"**{sc.nodes[d_id].name}** — {len(paths)} alternate route(s):")
                        for i,p in enumerate(paths,1):
                            st.markdown(f'<div class="card"><b>Route {i}</b> | Cost: {p["cost"]} | {" → ".join(p["path"])}</div>',unsafe_allow_html=True)
            else:
                st.markdown('<div class="al-r"><b>No alternate routes found</b> — This is a single point of failure!</div>',unsafe_allow_html=True)

            st.markdown("<br>",unsafe_allow_html=True)
            st.markdown('<div class="sh">Safety Stock Alternatives</div>',unsafe_allow_html=True)
            de=st.session_state.disrupted_edge
            for d_id,imp in result["impact"].items():
                if imp["drop_pct"]<=0: continue
                sf=imp["lost_units"]
                st.markdown(f"**{imp['demand_name']}** — shortfall: **{sf:.0f} units**")
                rows_s=[]
                for iid in list(inv.items.keys()):
                    alts=inv.find_alternatives(d_id,iid,sf,sc,de)
                    for alt in alts:
                        rows_s.append({"Source":alt["node_name"],"Item":inv.items[iid]["name"],
                            "Available":f"{alt['available']:.0f}","Covers":("OK Yes" if alt["can_cover"] else f"Partial {alt['coverage_pct']}%"),
                            "Route":" → ".join(alt["path"]),"Cost":alt["route_cost"],"Coverage Days":alt.get("coverage_days","—")})
                if rows_s:
                    df_s=pd.DataFrame(rows_s)
                    def cov_s(v): return "background-color:#D5F5E3;font-weight:700" if "Yes" in str(v) else "background-color:#FDEBD0"
                    st.dataframe(df_s.style.map(cov_s,subset=["Covers"]),use_container_width=True,hide_index=True)
                else:
                    st.markdown('<div class="al-a">No above-safety stock available.</div>',unsafe_allow_html=True)

            st.markdown("<br>",unsafe_allow_html=True)
            st.markdown('<div class="sh">Network View — Disruption Highlighted</div>',unsafe_allow_html=True)
            alt_hi=[]
            if result["alt_paths"]:
                first=[v for v in result["alt_paths"].values() if v]
                if first and first[0]: alt_hi=first[0][0].get("path_ids",[])
            st.plotly_chart(draw_network(sc,alt_hi,st.session_state.disrupted_edge,True,
                [d for d in st.session_state.dispatch_log if d.get("status")=="In Transit"]),use_container_width=True)
            if st.button(" Reset Simulation",key="res_dis"):
                st.session_state.disruption_result=None; st.session_state.disrupted_edge=None
                st.session_state.highlight_path=[]; st.rerun()


# ─────────────────────────────────────────────────────────────
# TAB 4 — RISK HEATMAP (Green → Red, proper colors)
# ─────────────────────────────────────────────────────────────
with t4:
    st.markdown('<div class="sh">Network Risk Heatmap</div>',unsafe_allow_html=True)
    st.caption("**Green = Safe | Yellow = Moderate | Red = High Risk | Dark Red = Critical**")
    if not sc.edges: st.info("Add connections first.")
    else:
        c1,c2=st.columns([3,1])
        if c2.button(" Run Stress Test",use_container_width=True,key="hm_btn"):
            with st.spinner("Testing all connections..."): st.session_state.ranking=sc.rank_critical_edges(); st.rerun()
        if st.session_state.ranking is None:
            st.info("Click **Run Stress Test** to populate the risk heatmap.")
        if st.session_state.ranking:
            ranking=st.session_state.ranking
            st.plotly_chart(draw_heatmap(sc,ranking),use_container_width=True)

            st.markdown("<br>",unsafe_allow_html=True)
            st.markdown('<div class="sh">Node Risk Summary</div>',unsafe_allow_html=True)
            rows=[]
            for nid,node in sc.nodes.items():
                outs=[r for r in ranking if r["source"]==nid]; ins=[r for r in ranking if r["target"]==nid]
                md=max((r["avg_fulfillment_drop"] for r in outs+ins),default=0)
                nc=sum(1 for r in outs+ins if r["severity"]=="critical")
                rl="Critical" if md>=50 else "High" if md>=25 else "Medium" if md>=5 else "Low"
                rows.append({"Node":node.name,"Type":node.node_type.capitalize(),
                             "Max Drop If Link Fails":f"**{md:.0f}%**","Critical Links":nc,"Risk Level":rl})
            df_r=pd.DataFrame(rows).sort_values("Max Drop If Link Fails",ascending=False)
            def rl_col(v):
                c={"Critical":"#FADBD8","High":"#FDEBD0","Medium":"#D6EAF8","Low":"#D5F5E3"}
                return f"background-color:{c.get(v,'#FFFFFF')}"
            st.dataframe(df_r.style.map(rl_col,subset=["Risk Level"]),use_container_width=True,hide_index=True)

            st.markdown("<br>",unsafe_allow_html=True)
            # Node risk summary chart
            if rows:
                fig_node_risk = go.Figure()
                risk_colors = {"Critical":"#DC2626","High":"#D97706","Medium":"#2563EB","Low":"#16A34A"}
                node_names_r  = [r["Node"]  for r in rows]
                node_drops_r  = [float(r["Max Drop"].replace("%","")) for r in rows]
                node_colors_r = [risk_colors.get(r["Risk"],"#94A3B8") for r in rows]
                fig_node_risk.add_trace(go.Bar(
                    y=node_names_r, x=node_drops_r, orientation="h",
                    marker_color=node_colors_r,
                    text=[f"{v:.0f}%" for v in node_drops_r],
                    textposition="outside",textfont=dict(size=11,color="#0F172A")))
                fig_node_risk.update_layout(
                    title=dict(text="Max Fulfillment Drop by Node",font=dict(size=12,color="#0F172A"),x=0.5),
                    paper_bgcolor="#FFFFFF",plot_bgcolor="#FFFFFF",height=max(260,len(rows)*38),
                    xaxis=dict(title="Drop (%)",range=[0,max(node_drops_r or [10])*1.3],gridcolor="#F1F5F9"),
                    yaxis=dict(autorange="reversed"),
                    margin=dict(l=10,r=60,t=36,b=10))
                st.plotly_chart(fig_node_risk, use_container_width=True)

            st.markdown("<br>",unsafe_allow_html=True)
            st.markdown('<div class="sh">Criticality Ranking</div>',unsafe_allow_html=True)
            st.plotly_chart(draw_criticality_chart(ranking),use_container_width=True)

            st.markdown('<div class="sh">Risk Mitigation Recommendations</div>',unsafe_allow_html=True)
            for sev,cls,label in [("critical","al-r"," Immediate Action"),("high","al-a"," High Priority"),("low","al-g","OK Well Redundant")]:
                items=[r for r in ranking if r["severity"]==sev]
                if items:
                    links="".join(f"<li><b>{r['label']}</b></li>" for r in items)
                    st.markdown(f'<div class="{cls}"><b>{label} ({len(items)} links):</b><ul style="margin:4px 0 0 16px">{links}</ul></div>',unsafe_allow_html=True)


# ─────────────────────────────────────────────────────────────
# TAB 5 — DEMAND FORECAST
# ─────────────────────────────────────────────────────────────
with t5:
    st.markdown('<div class="sh">AI-Powered Demand Forecasting</div>',unsafe_allow_html=True)
    st.caption("**RandomForest + GradientBoosting ensemble** · Forecast → Warehouse → Plant propagation")
    fc=st.session_state.forecaster
    demand_nodes=[n for n in sc.nodes.values() if n.node_type=="demand"]
    if not demand_nodes: st.info("Add demand nodes to enable forecasting.")
    else:
        c1,c2,c3=st.columns([2,1,1])
        sel_dem=c1.selectbox("**Demand node**",[n.name for n in demand_nodes],key="fc_node")
        horizon=c2.selectbox("**Horizon (days)**",[14,30,60,90],index=1,key="fc_h")
        use_syn=c3.checkbox("**Use generated history**",value=True,key="fc_syn")

        sel_node=next(n for n in demand_nodes if n.name==sel_dem)
        if not use_syn:
            st.caption("Upload CSV with columns: **date** (YYYY-MM-DD), **demand** (number)")
            hf=st.file_uploader("Upload historical CSV",type=["csv"],key="fc_up")
            if hf:
                try:
                    hdf=pd.read_csv(hf); hdf["date"]=pd.to_datetime(hdf["date"])
                    hdf["node_id"]=sel_node.id; hdf["node_name"]=sel_node.name
                    fc.history[sel_node.id]=hdf.sort_values("date").reset_index(drop=True)
                    st.success(f"OK Loaded {len(hdf)} days of data")
                except Exception as ex: st.error(f"Upload failed: {ex}")

        if st.button("Train Model & Forecast",use_container_width=True,key="train_btn"):
            nid=sel_node.id
            if nid not in fc.history or use_syn: fc.generate_synthetic_history(nid,sel_node.name,sel_node.capacity)
            with st.spinner("Training RF + GBM ensemble..."):
                fdf=fc.train(nid,horizon=int(horizon))
            if fdf is not None:
                st.session_state.forecast_trained.add(nid)
                m=fc.metrics[nid]
                st.success(f"OK Model trained! **RMSE:{m['rmse']:.1f}** | **MAE:{m['mae']:.1f}** | **MAPE:{m['mape']:.1f}%** | **R²:{m['r2']:.3f}**")
            st.rerun()

        nid=sel_node.id
        if nid in st.session_state.forecast_trained and nid in fc.forecasts:
            m=fc.metrics.get(nid,{})
            mc1,mc2,mc3,mc4=st.columns(4)
            mc1.metric("**RMSE**",f"{m.get('rmse','—')}")
            mc2.metric("**MAE**",f"{m.get('mae','—')}")
            mc3.metric("**MAPE**",f"{m.get('mape','—')}%")
            mc4.metric("**R²**",f"{m.get('r2','—')}")
            st.plotly_chart(draw_forecast_chart(fc,nid,sc.nodes),use_container_width=True)
            with st.expander(" **Forecast Values Table**"):
                fdf=fc.forecasts[nid]
                df_show=fdf[["date","forecast","lower","upper"]].copy()
                df_show.columns=["Date","Forecast","Lower CI","Upper CI"]
                df_show["Date"]=df_show["Date"].dt.strftime("%Y-%m-%d")
                for col in ["Forecast","Lower CI","Upper CI"]: df_show[col]=df_show[col].round(1)
                st.dataframe(df_show,use_container_width=True,hide_index=True)

        st.markdown("<br>",unsafe_allow_html=True)
        if st.session_state.forecast_trained:
            if st.button(" Aggregate to Warehouses & Plants",use_container_width=True,key="agg_btn"):
                with st.spinner("Aggregating forecasts..."):
                    for dn in demand_nodes:
                        if dn.id not in fc.history: fc.generate_synthetic_history(dn.id,dn.name,dn.capacity)
                        if dn.id not in fc.forecasts: fc.train(dn.id,horizon=int(horizon)); st.session_state.forecast_trained.add(dn.id)
                    wh_fc=fc.aggregate_to_warehouses(sc,int(horizon))
                    pr=fc.get_plant_requirements(sc,wh_fc,int(horizon))
                    st.session_state.wh_forecasts=wh_fc; st.session_state.plant_req=pr
                st.rerun()

            if st.session_state.wh_forecasts:
                st.markdown('<div class="sh">Warehouse Aggregated Demand</div>',unsafe_allow_html=True)
                wh_fc=st.session_state.wh_forecasts
                fig_wh=go.Figure()
                for i,(wid,wi) in enumerate(wh_fc.items()):
                    fig_wh.add_trace(go.Scatter(x=list(range(1,len(wi["forecast"])+1)),y=wi["forecast"],
                        mode="lines",name=wi["name"],line=dict(width=2)))
                fig_wh.update_layout(title=dict(text="<b>Aggregated Warehouse Demand Forecast</b>",font=dict(size=13,color="#2C3E50"),x=0.5),
                    xaxis=dict(title="Days",gridcolor="#F1F5F9"),yaxis=dict(title="Units",gridcolor="#F1F5F9"),
                    paper_bgcolor="#FFFFFF",plot_bgcolor="#FFFFFF",height=320,margin=dict(l=10,r=10,t=40,b=10))
                st.plotly_chart(fig_wh,use_container_width=True)

            if st.session_state.plant_req:
                st.markdown('<div class="sh">Plant Manufacturing Requirements vs Capacity</div>',unsafe_allow_html=True)
                pr=st.session_state.plant_req
                fig_pr=go.Figure()
                for i,(pid,pi) in enumerate(pr.items()):
                    x=list(range(1,len(pi["required"])+1))
                    fig_pr.add_trace(go.Scatter(x=x,y=pi["required"],mode="lines",name=pi["name"],line=dict(width=2)))
                    fig_pr.add_trace(go.Scatter(x=x,y=[pi["capacity"]]*len(x),mode="lines",name=f"{pi['name']} Capacity",line=dict(dash="dot",width=1.5),showlegend=False))
                fig_pr.update_layout(title=dict(text="<b>Plant Requirements vs Capacity</b>",font=dict(size=13,color="#2C3E50"),x=0.5),
                    xaxis=dict(title="Days",gridcolor="#F1F5F9"),yaxis=dict(title="Units",gridcolor="#F1F5F9"),
                    paper_bgcolor="#FFFFFF",plot_bgcolor="#FFFFFF",height=320,margin=dict(l=10,r=10,t=40,b=10))
                st.plotly_chart(fig_pr,use_container_width=True)

        if st.button("Train All Demand Nodes",use_container_width=True,key="train_all"):
            prog=st.progress(0)
            for i,dn in enumerate(demand_nodes):
                if dn.id not in fc.history: fc.generate_synthetic_history(dn.id,dn.name,dn.capacity)
                fc.train(dn.id,horizon=int(horizon)); st.session_state.forecast_trained.add(dn.id)
                prog.progress((i+1)/len(demand_nodes))
            st.success(f"OK Trained {len(demand_nodes)} demand nodes"); st.rerun()


# ─────────────────────────────────────────────────────────────
# TAB 6 — DISPATCH LOG
# ─────────────────────────────────────────────────────────────
with t6:
    st.markdown('<div class="sh">Dispatch Log — Live Goods Movement</div>',unsafe_allow_html=True)
    with st.form("dis_form"):
        st.markdown("**Log New Dispatch**")
        c1,c2,c3=st.columns(3)
        all_names=[n.name for n in sc.nodes.values()]
        df_from=c1.selectbox("**From Node**",all_names,key="df"); df_to=c2.selectbox("**To Node**",all_names,key="dt",index=min(1,len(all_names)-1))
        ai2={inv.items[iid]["name"]:iid for iid in inv.items} if inv.items else {}
        df_item=c3.selectbox("**Item**",list(ai2.keys()) or ["—"],key="di")
        c1b,c2b,c3b=st.columns(3)
        df_qty=c1b.number_input("**Quantity**",min_value=0.1,value=50.0,key="dq")
        df_st=c2b.selectbox("**Status**",["In Transit","Delivered","Delayed"],key="ds")
        df_no=c3b.text_input("**Notes**",placeholder="Optional",key="dno")
        if st.form_submit_button("Log Dispatch",use_container_width=True):
            fid=[n.id for n in sc.nodes.values() if n.name==df_from]
            tid=[n.id for n in sc.nodes.values() if n.name==df_to]
            iid=ai2.get(df_item,"")
            if fid and tid:
                if df_st in ("In Transit","Delivered") and iid: inv.update_stock(fid[0],iid,-df_qty)
                if df_st=="Delivered" and iid: inv.update_stock(tid[0],iid,df_qty)
                st.session_state.dispatch_log.append({"timestamp":datetime.now().strftime("%Y-%m-%d %H:%M"),
                    "from_node":df_from,"to_node":df_to,"from_id":fid[0],"to_id":tid[0],
                    "item":df_item,"item_id":iid,"quantity":df_qty,"status":df_st,"notes":df_no})
                st.success(f"OK Logged: **{df_qty:.0f} units** of **{df_item}**"); st.rerun()

    if st.session_state.dispatch_log:
        st.markdown("<br>",unsafe_allow_html=True)
        ldf=pd.DataFrame(st.session_state.dispatch_log[::-1])
        cs=[c for c in ["timestamp","from_node","to_node","item","quantity","status","notes"] if c in ldf.columns]
        ldf_s=ldf[cs].copy(); ldf_s.columns=["Time","From","To","Item","Qty","Status","Notes"][:len(cs)]
        def ss(v):
            c={"In Transit":"#D6EAF8","Delivered":"#D5F5E3","Delayed":"#FADBD8"}
            return f"background-color:{c.get(v,'#fff')};font-weight:600"
        st.dataframe(ldf_s.style.map(ss,subset=["Status"]),use_container_width=True,hide_index=True)
        it=len([d for d in st.session_state.dispatch_log if d["status"]=="In Transit"])
        dv=len([d for d in st.session_state.dispatch_log if d["status"]=="Delivered"])
        dl=len([d for d in st.session_state.dispatch_log if d["status"]=="Delayed"])
        c1,c2,c3=st.columns(3)
        c1.markdown(f'<div class="kpi" style="border-top-color:#0369A1"><div class="kpi-lbl">In Transit</div><div class="kpi-val">{it}</div></div>',unsafe_allow_html=True)
        c2.markdown(f'<div class="kpi" style="border-top-color:#15803D"><div class="kpi-lbl">Delivered</div><div class="kpi-val">{dv}</div></div>',unsafe_allow_html=True)
        c3.markdown(f'<div class="kpi" style="border-top-color:#DC2626"><div class="kpi-lbl">Delayed</div><div class="kpi-val">{dl}</div></div>',unsafe_allow_html=True)
        if st.button(" Clear Log",key="cl_log"): st.session_state.dispatch_log=[]; st.rerun()


# ─────────────────────────────────────────────────────────────
# TAB 7 — ATP & SCORECARD
# ─────────────────────────────────────────────────────────────
with t7:
    atp_tab,sc_tab=st.tabs([" Available-to-Promise"," Supplier Scorecard"])
    with atp_tab:
        st.markdown('<div class="sh">Available-to-Promise Analysis</div>',unsafe_allow_html=True)
        ai3={inv.items[iid]["name"]:iid for iid in inv.items} if inv.items else {}
        c1,c2,c3=st.columns([2,1,1])
        atp_i=c1.selectbox("**Item**",list(ai3.keys()) or ["No items"],key="atp_i")
        atp_q=c2.number_input("**Required Qty**",min_value=1.0,value=100.0,key="atp_q")
        atp_d=c3.selectbox("**Destination**",[n.name for n in sc.nodes.values() if n.node_type=="demand"],key="atp_d")
        if st.button(" Check Availability",use_container_width=True,key="atp_chk"):
            iid=ai3.get(atp_i,""); did=[n.id for n in sc.nodes.values() if n.name==atp_d]
            if iid and did:
                alts=inv.find_alternatives(did[0],iid,atp_q,sc)
                if alts:
                    for i,a in enumerate(alts,1):
                        cls="al-g" if a["can_cover"] else "al-a"; bdg="b-g" if a["can_cover"] else "b-a"
                        can_txt = "Can Fulfill" if a["can_cover"] else f"Partial {a['coverage_pct']}%"
                        route_txt = " → ".join(a["path"])
                        st.markdown(f'<div class="{cls}"><b>Option {i}: {a["node_name"]}</b> <span class="{bdg}">{can_txt}</span><br>Available: <b>{a["available"]:.0f}</b> | Coverage: {a["coverage_days"]} days | Route: {route_txt} | Cost: {a["route_cost"]}</div>',unsafe_allow_html=True)
                else:
                    st.markdown('<div class="al-r"><b>No nodes can fulfill this request</b> with current stock levels.</div>',unsafe_allow_html=True)

    with sc_tab:
        st.markdown('<div class="sh">Supplier & Node Scorecard</div>',unsafe_allow_html=True)
        scores=st.session_state.scores
        node_opts=[n for n in sc.nodes.values() if n.node_type in ("plant","warehouse")]
        if node_opts:
            snm=st.selectbox("**Select Node**",[n.name for n in node_opts],key="sc_sel")
            snid=[n.id for n in node_opts if n.name==snm]
            if snid:
                nid=snid[0]
                if nid not in scores: scores[nid]={"reliability":80,"lead_time":80,"quality":80,"cost_efficiency":80}
                s=scores[nid]
                c1,c2=st.columns([1,1])
                with c1:
                    st.markdown("**Edit Performance Scores (0–100)**")
                    s["reliability"]    =st.slider("**Reliability**",   0,100,int(s["reliability"]),   key=f"sr_{nid}")
                    s["lead_time"]      =st.slider("**Lead Time**",     0,100,int(s["lead_time"]),     key=f"sl_{nid}")
                    s["quality"]        =st.slider("**Quality**",       0,100,int(s["quality"]),       key=f"sq_{nid}")
                    s["cost_efficiency"]=st.slider("**Cost Efficiency**",0,100,int(s["cost_efficiency"]),key=f"sc_{nid}")
                    ov=round((s["reliability"]+s["lead_time"]+s["quality"]+s["cost_efficiency"])/4,1)
                    grade="A" if ov>=90 else "B" if ov>=75 else "C" if ov>=60 else "D"
                    gc_="b-g" if grade=="A" else "b-b" if grade=="B" else "b-a" if grade=="C" else "b-r"
                    st.markdown(f"<br><b>Overall: {ov}</b> <span class='{gc_}'>Grade {grade}</span>",unsafe_allow_html=True)
                with c2: st.plotly_chart(draw_scorecard_radar(snm,s),use_container_width=True)

            rows_sc=[]
            for n in node_opts:
                if n.id in scores:
                    s=scores[n.id]; ov=round((s["reliability"]+s["lead_time"]+s["quality"]+s["cost_efficiency"])/4,1)
                    g="A" if ov>=90 else "B" if ov>=75 else "C" if ov>=60 else "D"
                    rows_sc.append({"Node":n.name,"Type":n.node_type.capitalize(),"Reliability":s["reliability"],
                        "Lead Time":s["lead_time"],"Quality":s["quality"],"Cost Eff":s["cost_efficiency"],"Overall":ov,"Grade":g})
            if rows_sc:
                df_sc=pd.DataFrame(rows_sc).sort_values("Overall",ascending=False)
                def gc2(v):
                    c={"A":"#D5F5E3","B":"#D6EAF8","C":"#FDEBD0","D":"#FADBD8"}
                    return f"background-color:{c.get(v,'#fff')};font-weight:700"
                st.dataframe(df_sc.style.map(gc2,subset=["Grade"]),use_container_width=True,hide_index=True)
                # Scorecard bar chart
                if not df_sc.empty and len(df_sc)>0:
                    fig_sc_bar = go.Figure()
                    for metric, color in zip(
                        ["Reliability","Lead Time","Quality","Cost Eff"],
                        ["#1E40AF","#0369A1","#15803D","#7C3AED"]):
                        if metric in df_sc.columns:
                            fig_sc_bar.add_trace(go.Bar(
                                name=metric, x=df_sc["Node"].tolist(), y=df_sc[metric].tolist(),
                                marker_color=color))
                    fig_sc_bar.update_layout(
                        barmode="group",
                        title=dict(text="Performance Comparison by Node",font=dict(size=12,color="#0F172A"),x=0.5),
                        paper_bgcolor="#FFFFFF",plot_bgcolor="#FFFFFF",height=320,
                        yaxis=dict(range=[0,105],gridcolor="#F1F5F9",title="Score"),
                        xaxis=dict(tickfont=dict(size=11)),
                        legend=dict(font=dict(size=11),orientation="h",y=1.08),
                        margin=dict(l=10,r=10,t=50,b=10))
                    st.plotly_chart(fig_sc_bar, use_container_width=True)


# ─────────────────────────────────────────────────────────────
# TAB 8 — GEOGRAPHIC VIEW
# ─────────────────────────────────────────────────────────────
with t8:
    st.markdown('<div class="sh">Geographic Network View</div>',unsafe_allow_html=True)
    _,cr=st.columns([3,1]); mscope=cr.radio("**Scope**",["India","World"],key="geo_s")
    has_c=[n for n in sc.nodes.values() if n.x!=0 or n.y!=0]
    if not has_c: st.info("No coordinates found. Load the demo supply chain to see the India map.")
    else:
        st.plotly_chart(draw_geo_map(sc,mscope.lower()),use_container_width=True)
        with st.expander(" **Node Coordinates**"):
            st.dataframe(pd.DataFrame([{"Node":n.name,"Type":n.node_type,"Lat":n.y,"Lon":n.x,"Location":n.location} for n in sc.nodes.values()]),use_container_width=True,hide_index=True)


# ─────────────────────────────────────────────────────────────
# TAB 9 — REPORTS
# ─────────────────────────────────────────────────────────────
with t9:
    st.markdown('<div class="sh">Supply Chain Report Generator</div>',unsafe_allow_html=True)
    st.caption("Generate a **comprehensive Excel report** with charts, data, color coding, and recommendations.")

    with st.container():
        st.markdown('<div class="section-card">', unsafe_allow_html=True)
        st.markdown("** Report Period**")
        c1,c2=st.columns(2)
        start_dt=c1.date_input("**From Date**",value=date.today()-timedelta(days=30),key="rpt_start")
        end_dt  =c2.date_input("**To Date**",  value=date.today(),                   key="rpt_end")

        st.markdown("** Report Sections**")
        c1,c2,c3=st.columns(3)
        inc_network =c1.checkbox("**Network Summary**",value=True,key="r1")
        inc_demand  =c1.checkbox("**Demand Fulfillment**",value=True,key="r2")
        inc_inv     =c2.checkbox("**Inventory Status**",value=True,key="r3")
        inc_risk    =c2.checkbox("**Risk Analysis**",value=True,key="r4")
        inc_forecast=c3.checkbox("**Demand Forecast**",value=True,key="r5")
        inc_score   =c3.checkbox("**Supplier Scorecard**",value=True,key="r6")
        st.markdown('</div>', unsafe_allow_html=True)

    st.markdown("<br>",unsafe_allow_html=True)
    if st.button("Generate & Download Excel Report",use_container_width=True,type="primary",key="gen_report"):
        with st.spinner("Generating comprehensive report..."):
            try:
                ranking_for_report=st.session_state.ranking or []
                report_bytes=generate_excel_report(sc,inv,st.session_state.forecaster,
                                                   ranking_for_report,start_dt,end_dt)
                fname=f"SC_Report_{start_dt}_{end_dt}.xlsx"
                st.download_button(
                    label=f"Download: {fname}",
                    data=report_bytes,file_name=fname,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True)
                st.success(f"OK Report generated! Click above to download.")
            except Exception as ex:
                st.error(f"Report generation failed: {ex}")

    st.markdown("<br>",unsafe_allow_html=True)
    st.markdown('<div class="sh">Report Preview</div>',unsafe_allow_html=True)
    with st.expander("**Network Summary Preview**"):
        plants=[n for n in sc.nodes.values() if n.node_type=="plant"]
        whs_=[n for n in sc.nodes.values() if n.node_type=="warehouse"]
        dems_=[n for n in sc.nodes.values() if n.node_type=="demand"]
        st.markdown(f"""
        | Metric | Value |
        |--------|-------|
        | **Manufacturing Plants** | {len(plants)} |
        | **Warehouses** | {len(whs_)} |
        | **Demand Points** | {len(dems_)} |
        | **Network Connections** | {len(sc.edges)} |
        | **Total Plant Capacity** | {sum(n.capacity for n in plants):,.0f} |
        | **Total Demand** | {sum(n.capacity for n in dems_):,.0f} |
        | **Supply Coverage** | {round(sum(n.capacity for n in plants)/max(sum(n.capacity for n in dems_),1)*100,1)}% |
        """)
    with st.expander("**Inventory Status Preview**"):
        inv_df=inv.to_df(sc.nodes)
        if not inv_df.empty:
            def cs3(v): return f"color:{'#E74C3C' if v=='Critical' else '#E67E22' if v=='Low' else '#27AE60'};font-weight:700"
            st.dataframe(inv_df[["Node","Item","Current Stock","Safety Stock","Coverage Days","Status"]].style.map(cs3,subset=["Status"]),use_container_width=True,hide_index=True)
    with st.expander("**Risk Analysis Preview**"):
        if st.session_state.ranking:
            for r in st.session_state.ranking[:5]:
                badge={"critical":"b-r","high":"b-a","medium":"b-b","low":"b-g"}[r["severity"]]
                st.markdown(f'<span class="{badge}">{r["severity"].upper()}</span> **{r["label"]}** — Drop: **{r["avg_fulfillment_drop"]}%** | Resilience: **{r["resilience_score"]}%**',unsafe_allow_html=True)
        else:
            st.info("Run **Stress Test** in Risk Heatmap tab first.")


# ─────────────────────────────────────────────────────────────
# TAB 10 — AI ASSISTANT (Free via Groq)
# ─────────────────────────────────────────────────────────────
with t10:
    st.markdown('<div class="sh">AI Supply Chain Assistant</div>',unsafe_allow_html=True)

    if not st.session_state.ai_key:
        st.markdown(f"""
        <div class="al-a">
        <b> Free AI Setup Required</b><br><br>
        This assistant uses <b>Groq</b> — completely free, no credit card needed!<br><br>
        <b>Steps:</b><br>
        1. Go to <a href="{GROQ_SIGNUP}" target="_blank"><b>console.groq.com</b></a><br>
        2. Sign up free → API Keys → Create Key<br>
        3. Copy key (starts with <code>gsk_</code>)<br>
        4. Paste in sidebar → <b>AI Assistant Setup</b><br><br>
        <i>Models available: Llama 3.1 8B (fast) · Llama 3.3 70B (smart) · Gemma 2 9B</i>
        </div>""", unsafe_allow_html=True)

    # Voice
    st.markdown('<div class="sh">Voice Interface</div>', unsafe_allow_html=True)
    voice_component()
    st.markdown("<br>",unsafe_allow_html=True)

    # Quick actions
    st.markdown('<div class="sh">Quick Actions</div>', unsafe_allow_html=True)
    qa_cols=st.columns(4)
    for i,(label,query) in enumerate([
        ("Platform Tour","Give me a complete tour of this supply chain platform."),
        ("Network Status","What is my current supply chain network status?"),
        ("Stock Alerts","Check my inventory and tell me what needs urgent attention."),
        ("Best Practices","What supply chain resilience best practices should I implement?"),
    ]):
        if qa_cols[i].button(label,use_container_width=True,key=f"qa_{i}"):
            st.session_state.chat_history.append({"role":"user","content":query}); st.rerun()

    st.markdown("<br>",unsafe_allow_html=True)
    st.markdown('<div class="sh">Conversation</div>', unsafe_allow_html=True)

    if not st.session_state.chat_history:
        st.markdown("""
        <div style="background:#F8F9FA;border:1px solid #E5E8E8;border-radius:12px;padding:28px;
          text-align:center;color:#5D6D7E;margin:10px 0">
          <div style="font-size:32px;margin-bottom:8px"></div>
          <div style="font-size:15px;font-weight:700;color:#2C3E50;margin-bottom:6px">SR AI Assistant</div>
          <div style="font-size:13px">Ask anything, execute actions, or use quick buttons above.</div>
        </div>""", unsafe_allow_html=True)
    else:
        for msg in st.session_state.chat_history:
            role=msg["role"]; content=msg.get("display",msg["content"])
            if role=="user":
                st.markdown(f'<div class="chat-wrap chat-right"><div class="user-bubble">{content}</div></div>',unsafe_allow_html=True)
            else:
                ar=msg.get("action_result","")
                ar_html=f'<div class="al-g" style="margin-top:6px;font-size:12px">{ar}</div>' if ar else ""
                st.markdown(f'<div class="chat-wrap chat-left"><div class="ai-bubble">{content}{ar_html}</div></div>',unsafe_allow_html=True)

    # Auto-respond
    if st.session_state.chat_history and st.session_state.chat_history[-1]["role"]=="user":
        if not st.session_state.ai_key:
            um=st.session_state.chat_history[-1]["content"].lower()
            if any(w in um for w in ["status","network","supply"]):
                plants=[n for n in sc.nodes.values() if n.node_type=="plant"]
                al=inv.get_alerts(sc.nodes)
                reply=f"Network: **{len(plants)} plants**, **{len(sc.edges)} connections**. Stock alerts: **{len([a for a in al if a['level']=='critical'])} critical**. Add your free Groq API key in the sidebar for full AI responses!"
            else:
                reply="Add your **free Groq API key** in the sidebar (AI Assistant Setup) to enable full AI responses. Get it free at [console.groq.com](https://console.groq.com)!"
            disp=re.sub(r'\*\*(.*?)\*\*',r'<b>\1</b>',reply)
            st.session_state.chat_history.append({"role":"assistant","content":reply,"display":disp}); st.rerun()
        else:
            mc=FREE_AI_MODELS[st.session_state.ai_model]
            # Build context
            plants=[n for n in sc.nodes.values() if n.node_type=="plant"]
            al=inv.get_alerts(sc.nodes)
            ctx=(f"[NETWORK: {len(plants)} plants, {len(sc.nodes)-len(plants)} other nodes, "
                 f"{len(sc.edges)} connections, "
                 f"{len([a for a in al if a['level']=='critical'])} critical stock alerts]\n")
            messages=[]
            for m in st.session_state.chat_history:
                content=m["content"]
                if m["role"]=="user" and len(messages)==0: content=ctx+"User: "+content
                messages.append({"role":m["role"],"content":content})
            with st.spinner("Thinking..."):
                reply_text,status=call_free_ai(messages,st.session_state.ai_key,mc)
            if status=="ok" and reply_text:
                action=parse_action(reply_text); display=clean_response(reply_text)
                dhtml=re.sub(r'\*\*(.*?)\*\*',r'<b>\1</b>',display).replace('\n','<br>')
                ar=""
                if action:
                    modifying=action.get("action") in ("update_stock",)
                    if modifying:
                        st.session_state["pending_action"]=action
                        dhtml+="<br><br><i> I can execute this. Type <b>confirm</b> to proceed or <b>cancel</b> to abort.</i>"
                    else:
                        ok,msg=execute_action(action,sc,inv)
                        ar=("OK "+msg) if ok else ("X "+msg)
                st.session_state.chat_history.append({"role":"assistant","content":reply_text,"display":dhtml,"action_result":ar})
                st.session_state.last_ai_text=display; st.rerun()
            elif status=="rate_limit":
                st.session_state.chat_history.append({"role":"assistant","content":"Rate limit reached. Please wait a moment and try again.","display":"Rate limit reached. Please wait a moment and try again."}); st.rerun()
            elif status=="invalid_key":
                st.session_state.chat_history.append({"role":"assistant","content":"Invalid API key. Check your Groq key in the sidebar.","display":"Invalid API key. Check your Groq key in the sidebar."}); st.rerun()
            elif status=="timeout":
                st.session_state.chat_history.append({"role":"assistant","content":"Request timed out. Please try again.","display":"Request timed out. Please try again."}); st.rerun()
            else:
                st.session_state.chat_history.append({"role":"assistant","content":f"Error: {status}","display":f"Error: {status}"}); st.rerun()

    # Pending action confirmation
    if "pending_action" in st.session_state and st.session_state.chat_history and st.session_state.chat_history[-1]["role"]=="user":
        ut=st.session_state.chat_history[-1]["content"].strip().lower()
        if ut in ("confirm","yes","proceed","ok"):
            ok,msg=execute_action(st.session_state["pending_action"],sc,inv)
            st.session_state.chat_history.append({"role":"assistant","content":f"OK {msg}","display":f"OK {msg}","action_result":("OK" if ok else "X")+" "+msg})
            del st.session_state["pending_action"]; st.rerun()
        elif ut in ("cancel","no","abort"):
            st.session_state.chat_history.append({"role":"assistant","content":"Action cancelled.","display":"Action cancelled."})
            del st.session_state["pending_action"]; st.rerun()

    st.markdown("<br>",unsafe_allow_html=True)
    with st.form("chat_form",clear_on_submit=True):
        c1,c2=st.columns([5,1])
        ui=c1.text_input("**Message...**",placeholder="Ask anything or paste voice text...",label_visibility="collapsed",key="chat_in")
        send=c2.form_submit_button("Send",use_container_width=True)
        if send and ui.strip():
            st.session_state.chat_history.append({"role":"user","content":ui.strip()}); st.rerun()

    if st.session_state.last_ai_text:
        st.markdown(f"""<script>setTimeout(function(){{if(window.speakText)window.speakText({json.dumps(st.session_state.last_ai_text[:400])});}},800);</script>""",unsafe_allow_html=True)

    col_clr,_=st.columns([1,4])
    if col_clr.button(" Clear Chat",key="clr_chat"):
        st.session_state.chat_history=[]; st.session_state.last_ai_text=""; st.rerun()
